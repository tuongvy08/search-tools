from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any
from unicodedata import category, normalize

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from quote_workbook_export import WorkbookExportError, _read_valid_xlsx_entries


MAX_REQUEST_FILE_BYTES = 10 * 1024 * 1024
MAX_DATA_ROWS = 2000
MAX_ANALYZE_SCAN_ROWS = 50
MAX_PREVIEW_ROWS = 20
MAX_PREVIEW_COLS = 30
MAX_PREVIEW_CELL_CHARS = 120
CSV_SHEET_NAME = "CSV"


@dataclass
class RequestFile:
    filename: str
    size: int
    file_type: str
    raw: bytes


def _normalize_alias(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    text = "".join(
        ch for ch in normalize("NFKD", text.casefold()) if category(ch) != "Mn"
    )
    parts = re.findall(r"[0-9a-z]+", text)
    return " ".join(parts)


FIELD_ALIASES = {
    "requested_name": {
        "Name",
        "Product Name",
        "Tên",
        "Tên hàng",
        "Tên sản phẩm",
    },
    "code": {
        "Code",
        "Product Code",
        "Catalog No",
        "Catalog Number",
        "Cat No",
        "Mã",
        "Mã hàng",
    },
    "cas": {
        "CAS",
        "CAS No",
        "CAS Number",
        "CAS RN",
        "Số CAS",
    },
}

NORMALIZED_ALIASES = {
    field: {_normalize_alias(alias) for alias in aliases}
    for field, aliases in FIELD_ALIASES.items()
}


def excel_cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def read_request_upload(file_storage) -> RequestFile:
    if file_storage is None:
        raise ValueError("Thiếu file upload.")
    filename = _safe_filename(getattr(file_storage, "filename", "") or "")
    try:
        file_storage.seek(0)
    except Exception:
        pass
    raw = file_storage.read(MAX_REQUEST_FILE_BYTES + 1)
    if len(raw) > MAX_REQUEST_FILE_BYTES:
        raise OverflowError("File quá lớn, tối đa 10MB.")
    if not raw:
        raise ValueError("File upload rỗng.")
    file_type = _detect_file_type(filename, raw)
    return RequestFile(filename=filename, size=len(raw), file_type=file_type, raw=raw)


def analyze_request_file(file_storage, sheet: str | None = None, header_row: Any = None) -> dict:
    upload = read_request_upload(file_storage)
    if upload.file_type == "csv":
        return _analyze_csv(upload, sheet=sheet, header_row=header_row)
    return _analyze_xlsx(upload, sheet=sheet, header_row=header_row)


def parse_request_file(file_storage, mapping_text: str) -> dict:
    upload = read_request_upload(file_storage)
    mapping = _parse_mapping_json(mapping_text)
    if upload.file_type == "csv":
        return _parse_csv(upload, mapping)
    return _parse_xlsx(upload, mapping)


def _safe_filename(filename: str) -> str:
    base = filename.replace("\\", "/").rsplit("/", 1)[-1].strip()
    if not base:
        raise ValueError("Thiếu tên file.")
    lower = base.lower()
    if lower.endswith((".xls", ".xlsm")):
        raise ValueError("Chỉ hỗ trợ .xlsx hoặc .csv; không hỗ trợ .xls/.xlsm.")
    if not lower.endswith((".xlsx", ".csv")):
        raise ValueError("Chỉ hỗ trợ file .xlsx hoặc .csv.")
    return base


def _detect_file_type(filename: str, raw: bytes) -> str:
    lower = filename.lower()
    if lower.endswith(".csv"):
        if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            raise ValueError("File là .xls nhị phân, không phải CSV.")
        return "csv"
    try:
        _read_valid_xlsx_entries(raw)
    except WorkbookExportError as exc:
        raise ValueError(str(exc)) from exc
    return "xlsx"


def _decode_csv(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1258", "cp1252"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _csv_rows(raw: bytes) -> list[list[str]]:
    text = _decode_csv(raw)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";\t,")
    except csv.Error:
        dialect = csv.excel
    return [[excel_cell_to_str(cell) for cell in row] for row in csv.reader(StringIO(text), dialect)]


def _analyze_csv(upload: RequestFile, *, sheet: str | None = None, header_row: Any = None) -> dict:
    rows = _csv_rows(upload.raw)
    candidates = _header_candidates(rows[:MAX_ANALYZE_SCAN_ROWS])
    selected_sheet = (sheet or CSV_SHEET_NAME).strip()
    if selected_sheet != CSV_SHEET_NAME:
        raise ValueError(f"Sheet '{selected_sheet}' không tồn tại.")
    selected_header_row = _optional_header_row(header_row) or _suggest_header_row(candidates, rows)
    if selected_header_row > len(rows):
        raise ValueError("header_row không tồn tại trong file.")
    header = _row_at(rows, selected_header_row)
    warnings: list[str] = []
    return _analysis_response(
        upload,
        sheets=[CSV_SHEET_NAME],
        suggested_sheet=CSV_SHEET_NAME,
        header_candidates=candidates,
        suggested_header_row=selected_header_row,
        header=header,
        preview_rows=rows[:MAX_PREVIEW_ROWS],
        warnings=warnings,
    )


def _analyze_xlsx(upload: RequestFile, *, sheet: str | None = None, header_row: Any = None) -> dict:
    warnings: list[str] = []
    _validate_xlsx(upload.raw)
    wb_values = load_workbook(BytesIO(upload.raw), read_only=True, data_only=True)
    try:
        sheet_names = list(wb_values.sheetnames)
        best_sheet = sheet_names[0] if sheet_names else ""
        best_candidates: list[dict] = []
        best_score = -1
        best_row = 1
        for sheet_name in sheet_names:
            ws = wb_values[sheet_name]
            scan_rows = _worksheet_rows(ws, limit=MAX_ANALYZE_SCAN_ROWS)
            candidates = _header_candidates(scan_rows)
            candidate_header_row = _suggest_header_row(candidates, scan_rows)
            score = candidates[0]["score"] if candidates else 0
            if score > best_score:
                best_sheet = sheet_name
                best_candidates = candidates
                best_score = score
                best_row = candidate_header_row

        if sheet is not None and str(sheet).strip():
            selected_sheet = str(sheet).strip()
            if selected_sheet not in sheet_names:
                raise ValueError(f"Sheet '{selected_sheet}' không tồn tại.")
            best_sheet = selected_sheet
            scan_rows = _worksheet_rows(wb_values[best_sheet], limit=MAX_ANALYZE_SCAN_ROWS)
            best_candidates = _header_candidates(scan_rows)
            best_row = _suggest_header_row(best_candidates, scan_rows)
        requested_header_row = _optional_header_row(header_row)
        if requested_header_row is not None:
            best_row = requested_header_row

        ws = wb_values[best_sheet]
        if best_row > (ws.max_row or 0):
            raise ValueError("header_row không tồn tại trong file.")
        header_rows = _worksheet_rows(ws, min_rows=best_row, limit=best_row)
        header = _row_at(header_rows, best_row)
        preview_rows = _worksheet_rows(ws, limit=MAX_PREVIEW_ROWS)
        suggested = _suggest_mapping(header)
        _append_formula_warnings(upload.raw, best_sheet, best_row, suggested, warnings)
        return _analysis_response(
            upload,
            sheets=sheet_names,
            suggested_sheet=best_sheet,
            header_candidates=best_candidates,
            suggested_header_row=best_row,
            header=header,
            preview_rows=preview_rows,
            warnings=warnings,
        )
    finally:
        wb_values.close()


def _analysis_response(
    upload: RequestFile,
    *,
    sheets: list[str],
    suggested_sheet: str,
    header_candidates: list[dict],
    suggested_header_row: int,
    header: list[str],
    preview_rows: list[list[str]],
    warnings: list[str],
) -> dict:
    suggested = _suggest_mapping(header)
    return {
        "filename": upload.filename,
        "size": upload.size,
        "type": upload.file_type,
        "sheets": sheets,
        "suggested_sheet": suggested_sheet,
        "header_candidates": header_candidates,
        "suggested_header_row": suggested_header_row,
        "columns": _columns_from_header(header),
        "suggested_mapping": suggested,
        "preview": _preview(preview_rows),
        "warnings": warnings,
    }


def _validate_xlsx(raw: bytes) -> None:
    try:
        _read_valid_xlsx_entries(raw)
    except WorkbookExportError as exc:
        raise ValueError(str(exc)) from exc


def _worksheet_rows(ws, *, limit: int, min_rows: int = 0) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([excel_cell_to_str(cell) for cell in row])
        if len(rows) >= limit:
            break
    while len(rows) < min_rows:
        rows.append([])
    return rows


def _header_candidates(rows: list[list[str]]) -> list[dict]:
    candidates = []
    for idx, row in enumerate(rows, start=1):
        matches = _field_matches(row)
        non_empty = sum(1 for cell in row if str(cell or "").strip())
        score = sum(len(cols) for cols in matches.values())
        if score <= 0:
            continue
        candidates.append(
            {
                "row": idx,
                "score": score,
                "non_empty": non_empty,
                "matches": {field: [_column_info(i, row[i]) for i in cols] for field, cols in matches.items()},
            }
        )
    candidates.sort(key=lambda item: (-item["score"], item["row"]))
    return candidates


def _suggest_header_row(candidates: list[dict], rows: list[list[str]]) -> int:
    if candidates:
        return int(candidates[0]["row"])
    for idx, row in enumerate(rows, start=1):
        if any(str(cell or "").strip() for cell in row):
            return idx
    return 1


def _optional_header_row(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        header_row = int(value)
    except (TypeError, ValueError):
        raise ValueError("header_row không hợp lệ.")
    if header_row < 1:
        raise ValueError("header_row không hợp lệ.")
    return header_row


def _field_matches(header: list[str]) -> dict[str, list[int]]:
    matches = {field: [] for field in FIELD_ALIASES}
    for idx, text in enumerate(header):
        norm = _normalize_alias(text)
        if not norm:
            continue
        for field, aliases in NORMALIZED_ALIASES.items():
            if norm in aliases:
                matches[field].append(idx)
    return matches


def _suggest_mapping(header: list[str]) -> dict:
    matches = _field_matches(header)
    out = {}
    for field, cols in matches.items():
        candidates = [_column_info(idx, header[idx]) for idx in cols]
        out[field] = {
            "column": cols[0] if len(cols) == 1 else None,
            "letter": get_column_letter(cols[0] + 1) if len(cols) == 1 else None,
            "header": header[cols[0]] if len(cols) == 1 else None,
            "ambiguous": len(cols) > 1,
            "candidates": candidates,
        }
    return out


def _columns_from_header(header: list[str]) -> list[dict]:
    return [_column_info(idx, text) for idx, text in enumerate(header)]


def _column_info(index: int, header: Any) -> dict:
    return {
        "index": index,
        "letter": get_column_letter(index + 1),
        "header": excel_cell_to_str(header),
    }


def _preview(rows: list[list[str]]) -> list[list[str]]:
    out = []
    for row in rows[:MAX_PREVIEW_ROWS]:
        out.append([_truncate_cell(cell) for cell in row[:MAX_PREVIEW_COLS]])
    return out


def _truncate_cell(value: Any) -> str:
    text = excel_cell_to_str(value)
    if len(text) <= MAX_PREVIEW_CELL_CHARS:
        return text
    return text[:MAX_PREVIEW_CELL_CHARS] + "..."


def _row_at(rows: list[list[str]], row_number: int) -> list[str]:
    idx = row_number - 1
    if idx < 0 or idx >= len(rows):
        return []
    return rows[idx]


def _parse_mapping_json(mapping_text: str) -> dict:
    try:
        mapping = json.loads(mapping_text or "")
    except json.JSONDecodeError as exc:
        raise ValueError("mapping phải là JSON hợp lệ.") from exc
    if not isinstance(mapping, dict):
        raise ValueError("mapping phải là object JSON.")
    return mapping


def _parse_csv(upload: RequestFile, mapping: dict) -> dict:
    rows = _csv_rows(upload.raw)
    parsed_mapping = _validate_mapping(mapping, [CSV_SHEET_NAME], rows)
    return {
        "filename": upload.filename,
        "size": upload.size,
        "type": upload.file_type,
        "sheet": CSV_SHEET_NAME,
        "header_row": parsed_mapping["header_row"],
        "rows": _parse_rows(rows, parsed_mapping),
        "warnings": [],
    }


def _parse_xlsx(upload: RequestFile, mapping: dict) -> dict:
    _validate_xlsx(upload.raw)
    warnings: list[str] = []
    wb_values = load_workbook(BytesIO(upload.raw), read_only=True, data_only=True)
    try:
        sheet_names = list(wb_values.sheetnames)
        sheet = _mapping_sheet(mapping, sheet_names)
        ws = wb_values[sheet]
        header_row = _mapping_header_row(mapping)
        rows = _worksheet_rows_for_parse(ws, header_row)
        parsed_mapping = _validate_mapping(mapping, sheet_names, rows)
        parsed_rows = _parse_rows(rows, parsed_mapping)
        _append_formula_warnings(upload.raw, sheet, header_row, _mapping_as_suggestion(parsed_mapping, rows), warnings)
        return {
            "filename": upload.filename,
            "size": upload.size,
            "type": upload.file_type,
            "sheet": sheet,
            "header_row": header_row,
            "rows": parsed_rows,
            "warnings": warnings,
        }
    finally:
        wb_values.close()


def _validate_mapping(mapping: dict, sheet_names: list[str], rows: list[list[str]]) -> dict:
    sheet = _mapping_sheet(mapping, sheet_names)
    header_row = _mapping_header_row(mapping)
    if header_row > len(rows):
        raise ValueError("header_row không tồn tại trong file.")
    header = _row_at(rows, header_row)
    max_cols = len(header)
    parsed = {"sheet": sheet, "header_row": header_row}
    for field in ("requested_name", "code", "cas"):
        parsed[field] = _mapping_column(mapping, field, max_cols)
    if parsed["code"] is None and parsed["cas"] is None:
        raise ValueError("Cần map ít nhất Code hoặc CAS.")
    return parsed


def _worksheet_rows_for_parse(ws, header_row: int) -> list[list[str]]:
    rows: list[list[str]] = []
    data_seen = 0
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [excel_cell_to_str(cell) for cell in row]
        rows.append(values)
        if row_number <= header_row:
            continue
        if not any(values):
            continue
        data_seen += 1
        if data_seen > MAX_DATA_ROWS:
            break
    while len(rows) < header_row:
        rows.append([])
    return rows


def _mapping_sheet(mapping: dict, sheet_names: list[str]) -> str:
    sheet = mapping.get("sheet")
    if not isinstance(sheet, str) or not sheet.strip():
        raise ValueError("mapping thiếu sheet.")
    sheet = sheet.strip()
    if sheet not in sheet_names:
        raise ValueError(f"Sheet '{sheet}' không tồn tại.")
    return sheet


def _mapping_header_row(mapping: dict) -> int:
    try:
        header_row = int(mapping.get("header_row"))
    except (TypeError, ValueError):
        raise ValueError("header_row không hợp lệ.")
    if header_row < 1:
        raise ValueError("header_row không hợp lệ.")
    return header_row


def _mapping_column(mapping: dict, field: str, max_cols: int) -> int | None:
    value = mapping.get(field)
    if isinstance(value, dict):
        value = value.get("column")
    if value is None or value == "":
        return None
    if isinstance(value, str) and re.fullmatch(r"[A-Za-z]{1,3}", value.strip()):
        value = _letter_to_index(value.strip())
    else:
        try:
            value = int(value)
        except (TypeError, ValueError):
            raise ValueError(f"Cột {field} không hợp lệ.")
    if value < 0 or value >= max_cols:
        raise ValueError(f"Cột {field} không tồn tại trong header.")
    return value


def _letter_to_index(letter: str) -> int:
    total = 0
    for ch in letter.upper():
        total = total * 26 + (ord(ch) - ord("A") + 1)
    return total - 1


def _parse_rows(rows: list[list[str]], mapping: dict) -> list[dict]:
    output = []
    data_seen = 0
    for source_row, row in enumerate(rows[mapping["header_row"] :], start=mapping["header_row"] + 1):
        if not any(excel_cell_to_str(cell) for cell in row):
            continue
        data_seen += 1
        if data_seen > MAX_DATA_ROWS:
            raise OverflowError(f"Tối đa {MAX_DATA_ROWS} dòng dữ liệu mỗi lần.")
        item = {
            "source_row": source_row,
            "requested_name": _cell_by_mapping(row, mapping.get("requested_name")),
            "code": _cell_by_mapping(row, mapping.get("code")),
            "cas": _cell_by_mapping(row, mapping.get("cas")),
        }
        output.append(item)
    return output


def _cell_by_mapping(row: list[str], col: int | None) -> str:
    if col is None or col >= len(row):
        return ""
    return excel_cell_to_str(row[col])


def _mapping_as_suggestion(mapping: dict, rows: list[list[str]]) -> dict:
    header = _row_at(rows, mapping["header_row"])
    out = {}
    for field in ("requested_name", "code", "cas"):
        col = mapping.get(field)
        out[field] = {
            "column": col,
            "letter": get_column_letter(col + 1) if col is not None else None,
            "header": header[col] if col is not None and col < len(header) else None,
            "ambiguous": False,
            "candidates": [],
        }
    return out


def _append_formula_warnings(raw: bytes, sheet: str, header_row: int, suggested_mapping: dict, warnings: list[str]) -> None:
    formula_cols = []
    for field in ("code", "cas"):
        item = suggested_mapping.get(field) or {}
        col = item.get("column")
        if isinstance(col, int):
            formula_cols.append((field, col))
    if not formula_cols:
        return

    wb_formula = load_workbook(BytesIO(raw), read_only=True, data_only=False)
    wb_values = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    try:
        if sheet not in wb_formula.sheetnames or sheet not in wb_values.sheetnames:
            return
        ws_formula = wb_formula[sheet]
        ws_values = wb_values[sheet]
        for field, col in formula_cols:
            letter = get_column_letter(col + 1)
            for row_num in range(header_row + 1, min(ws_formula.max_row or header_row, header_row + MAX_DATA_ROWS) + 1):
                formula_cell = ws_formula[f"{letter}{row_num}"]
                if not (isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")):
                    continue
                value_cell = ws_values[f"{letter}{row_num}"]
                if value_cell.value is None:
                    warnings.append(
                        f"Ô {letter}{row_num} ({field}) là formula nhưng không có cached value."
                    )
                    break
    finally:
        wb_formula.close()
        wb_values.close()
