"""Đọc file Excel .xlsx → danh sách tuple (name..note). Dùng chung cho import/sync."""

from __future__ import annotations

import os
from typing import Any


def load_product_rows_from_xlsx(xlsx_path: str) -> list[tuple[Any, ...]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("Cần cài: pip install openpyxl") from e

    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(xlsx_path)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("File Excel rỗng.")

    def norm(s):
        if s is None:
            return ""
        return str(s).strip().lower()

    header_map = {
        norm(c): i for i, c in enumerate(header_row) if c is not None and str(c).strip()
    }
    required = ["name", "code", "cas", "brand", "size", "ship", "price", "note"]
    for col in required:
        if col not in header_map:
            wb.close()
            raise ValueError(
                f"Thiếu cột '{col}'. Các cột tìm thấy: {list(header_map.keys())}"
            )

    data_rows = []
    for row in rows_iter:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        def cell(name):
            i = header_map[name]
            v = row[i] if i < len(row) else None
            if v is None:
                return None
            return str(v).strip() if not isinstance(v, (int, float)) else str(v)

        data_rows.append(
            (
                cell("name"),
                cell("code"),
                cell("cas"),
                cell("brand"),
                cell("size"),
                cell("ship"),
                cell("price"),
                cell("note"),
            )
        )

    wb.close()

    if not data_rows:
        raise ValueError("Không có dòng dữ liệu nào sau dòng tiêu đề.")

    return data_rows
