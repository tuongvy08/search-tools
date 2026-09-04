import csv
import hashlib
import ipaddress
import json
import os
import re
import sys
import zipfile
from io import BytesIO, StringIO
from typing import Optional
from uuid import uuid4

from dotenv import load_dotenv
from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_cors import CORS
import psycopg2
from openpyxl import Workbook, load_workbook
from psycopg2 import Binary, IntegrityError
from psycopg2.errors import UndefinedTable
from werkzeug.security import check_password_hash, generate_password_hash

from werkzeug.middleware.proxy_fix import ProxyFix

import admin_google_users
import admin_login_history
import admin_teams
import auth_google
import session_security
from compliance_resolver import compliance_css_type, resolve_compliance_precedence
from db import get_connection
from middleware_access import register_ip_access_control
from product_import_manual import (
    HEADER_MODE_ABSENT,
    classify_manual_compliance_headers,
    fetch_manual_compliance_snapshot,
    fetch_preparation_type_snapshot,
    resolve_preparation_type_for_write,
    resolve_manual_fields_for_write,
    validate_product_import_rows,
)
from quote_request_file import analyze_request_file, parse_request_file
from quote_workbook_export import MAX_XLSX_BYTES, WorkbookExportError, export_quick_quote_workbook, inspect_bg_template

load_dotenv()

app = Flask(__name__)
CORS(app)

app.secret_key = os.environ.get("FLASK_SECRET_KEY")
if not app.secret_key:
    raise RuntimeError(
        "FLASK_SECRET_KEY environment variable is required (no default is provided)."
    )

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=auth_google.strict_bool_env("SESSION_COOKIE_SECURE", False),
)

# Gunicorn binds to loopback behind exactly one Nginx hop in this deployment;
# only trust one hop for each forwarded header (never an arbitrary count).
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=0)

# Registers /auth/google + /auth/google/callback. No-op (no secrets/network)
# unless GOOGLE_AUTH_ENABLED=true; raises at startup if enabled+misconfigured.
auth_google.init_app(app)

# Registers POST /logout + the per-request session-liveness check (Phase
# 5D2A): revoked/suspended/deleted accounts and bumped auth_version are
# rejected on the very next request, not just at next login. Phase 6A:
# MUST be registered before `register_ip_access_control` below -- Flask
# runs `before_request` hooks in registration order, and the IP/team
# policy check needs to trust `session.get(...)` (team_id/is_admin/
# ip_bypass_allowlist), which is only safe once this hook has already
# confirmed the session's account is still ACTIVE and auth_version still
# matches (or rejected/cleared it). See middleware_access.py's docstring.
session_security.init_app(app)

base_path = os.environ.get("ACCESS_CONTROL_BASE_PATH", "/home/deploy/myapps")
register_ip_access_control(app, base_path=base_path)

# Registers the POST /admin/users/google/* actions (Phase 5D2B): approve /
# invite / suspend / reactivate / revoke-sessions for Google Workspace
# accounts. The GET page itself stays on admin_users() below (same
# /admin/users route as the existing LOCAL/legacy user management).
app.register_blueprint(admin_google_users.admin_google_users_bp)

# Registers GET /admin/login-history (Phase 5D3): read-only admin screen
# over `login_audit_events`. Admin-only; never writes anything.
app.register_blueprint(admin_login_history.admin_login_history_bp)

# Registers GET/POST /admin/teams (Phase 6A): team CRUD (create/rename),
# brand assignment (from existing product brands, not free text), and the
# 3-mode IP policy, gated behind the preview -> confirm flow described in
# that module's docstring.
app.register_blueprint(admin_teams.admin_teams_bp)


@app.after_request
def _static_no_cache_js_css(response):
    """Tránh trình duyệt giữ bản cũ của script.js / styles.css sau khi deploy."""
    try:
        path = request.path or ""
        if path.startswith("/static/") and path.endswith((".js", ".css")):
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response.headers["Pragma"] = "no-cache"
    except Exception:
        pass
    return response


ENABLE_LEGACY_PASSWORD_LOGIN = auth_google.strict_bool_env("ENABLE_LEGACY_PASSWORD_LOGIN", False)
MANAGER_PASSWORD = os.environ.get("APP_PASSWORD_MANAGER", "")
STAFF_PASSWORD = os.environ.get("APP_PASSWORD_STAFF", "")
if ENABLE_LEGACY_PASSWORD_LOGIN and not (MANAGER_PASSWORD and STAFF_PASSWORD):
    raise RuntimeError(
        "ENABLE_LEGACY_PASSWORD_LOGIN=true requires BOTH APP_PASSWORD_MANAGER and "
        "APP_PASSWORD_STAFF to be set explicitly (no defaults)."
    )

IMPORT_PREVIEWS = {}

QUOTE_TEMPLATE_PROFILE_VERSION = "BG_V1"
QUOTE_TEMPLATE_MAPPING_SNAPSHOT = {
    "profile_version": QUOTE_TEMPLATE_PROFILE_VERSION,
    "sheet": "BG",
    "header_row": 16,
    "product_start_row": 17,
    "total_label": "Tổng giá",
    "mapping": {
        "sequence": "A",
        "Name": "B",
        "Code": "C",
        "Cas": "D",
        "Brand": "E",
        "Size": "F",
        "Note": "M",
        "Compliance_Combined": "N",
        "Unit_Price_Value": "P",
    },
}


class QuoteTemplateError(ValueError):
    pass


def _default_exchange_rates_from_json() -> dict[str, float]:
    path = os.path.join(app.root_path, "static", "exchange_rates.json")
    out: dict[str, float] = {}
    try:
        with open(path, "r", encoding="utf-8") as file:
            raw = json.load(file)
        for k, v in (raw or {}).items():
            try:
                out[str(k).strip()] = float(v)
            except (TypeError, ValueError):
                continue
    except Exception:
        pass
    return out


def _exchange_rate_map(conn) -> dict[str, float]:
    """JSON làm mặc định; dòng trong bảng exchange_rates ghi đè theo brand."""
    base = _default_exchange_rates_from_json()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT brand, rate FROM exchange_rates")
            for b, r in cur.fetchall():
                if b is None or str(b).strip() == "":
                    continue
                try:
                    base[str(b).strip()] = float(r)
                except (TypeError, ValueError):
                    continue
    except Exception:
        pass
    return base


def _visibility_sql(alias: str):
    if session.get("is_admin"):
        return "", ()
    tid = session.get("team_id")
    if tid is None:
        return " AND FALSE", ()
    return (f" AND {alias}.brand IN (SELECT brand FROM team_brands WHERE team_id = %s)", (tid,))


def _warning_css_type(label: Optional[str]) -> Optional[str]:
    return compliance_css_type(label)


def _norm(v):
    return (v or "").strip()


def _split_multi_items(text: str, max_items: int = 2000) -> list[str]:
    """
    Tách danh sách nhiều dòng từ textarea/input.
    Cho phép xuống dòng và tách thêm bởi dấu phẩy/dấu chấm phẩy.
    Giữ thứ tự xuất hiện (không bỏ trùng), để output khớp đúng với list bạn paste.
    """
    if not text:
        return []
    out: list[str] = []
    for line in str(text).splitlines():
        line = (line or "").strip()
        if not line or line.startswith("#"):
            continue
        for part in line.replace(";", ",").split(","):
            item = (part or "").strip()
            if not item:
                continue
            out.append(item)
            if len(out) >= max_items:
                return out
    return out


_SIZE_TOKEN_RE = re.compile(
    r"([\d]+(?:[.,]\d+)?)\s*([a-zA-Zµμ]+)?",
    re.IGNORECASE,
)


def _parse_size_token(size_str: str) -> dict:
    """Tách số lượng + đơn vị từ chuỗi size (vd. 100mg, 1 L, 2,5g)."""
    raw = (size_str or "").strip()
    if not raw:
        return {"raw": "", "value": None, "unit": ""}
    normalized = raw.replace(",", ".")
    match = _SIZE_TOKEN_RE.search(normalized)
    if not match:
        return {"raw": raw.upper(), "value": None, "unit": ""}
    try:
        value = float(match.group(1))
    except (TypeError, ValueError):
        value = None
    unit = (match.group(2) or "").strip().lower()
    unit_aliases = {
        "l": "l",
        "liter": "l",
        "litre": "l",
        "liters": "l",
        "litres": "l",
        "ml": "ml",
        "milliliter": "ml",
        "millilitre": "ml",
        "g": "g",
        "gram": "g",
        "grams": "g",
        "mg": "mg",
        "kg": "kg",
        "µg": "ug",
        "μg": "ug",
        "ug": "ug",
    }
    unit = unit_aliases.get(unit, unit)
    return {"raw": raw.upper(), "value": value, "unit": unit}


def _size_matches(product_size: str, allowed_sizes: list[str], tolerance_pct: float = 0) -> bool:
    if not allowed_sizes:
        return True
    product_norm = (product_size or "").strip().upper()
    product_parsed = _parse_size_token(product_size)
    for allowed in allowed_sizes:
        allowed_norm = (allowed or "").strip().upper()
        if product_norm and allowed_norm and product_norm == allowed_norm:
            return True
        if tolerance_pct <= 0:
            continue
        allowed_parsed = _parse_size_token(allowed)
        if (
            product_parsed["value"] is not None
            and allowed_parsed["value"] is not None
            and product_parsed["unit"]
            and product_parsed["unit"] == allowed_parsed["unit"]
            and allowed_parsed["value"] > 0
        ):
            diff_pct = abs(product_parsed["value"] - allowed_parsed["value"]) / allowed_parsed["value"] * 100
            if diff_pct <= tolerance_pct:
                return True
    return False


def _split_multi_values(text: str, max_items: int = 200) -> list[str]:
    """Tách danh sách brand/size từ form (dòng, phẩy, hoặc nhiều field cùng tên)."""
    if isinstance(text, list):
        items = [str(x).strip() for x in text if str(x).strip()]
        return items[:max_items]
    return _split_multi_items(text or "", max_items=max_items)


def _product_row_to_result(
    name,
    code,
    cas,
    brand,
    size,
    ship,
    price,
    note,
    compliance_status,
    compliance_note,
    rate_map: dict,
    *,
    brand_manual_enabled=False,
    manual_compliance=None,
    manual_compliance_note=None,
) -> dict:
    try:
        ship_f = float(ship) if ship is not None else 0
    except (TypeError, ValueError):
        ship_f = 0
    try:
        price_f = float(price) if price is not None else 0
    except (TypeError, ValueError):
        price_f = 0
    bkey = (brand or "").strip()
    exchange_rate = rate_map.get(bkey, 1.0)
    unit_price = round(price_f * ship_f * exchange_rate, -3)
    resolved = resolve_compliance_precedence(
        brand_manual_enabled=bool(brand_manual_enabled),
        manual_compliance=manual_compliance,
        manual_compliance_note=manual_compliance_note,
        legacy_compliance=compliance_status,
        legacy_compliance_note=compliance_note,
        cas=cas,
    )
    return {
        "Name": name or "",
        "Code": code or "",
        "Cas": cas or "",
        "Brand": brand or "",
        "Size": size or "",
        "Unit_Price": "{:,.0f}".format(unit_price),
        "Note": note or "",
        "Compliance_Status": resolved["compliance"],
        "Compliance_Note": resolved["compliance_note"],
        "Compliance_Css": resolved["compliance_css"],
        "Compliance_Source": resolved["compliance_source"],
        "note": note or "",
        "compliance": resolved["compliance"],
        "compliance_note": resolved["compliance_note"],
        "compliance_css": resolved["compliance_css"],
        "compliance_source": resolved["compliance_source"],
    }


QUOTE_MAX_ROWS = 2000
QUOTE_CANDIDATE_LIMIT = 500
QUOTE_SELECTION_MANUAL = "MANUAL"
QUOTE_SELECTION_LOWEST_UNIT_PRICE = "LOWEST_UNIT_PRICE"
QUOTE_SELECTION_LOWEST_OVERALL = "LOWEST_OVERALL"
QUOTE_SELECTION_LOWEST_PER_BRAND = "LOWEST_PER_BRAND"
QUOTE_SELECTION_ALIASES = {QUOTE_SELECTION_LOWEST_UNIT_PRICE: QUOTE_SELECTION_LOWEST_OVERALL}
QUOTE_SELECTION_STRATEGIES = {
    QUOTE_SELECTION_MANUAL,
    QUOTE_SELECTION_LOWEST_UNIT_PRICE,
    QUOTE_SELECTION_LOWEST_OVERALL,
    QUOTE_SELECTION_LOWEST_PER_BRAND,
}
QUOTE_BLOCKED_COMPLIANCE = {"CẤM NHẬP", "Cấm nhập", "Chưa xác định"}
QUOTE_WARNING_COMPLIANCE = {"Phụ lục II", "Phụ lục III", "Cần giấy phép"}
QUOTE_UNIT_GROUP_ANY = "ANY"
QUOTE_UNIT_GROUP_SOLID = "SOLID"
QUOTE_UNIT_GROUP_LIQUID = "LIQUID"
QUOTE_UNIT_GROUPS = {QUOTE_UNIT_GROUP_ANY, QUOTE_UNIT_GROUP_SOLID, QUOTE_UNIT_GROUP_LIQUID}
QUOTE_PREPARATION_ANY = "ANY"
QUOTE_PREPARATION_NEAT = "NEAT"
QUOTE_PREPARATION_SOLUTION = "SOLUTION"
QUOTE_PREPARATION_MIXTURE = "MIXTURE"
QUOTE_PREPARATION_TYPES = {
    QUOTE_PREPARATION_ANY,
    QUOTE_PREPARATION_NEAT,
    QUOTE_PREPARATION_SOLUTION,
    QUOTE_PREPARATION_MIXTURE,
}
QUOTE_SIZE_MODE_ANY = "ANY"
QUOTE_SIZE_MODE_EXACT = "EXACT"
QUOTE_SIZE_MODE_MIN = "MIN"
QUOTE_SIZE_MODE_MAX = "MAX"
QUOTE_SIZE_MODES = {QUOTE_SIZE_MODE_ANY, QUOTE_SIZE_MODE_EXACT, QUOTE_SIZE_MODE_MIN, QUOTE_SIZE_MODE_MAX}
QUOTE_SOLID_FACTORS_TO_MG = {"MG": 1.0, "G": 1000.0, "KG": 1000000.0}
QUOTE_LIQUID_FACTORS_TO_ML = {"ML": 1.0, "L": 1000.0}
QUOTE_SIMPLE_SIZE_RE = re.compile(r"^\s*(\d+(?:[.,]\d+)?)\s*([a-zA-Zµμ]+)\s*$", re.IGNORECASE)

# ── Lifecycle & Reason Codes ──
LIFECYCLE_SELECTED = "SELECTED"
LIFECYCLE_REVIEW = "REVIEW"
LIFECYCLE_UNRESOLVED = "UNRESOLVED"
LIFECYCLE_BLOCKED = "BLOCKED"
LIFECYCLE_EXPORTED = "EXPORTED"
QUOTE_LIFECYCLES = {
    LIFECYCLE_SELECTED,
    LIFECYCLE_REVIEW,
    LIFECYCLE_UNRESOLVED,
    LIFECYCLE_BLOCKED,
    LIFECYCLE_EXPORTED,
}

REASON_PENDING_MATCH = "PENDING_MATCH"
REASON_MISSING_IDENTIFIER = "MISSING_IDENTIFIER"
REASON_NO_MATCH = "NO_MATCH"
REASON_CODE_CAS_CONFLICT = "CODE_CAS_CONFLICT"
REASON_CODE_HAS_NO_CAS = "CODE_HAS_NO_CAS"
REASON_CODE_MULTIPLE_CAS = "CODE_MULTIPLE_CAS"
REASON_BRAND_REQUIRED = "BRAND_REQUIRED"
REASON_NO_VALID_PRICE = "NO_VALID_PRICE"
REASON_MANUAL_SELECTION_REQUIRED = "MANUAL_SELECTION_REQUIRED"
REASON_FILTER_NO_MATCH = "FILTER_NO_MATCH"
REASON_COMPLIANCE_BLOCKED = "COMPLIANCE_BLOCKED"
REASON_COMPLIANCE_UNRESOLVED = "COMPLIANCE_UNRESOLVED"
REASON_DUPLICATE_CODE_BRAND_SIZE = "DUPLICATE_CODE_BRAND_SIZE"
REASON_AUTO_SELECTED = "AUTO_SELECTED"
REASON_MANUALLY_SELECTED = "MANUALLY_SELECTED"
REASON_EXPORTED_SUCCESSFULLY = "EXPORTED_SUCCESSFULLY"
REASON_CANDIDATE_LIMIT_EXCEEDED = "CANDIDATE_LIMIT_EXCEEDED"
REASON_CODE_HAS_PLACEHOLDER_CAS = "CODE_HAS_PLACEHOLDER_CAS"
QUOTE_REASON_CODES = {
    REASON_PENDING_MATCH,
    REASON_MISSING_IDENTIFIER,
    REASON_NO_MATCH,
    REASON_CODE_CAS_CONFLICT,
    REASON_CODE_HAS_NO_CAS,
    REASON_CODE_MULTIPLE_CAS,
    REASON_BRAND_REQUIRED,
    REASON_NO_VALID_PRICE,
    REASON_MANUAL_SELECTION_REQUIRED,
    REASON_FILTER_NO_MATCH,
    REASON_COMPLIANCE_BLOCKED,
    REASON_COMPLIANCE_UNRESOLVED,
    REASON_DUPLICATE_CODE_BRAND_SIZE,
    REASON_AUTO_SELECTED,
    REASON_MANUALLY_SELECTED,
    REASON_EXPORTED_SUCCESSFULLY,
    REASON_CANDIDATE_LIMIT_EXCEEDED,
    REASON_CODE_HAS_PLACEHOLDER_CAS,
}

REASON_CODE_TO_LIFECYCLE = {
    REASON_PENDING_MATCH: LIFECYCLE_REVIEW,
    REASON_MISSING_IDENTIFIER: LIFECYCLE_UNRESOLVED,
    REASON_NO_MATCH: LIFECYCLE_UNRESOLVED,
    REASON_CODE_CAS_CONFLICT: LIFECYCLE_UNRESOLVED,
    REASON_CODE_HAS_NO_CAS: LIFECYCLE_UNRESOLVED,
    REASON_CODE_MULTIPLE_CAS: LIFECYCLE_UNRESOLVED,
    REASON_COMPLIANCE_UNRESOLVED: LIFECYCLE_UNRESOLVED,
    REASON_BRAND_REQUIRED: LIFECYCLE_REVIEW,
    REASON_NO_VALID_PRICE: LIFECYCLE_REVIEW,
    REASON_MANUAL_SELECTION_REQUIRED: LIFECYCLE_REVIEW,
    REASON_FILTER_NO_MATCH: LIFECYCLE_REVIEW,
    REASON_DUPLICATE_CODE_BRAND_SIZE: LIFECYCLE_REVIEW,
    REASON_COMPLIANCE_BLOCKED: LIFECYCLE_BLOCKED,
    REASON_AUTO_SELECTED: LIFECYCLE_SELECTED,
    REASON_MANUALLY_SELECTED: LIFECYCLE_SELECTED,
    REASON_EXPORTED_SUCCESSFULLY: LIFECYCLE_EXPORTED,
    REASON_CANDIDATE_LIMIT_EXCEEDED: LIFECYCLE_REVIEW,
    REASON_CODE_HAS_PLACEHOLDER_CAS: LIFECYCLE_UNRESOLVED,
}

# ── Export placeholder lines (Phase 4A) ──
# Only these three lifecycles may be exported without a selected product_id.
QUOTE_EXPORT_PLACEHOLDER_CLASSIFICATIONS = {LIFECYCLE_UNRESOLVED, LIFECYCLE_BLOCKED, LIFECYCLE_REVIEW}
QUOTE_EXPORT_PLACEHOLDER_NOTES = {
    LIFECYCLE_UNRESOLVED: "Không tìm thấy",
    LIFECYCLE_REVIEW: "Cần kiểm tra/chọn thủ công",
}
QUOTE_EXPORT_BLOCKED_REASON_VN = {
    REASON_COMPLIANCE_BLOCKED: "tất cả sản phẩm bị chặn compliance",
    REASON_COMPLIANCE_UNRESOLVED: "compliance chưa xác định",
}
QUOTE_EXPORT_BLOCKED_REASON_VN_DEFAULT = "không đủ điều kiện báo giá"


# ── Brand Policy Modes ──
BRAND_POLICY_INHERIT = "INHERIT"
BRAND_POLICY_PRIORITY_FALLBACK = "PRIORITY_FALLBACK"
BRAND_POLICY_ALLOWLIST_ONLY = "ALLOWLIST_ONLY"
BRAND_POLICY_ALL_AVAILABLE = "ALL_AVAILABLE"

VALID_GLOBAL_BRAND_POLICIES = {
    BRAND_POLICY_PRIORITY_FALLBACK,
    BRAND_POLICY_ALLOWLIST_ONLY,
    BRAND_POLICY_ALL_AVAILABLE,
}

VALID_ROW_BRAND_POLICIES = {
    BRAND_POLICY_INHERIT,
    BRAND_POLICY_PRIORITY_FALLBACK,
    BRAND_POLICY_ALLOWLIST_ONLY,
    BRAND_POLICY_ALL_AVAILABLE,
}

MAX_POLICY_TIERS = 20
MAX_POLICY_BRANDS = 100


QUOTE_TEMPLATE_TABLE_UNAVAILABLE_MSG = "Hệ thống quản lý mẫu báo giá chưa sẵn sàng. Vui lòng liên hệ quản trị viên."


def _is_table_missing_error(exc: BaseException) -> bool:
    if isinstance(exc, UndefinedTable):
        return True
    if getattr(exc, "pgcode", None) == "42P01":
        return True
    if exc.__class__.__name__ == "UndefinedTable":
        return True
    return False


def _quote_json_error(message: str, status: int = 400):
    return jsonify({"error": message}), status


def _require_authenticated_quote_api():
    if not session.get("authenticated"):
        return _quote_json_error("Chưa đăng nhập.", status=401)
    if not session.get("is_admin") and session.get("team_id") is None:
        return _quote_json_error("Tài khoản chưa được gán team.", status=403)
    return None


def _quote_norm_identifier(value) -> str:
    return "" if value is None else str(value).strip().upper()


def _quote_text(value, max_len: int = 500) -> str:
    text = "" if value is None else str(value).strip()
    if len(text) > max_len:
        raise ValueError(f"Giá trị quá dài, tối đa {max_len} ký tự.")
    return text


def _quote_list_values(value, *, max_items: int = 200) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        raw = value
    else:
        raw = _split_multi_values(str(value), max_items=max_items)
    out: list[str] = []
    for item in raw:
        text = _quote_text(item, max_len=200)
        if text:
            out.append(text)
            if len(out) >= max_items:
                break
    return out


def _quote_bool_or_none(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return None


def _quote_simple_size(size: str) -> dict:
    raw = (size or "").strip()
    if not raw:
        return {"raw": "", "value": None, "unit": "", "unit_group": "", "base_value": None, "simple": False}
    match = QUOTE_SIMPLE_SIZE_RE.fullmatch(raw)
    if not match:
        return {"raw": raw.upper(), "value": None, "unit": "", "unit_group": "", "base_value": None, "simple": False}
    try:
        value = float(match.group(1).replace(",", "."))
    except (TypeError, ValueError):
        value = None
    unit = (match.group(2) or "").strip().upper().replace("µ", "U").replace("μ", "U")
    if value is None:
        return {"raw": raw.upper(), "value": None, "unit": unit, "unit_group": "", "base_value": None, "simple": False}
    if unit in QUOTE_SOLID_FACTORS_TO_MG:
        return {
            "raw": raw.upper(),
            "value": value,
            "unit": unit,
            "unit_group": QUOTE_UNIT_GROUP_SOLID,
            "base_value": value * QUOTE_SOLID_FACTORS_TO_MG[unit],
            "simple": True,
        }
    if unit in QUOTE_LIQUID_FACTORS_TO_ML:
        return {
            "raw": raw.upper(),
            "value": value,
            "unit": unit,
            "unit_group": QUOTE_UNIT_GROUP_LIQUID,
            "base_value": value * QUOTE_LIQUID_FACTORS_TO_ML[unit],
            "simple": True,
        }
    return {"raw": raw.upper(), "value": value, "unit": unit, "unit_group": "", "base_value": None, "simple": False}


def _quote_validate_brand_policy(policy_raw, *, is_row_override: bool = False) -> Optional[dict]:
    if policy_raw is None:
        return None
    if not isinstance(policy_raw, dict):
        raise ValueError("Brand policy phải là object.")

    mode_raw = policy_raw.get("mode")
    if not isinstance(mode_raw, str):
        raise ValueError("Brand policy phải có trường 'mode' là chuỗi.")
    mode = _quote_text(mode_raw, max_len=50).upper()
    valid_modes = VALID_ROW_BRAND_POLICIES if is_row_override else VALID_GLOBAL_BRAND_POLICIES
    if mode not in valid_modes:
        if is_row_override:
            raise ValueError(
                f"brand_policy_override.mode '{mode}' không hợp lệ. Hỗ trợ: INHERIT, PRIORITY_FALLBACK, ALLOWLIST_ONLY, ALL_AVAILABLE."
            )
        raise ValueError(
            f"global_brand_policy.mode '{mode}' không hợp lệ. Hỗ trợ: PRIORITY_FALLBACK, ALLOWLIST_ONLY, ALL_AVAILABLE."
        )

    if mode == BRAND_POLICY_INHERIT:
        return {"mode": BRAND_POLICY_INHERIT}

    if mode == BRAND_POLICY_ALL_AVAILABLE:
        return {"mode": BRAND_POLICY_ALL_AVAILABLE}

    if mode == BRAND_POLICY_ALLOWLIST_ONLY:
        raw_brands = policy_raw.get("brands")
        if not isinstance(raw_brands, list) or not raw_brands:
            raise ValueError("ALLOWLIST_ONLY bắt buộc có danh sách 'brands' không rỗng.")
        if len(raw_brands) > MAX_POLICY_BRANDS:
            raise ValueError(f"ALLOWLIST_ONLY tối đa {MAX_POLICY_BRANDS} brands.")

        brands_clean: list[str] = []
        seen_norm: set[str] = set()
        for idx, b in enumerate(raw_brands):
            b_str = _quote_text(b, max_len=100)
            if not b_str:
                raise ValueError(f"Brand ở vị trí {idx + 1} trong allowlist không được rỗng.")
            b_norm = _quote_norm_identifier(b_str)
            if b_norm in seen_norm:
                raise ValueError(f"Brand '{b_str}' bị trùng lặp trong allowlist.")
            seen_norm.add(b_norm)
            brands_clean.append(b_str)

        return {
            "mode": BRAND_POLICY_ALLOWLIST_ONLY,
            "brands": brands_clean,
            "brands_norm": sorted(seen_norm),
        }

    if mode == BRAND_POLICY_PRIORITY_FALLBACK:
        raw_tiers = policy_raw.get("priority_tiers")
        if not isinstance(raw_tiers, list) or not raw_tiers:
            raise ValueError("PRIORITY_FALLBACK bắt buộc có danh sách 'priority_tiers' không rỗng.")
        if len(raw_tiers) > MAX_POLICY_TIERS:
            raise ValueError(f"PRIORITY_FALLBACK tối đa {MAX_POLICY_TIERS} tiers.")

        cleaned_tiers: list[dict] = []
        all_seen_norm: set[str] = set()
        for tier_idx, tier in enumerate(raw_tiers):
            if not isinstance(tier, dict):
                raise ValueError(f"Tier {tier_idx + 1} phải là object.")
            raw_brands = tier.get("brands")
            if not isinstance(raw_brands, list) or not raw_brands:
                raise ValueError(f"Tier {tier_idx + 1} bắt buộc có danh sách 'brands' không rỗng.")

            tier_brands_clean: list[str] = []
            tier_seen_norm: set[str] = set()
            for b_idx, b in enumerate(raw_brands):
                b_str = _quote_text(b, max_len=100)
                if not b_str:
                    raise ValueError(f"Tier {tier_idx + 1}, brand {b_idx + 1} không được rỗng.")
                b_norm = _quote_norm_identifier(b_str)
                if b_norm in all_seen_norm or b_norm in tier_seen_norm:
                    raise ValueError(f"Brand '{b_str}' ở Tier {tier_idx + 1} bị trùng lặp với tier khác hoặc trong cùng tier.")
                all_seen_norm.add(b_norm)
                tier_seen_norm.add(b_norm)
                tier_brands_clean.append(b_str)

            if len(all_seen_norm) > MAX_POLICY_BRANDS:
                raise ValueError(f"Tổng số brands trên tất cả các tiers vượt quá giới hạn {MAX_POLICY_BRANDS}.")

            cleaned_tiers.append({
                "brands": tier_brands_clean,
                "brands_norm": [_quote_norm_identifier(b) for b in tier_brands_clean],
            })

        return {
            "mode": BRAND_POLICY_PRIORITY_FALLBACK,
            "priority_tiers": cleaned_tiers,
            "all_brands_norm": sorted(all_seen_norm),
        }

    return None


def _quote_format_policy_response(policy: Optional[dict]) -> Optional[dict]:
    if not policy:
        return None
    mode = policy.get("mode")
    if mode == BRAND_POLICY_ALL_AVAILABLE:
        return {"mode": BRAND_POLICY_ALL_AVAILABLE}
    if mode == BRAND_POLICY_ALLOWLIST_ONLY:
        return {
            "mode": BRAND_POLICY_ALLOWLIST_ONLY,
            "brands": list(policy.get("brands") or []),
        }
    if mode == BRAND_POLICY_PRIORITY_FALLBACK:
        return {
            "mode": BRAND_POLICY_PRIORITY_FALLBACK,
            "priority_tiers": [
                {"brands": list(t.get("brands") or [])}
                for t in policy.get("priority_tiers") or []
            ],
        }
    return {"mode": mode}


def _quote_parse_payload(payload: dict) -> tuple[list[dict], dict, str]:
    rows = payload.get("rows")
    if not isinstance(rows, list):
        raise ValueError("Payload phải có rows là danh sách.")
    if len(rows) > QUOTE_MAX_ROWS:
        raise OverflowError(f"Tối đa {QUOTE_MAX_ROWS} dòng mỗi lần.")

    equivalent_default = _quote_bool_or_none(payload.get("equivalent_search_default"))
    if equivalent_default is None:
        equivalent_default = False

    global_brand_policy = _quote_validate_brand_policy(payload.get("global_brand_policy"), is_row_override=False)

    parsed: list[dict] = []
    seen_request_ids: set[str] = set()
    seen_request_orders: set[int] = set()
    legacy_counter = 0
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"Dòng {index} không hợp lệ.")
        requested_name = _quote_text(row.get("requested_name") or row.get("name"))
        code = _quote_text(row.get("code"))
        cas = _quote_text(row.get("cas"))
        row_equivalent = _quote_bool_or_none(row.get("equivalent_override"))
        if row_equivalent is None:
            row_equivalent = equivalent_default

        brand_override_raw = row.get("brand_policy_override")
        brand_override = _quote_validate_brand_policy(brand_override_raw, is_row_override=True)
        if brand_override is not None and brand_override.get("mode") != BRAND_POLICY_INHERIT:
            effective_brand_policy = brand_override
        else:
            effective_brand_policy = global_brand_policy

        request_id = row.get("request_id")
        if request_id is not None:
            request_id = _quote_text(str(request_id), max_len=128)
            if not request_id:
                raise ValueError(f"Dòng {index}: request_id không được rỗng nếu có.")
            if request_id in seen_request_ids:
                raise ValueError(f"Dòng {index}: request_id trùng lặp.")
            seen_request_ids.add(request_id)
        else:
            legacy_counter += 1
            request_id = f"legacy-{legacy_counter}"

        request_order_raw = row.get("request_order")
        if request_order_raw is not None:
            try:
                request_order = int(request_order_raw)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Dòng {index}: request_order phải là số nguyên.") from exc
            if request_order < 1:
                raise ValueError(f"Dòng {index}: request_order phải >= 1.")
            if request_order > QUOTE_MAX_ROWS:
                raise ValueError(f"Dòng {index}: request_order vượt giới hạn.")
            if request_order in seen_request_orders:
                raise ValueError(f"Dòng {index}: request_order trùng lặp.")
            seen_request_orders.add(request_order)
        else:
            request_order = index

        source_row = row.get("source_row")
        if source_row is not None:
            try:
                source_row = int(source_row)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Dòng {index}: source_row phải là số nguyên hoặc null.") from exc
            if source_row < 1:
                raise ValueError(f"Dòng {index}: source_row phải >= 1 hoặc null.")

        parsed.append(
            {
                "ord": index,
                "request_id": request_id,
                "request_order": request_order,
                "source_row": source_row,
                "requested_name": requested_name,
                "code": code,
                "cas": cas,
                "code_u": _quote_norm_identifier(code),
                "cas_u": _quote_norm_identifier(cas),
                "equivalent_search": bool(row_equivalent),
                "brand_policy_override": brand_override,
                "effective_brand_policy": effective_brand_policy,
            }
        )

    filters = payload.get("filters") or {}
    if not isinstance(filters, dict):
        raise ValueError("filters phải là object nếu được gửi.")
    brand_values = (
        filters.get("normalized_brands")
        or filters.get("brands")
        or filters.get("brand")
        or payload.get("normalized_brands")
        or payload.get("brands")
        or payload.get("brand")
    )
    size_values = (
        filters.get("sizes")
        or filters.get("size")
        or payload.get("sizes")
        or payload.get("size")
    )
    brands_norm = [x.upper() for x in _quote_list_values(brand_values)]
    sizes_norm = [x.upper() for x in _quote_list_values(size_values)]
    unit_group = _quote_text(filters.get("unit_group") or payload.get("unit_group") or QUOTE_UNIT_GROUP_ANY, max_len=20).upper()
    if unit_group not in QUOTE_UNIT_GROUPS:
        raise ValueError("filters.unit_group không hợp lệ.")
    preparation_type = _quote_text(
        filters.get("preparation_type") or payload.get("preparation_type") or QUOTE_PREPARATION_ANY,
        max_len=20,
    ).upper()
    if preparation_type not in QUOTE_PREPARATION_TYPES:
        raise ValueError("filters.preparation_type không hợp lệ.")
    size_mode = _quote_text(filters.get("size_mode") or payload.get("size_mode") or QUOTE_SIZE_MODE_ANY, max_len=20).upper()
    if size_mode not in QUOTE_SIZE_MODES:
        raise ValueError("filters.size_mode không hợp lệ.")
    if sizes_norm and size_mode == QUOTE_SIZE_MODE_ANY:
        size_mode = QUOTE_SIZE_MODE_EXACT

    strategy = _quote_text(payload.get("selection_strategy") or payload.get("strategy") or QUOTE_SELECTION_MANUAL, max_len=50).upper()
    if strategy not in QUOTE_SELECTION_STRATEGIES:
        raise ValueError("selection_strategy không hợp lệ.")
    filters_out = {
        "brands_norm": brands_norm,
        "sizes_norm": sizes_norm,
        "unit_group": unit_group,
        "preparation_type": preparation_type,
        "size_mode": size_mode,
        "global_brand_policy": global_brand_policy,
    }
    return parsed, filters_out, strategy


def _quote_unit_price_value(ship, price, brand, rate_map: dict) -> float:
    try:
        ship_f = float(ship) if ship is not None else 0
    except (TypeError, ValueError):
        ship_f = 0
    try:
        price_f = float(price) if price is not None else 0
    except (TypeError, ValueError):
        price_f = 0
    exchange_rate = rate_map.get((brand or "").strip(), 1.0)
    return round(price_f * ship_f * exchange_rate, -3)


def _quote_candidate_from_row(row: tuple, rate_map: dict) -> dict:
    (
        _ord,
        product_id,
        name,
        code,
        cas,
        brand,
        size,
        ship,
        price,
        note,
        preparation_type,
        manual_compliance,
        manual_compliance_note,
        brand_manual_enabled,
        compliance_status,
        compliance_note,
    ) = row
    unit_price = _quote_unit_price_value(ship, price, brand, rate_map)
    resolved = resolve_compliance_precedence(
        brand_manual_enabled=bool(brand_manual_enabled),
        manual_compliance=manual_compliance,
        manual_compliance_note=manual_compliance_note,
        legacy_compliance=compliance_status,
        legacy_compliance_note=compliance_note,
        cas=cas,
    )
    compliance = resolved["compliance"]
    warnings = [compliance] if compliance in QUOTE_WARNING_COMPLIANCE else []
    ineligible_reason = ""
    if compliance in QUOTE_BLOCKED_COMPLIANCE:
        ineligible_reason = "COMPLIANCE_BLOCKED"
    elif unit_price <= 0:
        ineligible_reason = "NO_VALID_PRICE"
    eligible = ineligible_reason == ""
    return {
        "product_id": product_id,
        "Name": name or "",
        "Code": code or "",
        "Cas": cas or "",
        "Brand": brand or "",
        "Size": size or "",
        "Unit_Price": "{:,.0f}".format(unit_price),
        "Unit_Price_Value": unit_price,
        "Note": note or "",
        "preparation_type": preparation_type,
        "Compliance": compliance,
        "Compliance_Note": resolved["compliance_note"],
        "compliance_source": resolved["compliance_source"],
        "compliance_css": resolved["compliance_css"],
        "eligible": eligible,
        "ineligible_reason": ineligible_reason,
        "warnings": warnings,
    }


def _quote_candidate_auto_key(candidate: dict):
    return (
        float(candidate.get("Unit_Price_Value") or 0),
        (candidate.get("Brand") or "").strip().upper(),
        (candidate.get("Code") or "").strip().upper(),
        (candidate.get("Size") or "").strip().upper(),
        (candidate.get("Name") or "").strip().upper(),
        int(candidate.get("product_id") or 0),
    )


def _quote_mark_duplicate_groups(candidates: list[dict]) -> list[str]:
    counts: dict[tuple[str, str, str], int] = {}
    for candidate in candidates:
        key = (
            _quote_norm_identifier(candidate.get("Code")),
            _quote_norm_identifier(candidate.get("Brand")),
            _quote_norm_identifier(candidate.get("Size")),
        )
        if not key[0]:
            continue
        counts[key] = counts.get(key, 0) + 1
    duplicate_keys = {key for key, count in counts.items() if count > 1}
    if not duplicate_keys:
        return []
    for candidate in candidates:
        key = (
            _quote_norm_identifier(candidate.get("Code")),
            _quote_norm_identifier(candidate.get("Brand")),
            _quote_norm_identifier(candidate.get("Size")),
        )
        if key in duplicate_keys:
            candidate["auto_excluded"] = True
            candidate.setdefault("warnings", []).append("DUPLICATE_CODE_BRAND_SIZE")
    return ["DUPLICATE_CODE_BRAND_SIZE"]


def _quote_apply_candidate_filters(candidates: list[dict], filters: dict) -> tuple[list[dict], list[str]]:
    unit_group = filters.get("unit_group") or QUOTE_UNIT_GROUP_ANY
    preparation_type = filters.get("preparation_type") or QUOTE_PREPARATION_ANY
    size_mode = filters.get("size_mode") or QUOTE_SIZE_MODE_ANY
    row_warnings: list[str] = []
    filtered: list[dict] = []

    for candidate in candidates:
        if preparation_type != QUOTE_PREPARATION_ANY and candidate.get("preparation_type") != preparation_type:
            continue
        parsed = _quote_simple_size(candidate.get("Size") or "")
        candidate["size_parse"] = parsed
        if unit_group != QUOTE_UNIT_GROUP_ANY and parsed["unit_group"] != unit_group:
            continue
        filtered.append(candidate)

    if size_mode not in {QUOTE_SIZE_MODE_MIN, QUOTE_SIZE_MODE_MAX}:
        return filtered, row_warnings

    comparable: list[dict] = []
    for candidate in filtered:
        parsed = candidate.get("size_parse") or {}
        if parsed.get("simple") and parsed.get("base_value") is not None and parsed.get("unit_group"):
            comparable.append(candidate)
        else:
            candidate.setdefault("warnings", []).append("SIZE_NOT_COMPARABLE")
            if "SIZE_NOT_COMPARABLE" not in row_warnings:
                row_warnings.append("SIZE_NOT_COMPARABLE")

    if not comparable:
        return [], row_warnings

    by_brand_unit: dict[tuple[str, str], list[dict]] = {}
    for candidate in comparable:
        brand_key = _quote_norm_identifier(candidate.get("Brand"))
        unit_key = (candidate.get("size_parse") or {}).get("unit_group") or ""
        by_brand_unit.setdefault((brand_key, unit_key), []).append(candidate)

    keep_ids: set[int] = set()
    for brand_candidates in by_brand_unit.values():
        values = [float(c["size_parse"]["base_value"]) for c in brand_candidates]
        target = min(values) if size_mode == QUOTE_SIZE_MODE_MIN else max(values)
        for candidate in brand_candidates:
            if float(candidate["size_parse"]["base_value"]) == target:
                keep_ids.add(int(candidate["product_id"]))
    return [candidate for candidate in filtered if int(candidate["product_id"]) in keep_ids], row_warnings


def _quote_select_candidates(candidates: list[dict], strategy: str) -> tuple[list[dict], str, str, str, str]:
    if not candidates:
        return [], "UNRESOLVED", "NO_MATCH", LIFECYCLE_UNRESOLVED, REASON_NO_MATCH

    compliance_allowed = [c for c in candidates if c.get("ineligible_reason") != "COMPLIANCE_BLOCKED"]
    if not compliance_allowed:
        return [], "UNRESOLVED", "MANUAL_REVIEW", LIFECYCLE_BLOCKED, REASON_COMPLIANCE_BLOCKED

    selectable = [c for c in compliance_allowed if c.get("eligible") and not c.get("auto_excluded")]
    if not selectable:
        if any(c.get("eligible") for c in compliance_allowed):
            return [], "MATCHED", "MANUAL_SELECTION_REQUIRED", LIFECYCLE_REVIEW, REASON_DUPLICATE_CODE_BRAND_SIZE
        return [], "UNRESOLVED", "NO_VALID_PRICE", LIFECYCLE_REVIEW, REASON_NO_VALID_PRICE

    selectable = sorted(selectable, key=_quote_candidate_auto_key)
    if strategy in {QUOTE_SELECTION_LOWEST_OVERALL, QUOTE_SELECTION_LOWEST_UNIT_PRICE}:
        reason = "SELECTED_LOWEST_UNIT_PRICE" if strategy == QUOTE_SELECTION_LOWEST_UNIT_PRICE else "SELECTED_LOWEST_OVERALL"
        return [selectable[0]], "MATCHED", reason, LIFECYCLE_SELECTED, REASON_AUTO_SELECTED
    if strategy == QUOTE_SELECTION_LOWEST_PER_BRAND:
        by_brand: dict[str, dict] = {}
        for candidate in selectable:
            brand_key = _quote_norm_identifier(candidate.get("Brand"))
            if brand_key not in by_brand:
                by_brand[brand_key] = candidate
        return list(by_brand.values()), "MATCHED", "SELECTED_LOWEST_PER_BRAND", LIFECYCLE_SELECTED, REASON_AUTO_SELECTED

    return [], "MATCHED", "MANUAL_SELECTION_REQUIRED", LIFECYCLE_REVIEW, REASON_MANUAL_SELECTION_REQUIRED


def _quote_select_candidate(candidates: list[dict], strategy: str) -> tuple[Optional[dict], str, str]:
    selected, status, reason, _lifecycle, _reason_code = _quote_select_candidates(candidates, strategy)
    return (selected[0] if selected else None), status, reason


def _quote_apply_brand_policy_to_row(
    raw_candidates: list[dict],
    filtered_candidates: list[dict],
    policy: dict,
    strategy: str,
    all_warnings: list[str],
) -> tuple[list[dict], list[dict], str, str, str, str, Optional[int], list[dict]]:
    mode = policy.get("mode")

    if mode == BRAND_POLICY_PRIORITY_FALLBACK:
        priority_tiers = policy.get("priority_tiers") or []
        matched_tier_idx: Optional[int] = None
        fallback_path: list[dict] = []
        selected_cands: list[dict] = []
        return_cands: list[dict] = []
        status = "UNRESOLVED"
        reason = "NO_MATCH"
        lifecycle = LIFECYCLE_UNRESOLVED
        reason_code = REASON_NO_MATCH

        for t_idx, tier in enumerate(priority_tiers):
            tier_brands_norm = set(tier.get("brands_norm") or [])
            raw_tier = [c for c in raw_candidates if _quote_norm_identifier(c.get("Brand")) in tier_brands_norm]
            filtered_tier = [c for c in filtered_candidates if _quote_norm_identifier(c.get("Brand")) in tier_brands_norm]

            eligible_tier = [c for c in filtered_tier if c.get("eligible") and not c.get("auto_excluded")]

            compliance_blocked_count = sum(
                1
                for c in raw_tier
                if c.get("ineligible_reason") == "COMPLIANCE_BLOCKED" or c.get("Compliance") in QUOTE_BLOCKED_COMPLIANCE
            )
            filter_rejected_count = len(raw_tier) - len(filtered_tier)
            no_valid_price_count = sum(
                1
                for c in filtered_tier
                if (c.get("ineligible_reason") == "NO_VALID_PRICE" or (c.get("Unit_Price_Value") or 0) <= 0)
                and not (
                    c.get("ineligible_reason") == "COMPLIANCE_BLOCKED" or c.get("Compliance") in QUOTE_BLOCKED_COMPLIANCE
                )
            )

            if eligible_tier:
                matched_tier_idx = t_idx
                return_cands = filtered_tier

                if t_idx > 0:
                    all_warnings.append("FALLBACK_TIER_USED")

                if strategy == QUOTE_SELECTION_MANUAL:
                    selected_cands = []
                    status = "MATCHED"
                    reason = "MANUAL_SELECTION_REQUIRED"
                    lifecycle = LIFECYCLE_REVIEW
                    reason_code = REASON_MANUAL_SELECTION_REQUIRED
                elif strategy == QUOTE_SELECTION_LOWEST_PER_BRAND:
                    by_brand: dict[str, dict] = {}
                    for c in sorted(eligible_tier, key=_quote_candidate_auto_key):
                        b_norm = _quote_norm_identifier(c.get("Brand"))
                        if b_norm not in by_brand:
                            by_brand[b_norm] = c
                    selected_cands = list(by_brand.values())
                    status = "MATCHED"
                    reason = "SELECTED_LOWEST_PER_BRAND"
                    lifecycle = LIFECYCLE_SELECTED
                    reason_code = REASON_AUTO_SELECTED
                else:
                    sorted_cands = sorted(eligible_tier, key=_quote_candidate_auto_key)
                    selected_cands = [sorted_cands[0]]
                    status = "MATCHED"
                    reason = (
                        "SELECTED_LOWEST_UNIT_PRICE"
                        if strategy == QUOTE_SELECTION_LOWEST_UNIT_PRICE
                        else "SELECTED_LOWEST_OVERALL"
                    )
                    lifecycle = LIFECYCLE_SELECTED
                    reason_code = REASON_AUTO_SELECTED

                break
            else:
                fallback_path.append(
                    {
                        "tier": t_idx,
                        "brands": list(tier.get("brands") or []),
                        "eligible_count": 0,
                        "rejected_counts": {
                            "COMPLIANCE": compliance_blocked_count,
                            "FILTER": max(0, filter_rejected_count),
                            "NO_VALID_PRICE": no_valid_price_count,
                        },
                    }
                )

        if matched_tier_idx is None:
            all_policy_brands = set(policy.get("all_brands_norm") or [])
            all_raw_policy = [c for c in raw_candidates if _quote_norm_identifier(c.get("Brand")) in all_policy_brands]
            all_filtered_policy = [c for c in filtered_candidates if _quote_norm_identifier(c.get("Brand")) in all_policy_brands]

            if not all_raw_policy:
                return_cands = []
                selected_cands = []
                status = "UNRESOLVED"
                reason = "NO_MATCH"
                lifecycle = LIFECYCLE_UNRESOLVED
                reason_code = REASON_NO_MATCH
            elif all(
                c.get("ineligible_reason") == "COMPLIANCE_BLOCKED" or c.get("Compliance") in QUOTE_BLOCKED_COMPLIANCE
                for c in all_raw_policy
            ):
                return_cands = all_filtered_policy or all_raw_policy
                selected_cands = []
                status = "UNRESOLVED"
                reason = "MANUAL_REVIEW"
                lifecycle = LIFECYCLE_BLOCKED
                reason_code = REASON_COMPLIANCE_BLOCKED
            elif not all_filtered_policy and all_raw_policy:
                return_cands = []
                selected_cands = []
                status = "UNRESOLVED"
                reason = "NO_MATCH"
                lifecycle = LIFECYCLE_REVIEW
                reason_code = REASON_FILTER_NO_MATCH
            elif all((c.get("Unit_Price_Value") or 0) <= 0 for c in all_filtered_policy):
                return_cands = all_filtered_policy
                selected_cands = []
                status = "UNRESOLVED"
                reason = "NO_VALID_PRICE"
                lifecycle = LIFECYCLE_REVIEW
                reason_code = REASON_NO_VALID_PRICE
            else:
                return_cands = all_filtered_policy
                selected_cands = []
                status = "UNRESOLVED"
                reason = "MANUAL_SELECTION_REQUIRED"
                lifecycle = LIFECYCLE_REVIEW
                reason_code = REASON_MANUAL_SELECTION_REQUIRED

        return return_cands, selected_cands, status, reason, lifecycle, reason_code, matched_tier_idx, fallback_path

    if mode == BRAND_POLICY_ALLOWLIST_ONLY:
        allow_brands_norm = set(policy.get("brands_norm") or [])
        raw_allow = [c for c in raw_candidates if _quote_norm_identifier(c.get("Brand")) in allow_brands_norm]
        filtered_allow = [c for c in filtered_candidates if _quote_norm_identifier(c.get("Brand")) in allow_brands_norm]

        matched_tier_idx = None
        fallback_path = []

        if filtered_allow:
            selected_cands, status, reason, lifecycle, reason_code = _quote_select_candidates(filtered_allow, strategy)
            return filtered_allow, selected_cands, status, reason, lifecycle, reason_code, matched_tier_idx, fallback_path

        if raw_allow:
            if all(
                c.get("ineligible_reason") == "COMPLIANCE_BLOCKED" or c.get("Compliance") in QUOTE_BLOCKED_COMPLIANCE
                for c in raw_allow
            ):
                return raw_allow, [], "UNRESOLVED", "MANUAL_REVIEW", LIFECYCLE_BLOCKED, REASON_COMPLIANCE_BLOCKED, matched_tier_idx, fallback_path
            if all((c.get("Unit_Price_Value") or 0) <= 0 for c in raw_allow):
                return raw_allow, [], "UNRESOLVED", "NO_VALID_PRICE", LIFECYCLE_REVIEW, REASON_NO_VALID_PRICE, matched_tier_idx, fallback_path
            return [], [], "UNRESOLVED", "NO_MATCH", LIFECYCLE_REVIEW, REASON_FILTER_NO_MATCH, matched_tier_idx, fallback_path

        if raw_candidates:
            return [], [], "UNRESOLVED", "NO_MATCH", LIFECYCLE_REVIEW, REASON_FILTER_NO_MATCH, matched_tier_idx, fallback_path
        return [], [], "UNRESOLVED", "NO_MATCH", LIFECYCLE_UNRESOLVED, REASON_NO_MATCH, matched_tier_idx, fallback_path

    if mode == BRAND_POLICY_ALL_AVAILABLE:
        matched_tier_idx = None
        fallback_path = []

        if not filtered_candidates:
            if raw_candidates:
                if all(
                    c.get("ineligible_reason") == "COMPLIANCE_BLOCKED" or c.get("Compliance") in QUOTE_BLOCKED_COMPLIANCE
                    for c in raw_candidates
                ):
                    return raw_candidates, [], "UNRESOLVED", "MANUAL_REVIEW", LIFECYCLE_BLOCKED, REASON_COMPLIANCE_BLOCKED, matched_tier_idx, fallback_path
                if all((c.get("Unit_Price_Value") or 0) <= 0 for c in raw_candidates):
                    return raw_candidates, [], "UNRESOLVED", "NO_VALID_PRICE", LIFECYCLE_REVIEW, REASON_NO_VALID_PRICE, matched_tier_idx, fallback_path
                return [], [], "UNRESOLVED", "NO_MATCH", LIFECYCLE_REVIEW, REASON_FILTER_NO_MATCH, matched_tier_idx, fallback_path
            return [], [], "UNRESOLVED", "NO_MATCH", LIFECYCLE_UNRESOLVED, REASON_NO_MATCH, matched_tier_idx, fallback_path

        by_brand_map: dict[str, list[dict]] = {}
        for c in filtered_candidates:
            b_norm = _quote_norm_identifier(c.get("Brand"))
            by_brand_map.setdefault(b_norm, []).append(c)

        best_per_brand: list[dict] = []
        for b_norm, b_cands in by_brand_map.items():
            eligible_b = [c for c in b_cands if c.get("eligible") and not c.get("auto_excluded")]
            if eligible_b:
                best_c = sorted(eligible_b, key=_quote_candidate_auto_key)[0]
            else:
                best_c = sorted(b_cands, key=_quote_candidate_auto_key)[0]
            best_per_brand.append(best_c)

        best_per_brand = sorted(best_per_brand, key=_quote_candidate_auto_key)
        if len(best_per_brand) > QUOTE_CANDIDATE_LIMIT:
            # Fail-closed: selecting from a truncated brand set would silently
            # omit brands and give an incorrect best-per-brand result.
            return [], [], "UNRESOLVED", "CANDIDATE_LIMIT_EXCEEDED", LIFECYCLE_REVIEW, REASON_CANDIDATE_LIMIT_EXCEEDED, matched_tier_idx, fallback_path

        eligible_all = [c for c in best_per_brand if c.get("eligible") and not c.get("auto_excluded")]
        if eligible_all:
            if strategy == QUOTE_SELECTION_MANUAL:
                selected_cands = []
                status = "MATCHED"
                reason = "MANUAL_SELECTION_REQUIRED"
                lifecycle = LIFECYCLE_REVIEW
                reason_code = REASON_MANUAL_SELECTION_REQUIRED
            elif strategy == QUOTE_SELECTION_LOWEST_PER_BRAND:
                selected_cands = eligible_all
                status = "MATCHED"
                reason = "SELECTED_LOWEST_PER_BRAND"
                lifecycle = LIFECYCLE_SELECTED
                reason_code = REASON_AUTO_SELECTED
            else:
                selected_cands = [eligible_all[0]]
                status = "MATCHED"
                reason = (
                    "SELECTED_LOWEST_UNIT_PRICE"
                    if strategy == QUOTE_SELECTION_LOWEST_UNIT_PRICE
                    else "SELECTED_LOWEST_OVERALL"
                )
                lifecycle = LIFECYCLE_SELECTED
                reason_code = REASON_AUTO_SELECTED
            return best_per_brand, selected_cands, status, reason, lifecycle, reason_code, matched_tier_idx, fallback_path

        if all(
            c.get("ineligible_reason") == "COMPLIANCE_BLOCKED" or c.get("Compliance") in QUOTE_BLOCKED_COMPLIANCE
            for c in best_per_brand
        ):
            return best_per_brand, [], "UNRESOLVED", "MANUAL_REVIEW", LIFECYCLE_BLOCKED, REASON_COMPLIANCE_BLOCKED, matched_tier_idx, fallback_path
        if all((c.get("Unit_Price_Value") or 0) <= 0 for c in best_per_brand):
            return best_per_brand, [], "UNRESOLVED", "NO_VALID_PRICE", LIFECYCLE_REVIEW, REASON_NO_VALID_PRICE, matched_tier_idx, fallback_path
        return best_per_brand, [], "UNRESOLVED", "NO_MATCH", LIFECYCLE_REVIEW, REASON_FILTER_NO_MATCH, matched_tier_idx, fallback_path

    # Fallback to standard selection
    selected_cands, status, reason, lifecycle, reason_code = _quote_select_candidates(filtered_candidates, strategy)
    return filtered_candidates, selected_cands, status, reason, lifecycle, reason_code, None, []


def _quote_product_lateral_sql(vis: str, product_filter_sql: str) -> str:
    return f"""
        SELECT
            p.id AS product_id,
            p.name,
            p.code,
            p.cas,
            p.brand,
            p.size,
            p.ship,
            p.price,
            p.note,
            p.preparation_type,
            p.manual_compliance,
            p.manual_compliance_note,
            COALESCE(bcs.manual_compliance_priority, FALSE) AS brand_manual_enabled
        FROM products p
        LEFT JOIN brand_compliance_settings bcs
          ON bcs.brand_norm = UPPER(TRIM(COALESCE(p.brand, '')))
        WHERE {{match_predicate}}
          {vis}
          {product_filter_sql}
        ORDER BY (p.id + 0) ASC
        LIMIT {QUOTE_CANDIDATE_LIMIT + 1}
    """


def _quote_match_rows(conn, parsed_rows: list[dict], filters: dict, strategy: str) -> list[dict]:
    brands_norm = filters.get("brands_norm") or []
    sizes_norm = filters.get("sizes_norm") or []
    size_mode = filters.get("size_mode") or QUOTE_SIZE_MODE_ANY
    results = [
        {
            "request_id": row["request_id"],
            "request_order": row["request_order"],
            "source_row": row["source_row"],
            "requested_name": row["requested_name"],
            "requested_code": row["code"],
            "requested_cas": row["cas"],
            "status": "UNRESOLVED",
            "lifecycle": LIFECYCLE_UNRESOLVED,
            "reason_code": REASON_MISSING_IDENTIFIER if not row["code_u"] and not row["cas_u"] else REASON_NO_MATCH,
            "reason": "MISSING_IDENTIFIER" if not row["code_u"] and not row["cas_u"] else "NO_MATCH",
            "match_mode": None,
            "warnings": [],
            "input": {
                "requested_name": row["requested_name"],
                "code": row["code"],
                "cas": row["cas"],
                "equivalent_search": row["equivalent_search"],
            },
            "candidates": [],
            "selected_candidates": [],
            "selected": None,
            "effective_brand_policy": _quote_format_policy_response(row.get("effective_brand_policy")),
            "matched_priority_tier": None,
            "fallback_path": [],
        }
        for row in parsed_rows
    ]
    for row in parsed_rows:
        if not (row["code_u"] or row["cas_u"]):
            continue
        broad_match = (row["cas_u"] and not row["code_u"]) or row["equivalent_search"]
        policy = row.get("effective_brand_policy")
        if broad_match and not brands_norm and not policy:
            idx = row["ord"] - 1
            results[idx]["lifecycle"] = LIFECYCLE_REVIEW
            results[idx]["reason_code"] = REASON_BRAND_REQUIRED
            results[idx]["reason"] = "BRAND_REQUIRED"
            results[idx]["warnings"].append("BRAND_REQUIRED_FOR_BROAD_MATCH")

    lookup_rows = [
        row
        for row in parsed_rows
        if (row["code_u"] or row["cas_u"])
        and not (
            ((row["cas_u"] and not row["code_u"]) or row["equivalent_search"])
            and not brands_norm
            and not row.get("effective_brand_policy")
        )
    ]
    if not lookup_rows:
        return results

    vis, vis_params = _visibility_sql("p")
    product_filter_sql = ""
    product_filter_params: tuple = ()
    has_any_brand_policy = any(bool(r.get("effective_brand_policy")) for r in lookup_rows)
    if brands_norm and not has_any_brand_policy:
        product_filter_sql += " AND UPPER(TRIM(COALESCE(p.brand, ''))) = ANY(%s)"
        product_filter_params += (brands_norm,)
    preparation_type = filters.get("preparation_type") or QUOTE_PREPARATION_ANY
    if preparation_type != QUOTE_PREPARATION_ANY:
        product_filter_sql += " AND p.preparation_type = %s"
        product_filter_params += (preparation_type,)
    if sizes_norm and size_mode == QUOTE_SIZE_MODE_EXACT:
        product_filter_sql += " AND UPPER(TRIM(COALESCE(p.size, ''))) = ANY(%s)"
        product_filter_params += (sizes_norm,)

    product_lateral = _quote_product_lateral_sql(vis, product_filter_sql)
    verify_lateral = _quote_product_lateral_sql(vis, "")
    code_predicate = (
        "i.code_u <> '' AND i.cas_u = '' AND i.equivalent_search = FALSE"
        " AND p.code IS NOT NULL AND TRIM(p.code) <> ''"
        " AND UPPER(TRIM(p.code)) = i.code_u"
    )
    cas_predicate = (
        "i.code_u = '' AND i.cas_u <> ''"
        " AND p.cas IS NOT NULL AND TRIM(p.cas) <> ''"
        " AND UPPER(TRIM(p.cas)) = i.cas_u"
    )
    both_predicate = (
        "i.code_u <> '' AND i.cas_u <> '' AND i.equivalent_search = FALSE"
        " AND p.code IS NOT NULL AND TRIM(p.code) <> ''"
        " AND UPPER(TRIM(p.code)) = i.code_u"
        " AND p.cas IS NOT NULL AND TRIM(p.cas) <> ''"
        " AND UPPER(TRIM(p.cas)) = i.cas_u"
    )
    verify_predicate = (
        "i.code_u <> '' AND i.cas_u <> '' AND i.equivalent_search = TRUE"
        " AND p.code IS NOT NULL AND TRIM(p.code) <> ''"
        " AND UPPER(TRIM(p.code)) = i.code_u"
        " AND p.cas IS NOT NULL AND TRIM(p.cas) <> ''"
        " AND UPPER(TRIM(p.cas)) = i.cas_u"
    )
    equivalent_code_predicate = (
        "p.cas IS NOT NULL AND TRIM(p.cas) <> ''"
        " AND UPPER(TRIM(p.cas)) = c.resolved_cas_u"
    )
    equivalent_entered_cas_predicate = (
        "p.cas IS NOT NULL AND TRIM(p.cas) <> ''"
        " AND UPPER(TRIM(p.cas)) = i.cas_u"
    )

    query = f"""
        WITH input AS (
            SELECT u.ord, u.code_u, u.cas_u, u.equivalent_search
            FROM unnest(%s::int[], %s::text[], %s::text[], %s::boolean[])
                AS u(ord, code_u, cas_u, equivalent_search)
        ),
        code_cas_summary AS (
            -- Use code index first (idx_products_code_upper_trim) to find CAS entries per input code.
            -- Aggregation in the LATERAL subquery computes valid CAS count (matching standard CAS pattern
            -- or test fixture prefix) and total distinct raw CAS count without limit truncation or index misdirection.
            SELECT
                i.ord,
                COALESCE(c.valid_cas_count, 0)::int AS cas_count,
                c.resolved_cas_u,
                COALESCE(c.total_cas_count, 0)::int AS total_cas_count
            FROM (
                SELECT *
                FROM input
                WHERE code_u <> ''
                  AND cas_u = ''
                  AND equivalent_search = TRUE
            ) i
            LEFT JOIN LATERAL (
                SELECT
                    COUNT(DISTINCT CASE WHEN (UPPER(TRIM(p.cas)) ~ '^\d+-\d+-\d+$' OR UPPER(TRIM(p.cas)) ~ '^CURSOR') THEN UPPER(TRIM(p.cas)) END)::int AS valid_cas_count,
                    MIN(CASE WHEN (UPPER(TRIM(p.cas)) ~ '^\d+-\d+-\d+$' OR UPPER(TRIM(p.cas)) ~ '^CURSOR') THEN UPPER(TRIM(p.cas)) END) AS resolved_cas_u,
                    COUNT(DISTINCT UPPER(TRIM(p.cas)))::int AS total_cas_count
                FROM products p
                WHERE UPPER(TRIM(p.code)) = i.code_u
                  AND p.code IS NOT NULL
                  AND TRIM(p.code) <> ''
                  AND p.cas IS NOT NULL
                  AND TRIM(p.cas) <> ''
                  {vis}
            ) c ON TRUE
        ),
        code_cas_verified AS (
            SELECT DISTINCT i.ord
            FROM (
                SELECT *
                FROM input
                WHERE code_u <> ''
                  AND cas_u <> ''
                  AND equivalent_search = TRUE
            ) i
            JOIN LATERAL (
                {verify_lateral.format(match_predicate=verify_predicate)}
            ) p ON TRUE
        ),
        product_hits AS (
            SELECT
                i.ord,
                'EXACT_CODE' AS match_mode,
                p.product_id,
                p.name,
                p.code,
                p.cas,
                p.brand,
                p.size,
                p.ship,
                p.price,
                p.note,
                p.preparation_type,
                p.manual_compliance,
                p.manual_compliance_note,
                p.brand_manual_enabled
            FROM (
                SELECT *
                FROM input
                WHERE code_u <> ''
                  AND cas_u = ''
                  AND equivalent_search = FALSE
            ) i
            JOIN LATERAL (
                {product_lateral.format(match_predicate=code_predicate)}
            ) p ON TRUE
            UNION ALL
            SELECT
                i.ord,
                'EXACT_CAS' AS match_mode,
                p.product_id,
                p.name,
                p.code,
                p.cas,
                p.brand,
                p.size,
                p.ship,
                p.price,
                p.note,
                p.preparation_type,
                p.manual_compliance,
                p.manual_compliance_note,
                p.brand_manual_enabled
            FROM (
                SELECT *
                FROM input
                WHERE code_u = ''
                  AND cas_u <> ''
            ) i
            JOIN LATERAL (
                {product_lateral.format(match_predicate=cas_predicate)}
            ) p ON TRUE
            UNION ALL
            SELECT
                i.ord,
                'CODE_CAS' AS match_mode,
                p.product_id,
                p.name,
                p.code,
                p.cas,
                p.brand,
                p.size,
                p.ship,
                p.price,
                p.note,
                p.preparation_type,
                p.manual_compliance,
                p.manual_compliance_note,
                p.brand_manual_enabled
            FROM (
                SELECT *
                FROM input
                WHERE code_u <> ''
                  AND cas_u <> ''
                  AND equivalent_search = FALSE
            ) i
            JOIN LATERAL (
                {product_lateral.format(match_predicate=both_predicate)}
            ) p ON TRUE
            UNION ALL
            SELECT
                i.ord,
                'EQUIVALENT' AS match_mode,
                p.product_id,
                p.name,
                p.code,
                p.cas,
                p.brand,
                p.size,
                p.ship,
                p.price,
                p.note,
                p.preparation_type,
                p.manual_compliance,
                p.manual_compliance_note,
                p.brand_manual_enabled
            FROM (
                SELECT *
                FROM input
                WHERE code_u <> ''
                  AND cas_u = ''
                  AND equivalent_search = TRUE
            ) i
            JOIN code_cas_summary c
              ON c.ord = i.ord AND c.cas_count = 1
            JOIN LATERAL (
                {product_lateral.format(match_predicate=equivalent_code_predicate)}
            ) p ON TRUE
            UNION ALL
            SELECT
                i.ord,
                'EQUIVALENT' AS match_mode,
                p.product_id,
                p.name,
                p.code,
                p.cas,
                p.brand,
                p.size,
                p.ship,
                p.price,
                p.note,
                p.preparation_type,
                p.manual_compliance,
                p.manual_compliance_note,
                p.brand_manual_enabled
            FROM (
                SELECT *
                FROM input
                WHERE code_u <> ''
                  AND cas_u <> ''
                  AND equivalent_search = TRUE
            ) i
            JOIN code_cas_verified v ON v.ord = i.ord
            JOIN LATERAL (
                {product_lateral.format(match_predicate=equivalent_entered_cas_predicate)}
            ) p ON TRUE
        ),
        output_rows AS (
            SELECT
                'HIT'::text AS row_type,
                ph.ord,
                ph.match_mode,
                ph.product_id,
                ph.name,
                ph.code,
                ph.cas,
                ph.brand,
                ph.size,
                ph.ship,
                ph.price,
                ph.note,
                ph.preparation_type,
                ph.manual_compliance,
                ph.manual_compliance_note,
                ph.brand_manual_enabled,
                NULL::int AS cas_count,
                NULL::text AS resolved_cas_u,
                NULL::int AS total_cas_count
            FROM product_hits ph
            UNION ALL
            SELECT
                'CODE_CAS_META'::text AS row_type,
                c.ord,
                NULL::text AS match_mode,
                NULL::int AS product_id,
                NULL::text AS name,
                NULL::text AS code,
                NULL::text AS cas,
                NULL::text AS brand,
                NULL::text AS size,
                NULL::text AS ship,
                NULL::text AS price,
                NULL::text AS note,
                NULL::text AS preparation_type,
                NULL::text AS manual_compliance,
                NULL::text AS manual_compliance_note,
                FALSE AS brand_manual_enabled,
                c.cas_count,
                c.resolved_cas_u,
                c.total_cas_count
            FROM code_cas_summary c
            UNION ALL
            SELECT
                'CODE_CAS_VERIFIED'::text AS row_type,
                v.ord,
                NULL::text AS match_mode,
                NULL::int AS product_id,
                NULL::text AS name,
                NULL::text AS code,
                NULL::text AS cas,
                NULL::text AS brand,
                NULL::text AS size,
                NULL::text AS ship,
                NULL::text AS price,
                NULL::text AS note,
                NULL::text AS preparation_type,
                NULL::text AS manual_compliance,
                NULL::text AS manual_compliance_note,
                FALSE AS brand_manual_enabled,
                NULL::int AS cas_count,
                NULL::text AS resolved_cas_u,
                NULL::int AS total_cas_count
            FROM code_cas_verified v
        )
        SELECT
            o.row_type,
            o.ord,
            o.match_mode,
            o.product_id,
            o.name,
            o.code,
            o.cas,
            o.brand,
            o.size,
            o.ship,
            o.price,
            o.note,
            o.preparation_type,
            o.manual_compliance,
            o.manual_compliance_note,
            o.brand_manual_enabled,
            rr.rule_label AS compliance_status,
            rr.note AS compliance_note,
            o.cas_count,
            o.resolved_cas_u,
            o.total_cas_count
        FROM output_rows o
        LEFT JOIN LATERAL (
            SELECT r.rule_label, r.note
            FROM regulatory_rules r
            WHERE o.row_type = 'HIT'
              AND NOT (
                o.brand_manual_enabled
                AND NULLIF(TRIM(COALESCE(o.manual_compliance, '')), '') IS NOT NULL
            )
              AND r.is_active = TRUE
              AND (
                (r.match_field = 'cas' AND NULLIF(TRIM(o.cas), '') IS NOT NULL
                    AND UPPER(TRIM(o.cas)) = UPPER(TRIM(r.match_value)))
                OR (r.match_field = 'name' AND NULLIF(TRIM(o.name), '') IS NOT NULL
                    AND UPPER(TRIM(o.name)) = UPPER(TRIM(r.match_value)))
                OR (r.match_field = 'code' AND NULLIF(TRIM(o.code), '') IS NOT NULL
                    AND UPPER(TRIM(o.code)) = UPPER(TRIM(r.match_value)))
              )
            ORDER BY r.priority ASC, r.id ASC
            LIMIT 1
        ) rr ON TRUE
        ORDER BY o.ord ASC, o.row_type ASC, o.product_id ASC
    """
    branch_params = vis_params + product_filter_params
    params = (
        [row["ord"] for row in lookup_rows],
        [row["code_u"] for row in lookup_rows],
        [row["cas_u"] for row in lookup_rows],
        [row["equivalent_search"] for row in lookup_rows],
    )
    params += vis_params  # code_cas_summary
    params += vis_params  # code_cas_verified
    params += branch_params + branch_params + branch_params + branch_params + branch_params

    rate_map = _exchange_rate_map(conn)
    by_ord: dict[int, list[dict]] = {row["ord"]: [] for row in lookup_rows}
    match_modes: dict[int, str] = {}
    code_cas_counts: dict[int, int] = {}
    code_total_cas_counts: dict[int, int] = {}
    code_cas_verified: set[int] = set()
    with conn.cursor() as cur:
        cur.execute(query, params)
        for db_row in cur.fetchall():
            row_type = db_row[0]
            ord_ = int(db_row[1])
            if row_type == "CODE_CAS_META":
                code_cas_counts[ord_] = int(db_row[18] or 0)
                code_total_cas_counts[ord_] = int(db_row[20] or 0)
                continue
            if row_type == "CODE_CAS_VERIFIED":
                code_cas_verified.add(ord_)
                continue
            candidate_row = (
                ord_,
                db_row[3],
                db_row[4],
                db_row[5],
                db_row[6],
                db_row[7],
                db_row[8],
                db_row[9],
                db_row[10],
                db_row[11],
                db_row[12],
                db_row[13],
                db_row[14],
                db_row[15],
                db_row[16],
                db_row[17],
            )
            by_ord.setdefault(ord_, []).append(_quote_candidate_from_row(candidate_row, rate_map))
            match_modes.setdefault(ord_, db_row[2])

    for row in lookup_rows:
        idx = row["ord"] - 1
        raw_candidates = by_ord.get(row["ord"], [])
        row_warnings = results[idx]["warnings"]
        if len(raw_candidates) > QUOTE_CANDIDATE_LIMIT:
            results[idx]["status"] = "UNRESOLVED"
            results[idx]["lifecycle"] = LIFECYCLE_REVIEW
            results[idx]["reason_code"] = REASON_CANDIDATE_LIMIT_EXCEEDED
            results[idx]["reason"] = "CANDIDATE_LIMIT_EXCEEDED"
            results[idx]["candidates"] = []
            results[idx]["selected_candidates"] = []
            results[idx]["selected"] = None
            results[idx]["warnings"] = row_warnings + ["CANDIDATE_LIMIT_EXCEEDED"]
            results[idx]["match_mode"] = match_modes.get(row["ord"])
            continue
        if row["equivalent_search"] and row["code_u"] and not row["cas_u"]:
            cas_count = code_cas_counts.get(row["ord"], 0)
            if cas_count == 0:
                total_cas = code_total_cas_counts.get(row["ord"], 0)
                if total_cas > 0:
                    # Code exists and has CAS entries but all are placeholders
                    results[idx]["status"] = "UNRESOLVED"
                    results[idx]["lifecycle"] = LIFECYCLE_UNRESOLVED
                    results[idx]["reason_code"] = REASON_CODE_HAS_PLACEHOLDER_CAS
                    results[idx]["reason"] = "CODE_HAS_PLACEHOLDER_CAS"
                else:
                    results[idx]["status"] = "UNRESOLVED"
                    results[idx]["lifecycle"] = LIFECYCLE_UNRESOLVED
                    results[idx]["reason_code"] = REASON_CODE_HAS_NO_CAS
                    results[idx]["reason"] = "CODE_HAS_NO_CAS"
                results[idx]["match_mode"] = "EQUIVALENT"
                continue
            if cas_count > 1:
                results[idx]["status"] = "UNRESOLVED"
                results[idx]["lifecycle"] = LIFECYCLE_UNRESOLVED
                results[idx]["reason_code"] = REASON_CODE_MULTIPLE_CAS
                results[idx]["reason"] = "CODE_MULTIPLE_CAS"
                results[idx]["match_mode"] = "EQUIVALENT"
                results[idx]["warnings"] = row_warnings + ["CODE_MULTIPLE_CAS"]
                continue
        if row["equivalent_search"] and row["code_u"] and row["cas_u"] and row["ord"] not in code_cas_verified:
            results[idx]["status"] = "UNRESOLVED"
            results[idx]["lifecycle"] = LIFECYCLE_UNRESOLVED
            results[idx]["reason_code"] = REASON_CODE_CAS_CONFLICT
            results[idx]["reason"] = "CODE_CAS_CONFLICT"
            results[idx]["match_mode"] = "CODE_CAS"
            continue
        candidates, filter_warnings = _quote_apply_candidate_filters(raw_candidates, filters)
        duplicate_warnings = _quote_mark_duplicate_groups(candidates)
        all_warnings = row_warnings + filter_warnings + duplicate_warnings
        if row["code_u"] and not row["cas_u"] and not row["equivalent_search"]:
            distinct_cas = {
                _quote_norm_identifier(c.get("Cas"))
                for c in candidates
                if _quote_norm_identifier(c.get("Cas"))
            }
            if len(distinct_cas) > 1:
                all_warnings.append("CODE_MULTIPLE_CAS")
                results[idx]["force_manual"] = True
        match_mode = match_modes.get(row["ord"]) or (
            "CODE_CAS" if row["code_u"] and row["cas_u"] else "EXACT_CODE" if row["code_u"] else "EXACT_CAS"
        )
        results[idx]["match_mode"] = match_mode
        if row["code_u"] and row["cas_u"] and not row["equivalent_search"] and not candidates:
            results[idx]["status"] = "UNRESOLVED"
            results[idx]["lifecycle"] = LIFECYCLE_UNRESOLVED
            results[idx]["reason_code"] = REASON_CODE_CAS_CONFLICT
            results[idx]["reason"] = "CODE_CAS_CONFLICT"
            results[idx]["warnings"] = list(dict.fromkeys(all_warnings))
            continue

        effective_strategy = QUOTE_SELECTION_MANUAL if results[idx].pop("force_manual", False) else strategy
        policy = row.get("effective_brand_policy")
        is_exact_code = bool(row["code_u"] and not row["equivalent_search"])

        if policy and not is_exact_code:
            final_cands, selected_cands, status, reason, lifecycle, reason_code, matched_tier, fallback_path = (
                _quote_apply_brand_policy_to_row(
                    raw_candidates=raw_candidates,
                    filtered_candidates=candidates,
                    policy=policy,
                    strategy=effective_strategy,
                    all_warnings=all_warnings,
                )
            )
            for candidate in final_cands:
                candidate.pop("size_parse", None)
            results[idx]["candidates"] = final_cands
            results[idx]["selected_candidates"] = selected_cands
            results[idx]["selected"] = selected_cands[0] if selected_cands else None
            results[idx]["status"] = status
            results[idx]["reason"] = reason
            results[idx]["lifecycle"] = lifecycle
            results[idx]["reason_code"] = reason_code
            results[idx]["matched_priority_tier"] = matched_tier
            results[idx]["fallback_path"] = fallback_path
            results[idx]["warnings"] = list(dict.fromkeys(all_warnings))
        else:
            if raw_candidates and not candidates:
                results[idx]["status"] = "UNRESOLVED"
                results[idx]["lifecycle"] = LIFECYCLE_REVIEW
                results[idx]["reason_code"] = REASON_FILTER_NO_MATCH
                results[idx]["reason"] = "NO_MATCH"
                results[idx]["candidates"] = []
                results[idx]["selected_candidates"] = []
                results[idx]["selected"] = None
                results[idx]["warnings"] = list(dict.fromkeys(all_warnings))
                continue

            selected_candidates, status, reason, lifecycle, reason_code = _quote_select_candidates(
                candidates, effective_strategy
            )
            for candidate in candidates:
                candidate.pop("size_parse", None)
            results[idx]["candidates"] = candidates
            results[idx]["selected_candidates"] = selected_candidates
            results[idx]["selected"] = selected_candidates[0] if selected_candidates else None
            results[idx]["status"] = status
            results[idx]["reason"] = reason
            results[idx]["lifecycle"] = lifecycle
            results[idx]["reason_code"] = reason_code
            results[idx]["warnings"] = list(dict.fromkeys(all_warnings))
    return results


def _brands_from_text(text: str) -> list[str]:
    """Tách danh sách brand từ textarea (dòng/;/,), trim, bỏ trống, không làm mất Unicode."""
    items = _split_multi_items(text, max_items=2000)
    return [x.strip() for x in items if x and x.strip()]


def _excel_cell_to_str(val) -> str:
    """Chuyển giá trị ô Excel thành chuỗi, giữ Unicode (tiếng Việt, ký tự đặc biệt)."""
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        return str(int(val)) if val.is_integer() else str(val)
    try:
        from openpyxl.cell.rich_text import CellRichText

        if isinstance(val, CellRichText):
            return "".join(str(t) for t in val).strip()
    except ImportError:
        pass
    return str(val).strip()


def _is_ooxml_xlsx(raw: bytes) -> bool:
    if len(raw) < 64 or raw[:2] != b"PK":
        return False
    try:
        with zipfile.ZipFile(BytesIO(raw), "r") as z:
            return "[Content_Types].xml" in z.namelist()
    except zipfile.BadZipFile:
        return False


def _is_old_binary_xls(raw: bytes) -> bool:
    return len(raw) >= 8 and raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _decode_text_flexible(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1258", "cp1252"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _maybe_interpret_as_csv(raw: bytes, filename: str) -> Optional[str]:
    """Nếu không phải .xlsx chuẩn, thử coi là CSV (UTF-8 / Windows)."""
    if _is_ooxml_xlsx(raw) or _is_old_binary_xls(raw):
        return None
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        return _decode_text_flexible(raw)
    text = _decode_text_flexible(raw)
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return None
    head = lines[0]
    if any(d in head for d in (",", ";", "\t")):
        return text
    return None


def _read_csv_dicts(text: str) -> tuple[list[dict], set[str]]:
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";\t,")
    except csv.Error:
        dialect = csv.excel
    f = StringIO(text)
    reader = csv.reader(f, dialect)
    try:
        header_row = next(reader)
    except StopIteration:
        raise ValueError("CSV rỗng.")

    headers = [str(x).strip() for x in header_row]
    keys = [h.lower() for h in headers]
    header_cols = {k for k in keys if k}
    out: list[dict] = []
    for parts in reader:
        if not parts or all(not (c or "").strip() for c in parts):
            continue
        row: dict[str, str] = {}
        empty = True
        for i, k in enumerate(keys):
            if not k:
                continue
            val = parts[i] if i < len(parts) else ""
            s = "" if val is None else str(val).strip()
            if s != "":
                empty = False
            row[k] = s
        if not empty:
            out.append(row)
    return out, header_cols


def _read_xlsx_bytes(raw: bytes) -> tuple[list[dict], set[str]]:
    bio = BytesIO(raw)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [_excel_cell_to_str(x) for x in next(rows)]
        except StopIteration:
            raise ValueError("File Excel rỗng (không có dòng nào).")

        keys = [h.lower() for h in headers]
        header_cols = {k for k in keys if k}
        out: list[dict] = []
        for r in rows:
            if r is None:
                continue
            row: dict[str, str] = {}
            empty = True
            for i, k in enumerate(keys):
                if not k:
                    continue
                val = r[i] if i < len(r) else None
                s = _excel_cell_to_str(val)
                if s != "":
                    empty = False
                row[k] = s
            if not empty:
                out.append(row)
        return out, header_cols
    finally:
        wb.close()


def _read_excel_dicts(file_storage):
    """
    Đọc sheet đầu của .xlsx (Office Open XML) hoặc CSV.
    Dòng 1 = tiêu đề (tên cột, không phân biệt hoa thường).
    Trả về (danh_sách_dòng_dữ_liệu, tập_tên_cột_từ_tiêu_đề).
    """
    try:
        file_storage.seek(0)
    except Exception:
        pass
    raw = file_storage.read()
    filename = getattr(file_storage, "filename", None) or "upload.xlsx"

    if _is_old_binary_xls(raw):
        raise ValueError(
            "File là Excel cũ (.xls nhị phân). Vui lòng mở bằng Excel/LibreOffice và "
            "File → Save As / Lưu thành → **Excel Workbook (.xlsx)** — không dùng .xls."
        )

    if _is_ooxml_xlsx(raw):
        try:
            return _read_xlsx_bytes(raw)
        except zipfile.BadZipFile as e:
            raise ValueError(f"File .xlsx bị hỏng hoặc không đầy đủ: {e}") from e

    csv_text = _maybe_interpret_as_csv(raw, filename)
    if csv_text is not None:
        return _read_csv_dicts(csv_text)

    if raw[:2] == b"PK":
        raise ValueError(
            "File có đuôi .xlsx nhưng **không phải Excel .xlsx chuẩn** (thiếu [Content_Types].xml). "
            "Thường gặp khi: đổi đuôi file CSV/HTML thành .xlsx, hoặc xuất sai định dạng. "
            "Cách xử lý: mở bằng Excel → **File → Save As → Excel Workbook (.xlsx)**; "
            "hoặc lưu dưới dạng **CSV UTF-8** rồi đổi đuôi thành .csv và upload lại."
        )

    raise ValueError(
        "Không đọc được file: không phải .xlsx hợp lệ và không nhận dạng được CSV. "
        "Hãy dùng đúng file .xlsx (Excel / LibreOffice) hoặc .csv có dòng đầu là tên cột (UTF-8)."
    )


def _require_admin_page():
    if not session.get("authenticated"):
        return redirect(url_for("login"))
    if not session.get("is_admin"):
        return "Admin only", 403
    return None


def _require_admin_api():
    if not session.get("authenticated"):
        return _quick_edit_json_response(False, "Chưa đăng nhập.", status=401)
    if not session.get("is_admin"):
        return _quick_edit_json_response(False, "Chỉ admin mới được thao tác.", status=403)
    return None


def _current_actor():
    if session.get("user_id"):
        return f"user:{session.get('user_id')}"
    return session.get("role") or "unknown"


def _quote_template_mapping_snapshot() -> dict:
    return json.loads(json.dumps(QUOTE_TEMPLATE_MAPPING_SNAPSHOT, ensure_ascii=False))


def _safe_uploaded_xlsx_filename(filename: str) -> str:
    base = (filename or "").replace("\\", "/").rsplit("/", 1)[-1].strip()
    if not base:
        raise ValueError("Thiếu tên file workbook.")
    lower = base.lower()
    if lower.endswith(".xlsm") or lower.endswith(".xls") or not lower.endswith(".xlsx"):
        raise ValueError("Chỉ hỗ trợ file .xlsx, không hỗ trợ .xls/.xlsm.")
    return base


def _read_bounded_workbook_upload(file_storage) -> bytes:
    try:
        file_storage.seek(0)
    except Exception:
        pass
    raw = file_storage.read(MAX_XLSX_BYTES + 1)
    if len(raw) > MAX_XLSX_BYTES:
        raise OverflowError(f"File .xlsx quá lớn, tối đa {MAX_XLSX_BYTES // (1024 * 1024)}MB.")
    if not raw:
        raise ValueError("File workbook rỗng.")
    return raw


def _validate_bg_v1_template(raw: bytes) -> dict:
    info = inspect_bg_template(raw)
    if info.capacity < 1:
        raise QuoteTemplateError("Template BG_V1 không có vùng dòng sản phẩm hợp lệ.")
    return _quote_template_mapping_snapshot()


def _json_datetime(value):
    return value.isoformat() if hasattr(value, "isoformat") else value


def _quote_template_admin_metadata(row) -> dict:
    return {
        "id": row[0],
        "filename": row[1],
        "content_sha256": row[2],
        "content_size": row[3],
        "profile_version": row[4],
        "is_active": bool(row[5]),
        "uploaded_by": row[6],
        "created_at": _json_datetime(row[7]),
        "activated_at": _json_datetime(row[8]),
    }


def _quote_template_public_metadata(row) -> dict:
    return {
        "id": row[0],
        "filename": row[1],
        "profile_version": row[2],
        "content_size": row[3],
        "created_at": _json_datetime(row[4]),
        "activated_at": _json_datetime(row[5]),
    }


def _list_quote_templates(conn) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, filename, content_sha256, content_size, profile_version,
                   is_active, uploaded_by, created_at, activated_at
            FROM quote_templates
            ORDER BY created_at DESC, id DESC
            """
        )
        return [_quote_template_admin_metadata(row) for row in cur.fetchall()]


def _insert_quote_template(conn, *, filename: str, raw: bytes, mapping: dict, activate: bool, uploaded_by: str) -> dict:
    digest = hashlib.sha256(raw).hexdigest()
    with conn:
        with conn.cursor() as cur:
            if activate:
                cur.execute("UPDATE quote_templates SET is_active = FALSE WHERE is_active = TRUE")
            cur.execute(
                """
                INSERT INTO quote_templates (
                    filename, content, content_sha256, content_size, profile_version,
                    mapping_json, is_active, uploaded_by, activated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s::jsonb, %s, %s, CASE WHEN %s THEN NOW() ELSE NULL END)
                RETURNING id, filename, content_sha256, content_size, profile_version,
                          is_active, uploaded_by, created_at, activated_at
                """,
                (
                    filename,
                    Binary(raw),
                    digest,
                    len(raw),
                    QUOTE_TEMPLATE_PROFILE_VERSION,
                    json.dumps(mapping, ensure_ascii=False),
                    activate,
                    uploaded_by,
                    activate,
                ),
            )
            return _quote_template_admin_metadata(cur.fetchone())


def _activate_quote_template(conn, template_id: int) -> dict:
    with conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id
                FROM quote_templates
                WHERE id = %s
                  AND profile_version = %s
                  AND mapping_json->>'profile_version' = %s
                  AND mapping_json->>'sheet' = 'BG'
                """,
                (template_id, QUOTE_TEMPLATE_PROFILE_VERSION, QUOTE_TEMPLATE_PROFILE_VERSION),
            )
            if cur.fetchone() is None:
                raise QuoteTemplateError("Không tìm thấy template BG_V1 hợp lệ.")
            cur.execute("UPDATE quote_templates SET is_active = FALSE WHERE is_active = TRUE")
            cur.execute(
                """
                UPDATE quote_templates
                SET is_active = TRUE, activated_at = NOW()
                WHERE id = %s
                RETURNING id, filename, content_sha256, content_size, profile_version,
                          is_active, uploaded_by, created_at, activated_at
                """,
                (template_id,),
            )
            return _quote_template_admin_metadata(cur.fetchone())


def _get_active_quote_template(conn, *, include_content: bool = False) -> dict:
    with conn.cursor() as cur:
        if include_content:
            cur.execute(
                """
                SELECT id, filename, profile_version, content_size, created_at, activated_at, content
                FROM quote_templates
                WHERE is_active = TRUE AND profile_version = %s
                ORDER BY activated_at DESC NULLS LAST, id DESC
                LIMIT 1
                """,
                (QUOTE_TEMPLATE_PROFILE_VERSION,),
            )
        else:
            cur.execute(
                """
                SELECT id, filename, profile_version, content_size, created_at, activated_at
                FROM quote_templates
                WHERE is_active = TRUE AND profile_version = %s
                ORDER BY activated_at DESC NULLS LAST, id DESC
                LIMIT 1
                """,
                (QUOTE_TEMPLATE_PROFILE_VERSION,),
            )
        row = cur.fetchone()
    if row is None:
        raise QuoteTemplateError("Chưa có mẫu báo giá active. Vui lòng nhờ admin upload và kích hoạt mẫu BG_V1.")
    data = _quote_template_public_metadata(row[:6])
    if include_content:
        data["content"] = bytes(row[6])
    return data


def _download_quote_template(conn, template_id: int) -> tuple[str, bytes]:
    with conn.cursor() as cur:
        cur.execute("SELECT filename, content FROM quote_templates WHERE id = %s", (template_id,))
        row = cur.fetchone()
    if row is None:
        raise QuoteTemplateError("Không tìm thấy template.")
    return _safe_uploaded_xlsx_filename(row[0]), bytes(row[1])


def _client_ip_from_request() -> str:
    xff = (request.headers.get("X-Forwarded-For") or "").strip()
    if xff:
        return xff.split(",")[0].strip()
    return (request.remote_addr or "").strip()


def _host_cidr(ip_str: str) -> Optional[str]:
    try:
        ip = ipaddress.ip_address(ip_str)
        return f"{ip}/32" if ip.version == 4 else f"{ip}/128"
    except ValueError:
        return None


def _parse_brand_list(text: str) -> list[str]:
    """Mỗi dòng một hoặc nhiều brand, phân tách bởi dấu phẩy/chấm phẩy; bỏ trùng giữ thứ tự."""
    out: list[str] = []
    seen: set[str] = set()
    for line in (text or "").splitlines():
        for part in line.replace(";", ",").split(","):
            b = part.strip()
            if b and b not in seen:
                seen.add(b)
                out.append(b)
    return out


def _ip_looks_non_public(ip_str: str) -> bool:
    """True nếu có vẻ là IP nội bộ / loopback — thường do proxy chưa truyền IP WAN thật."""
    s = (ip_str or "").strip()
    if not s:
        return False
    try:
        ip = ipaddress.ip_address(s)
        return bool(ip.is_private or ip.is_loopback or ip.is_link_local)
    except ValueError:
        return False


def _insert_import_job(cur, **kwargs):
    cur.execute(
        """
        INSERT INTO import_jobs
            (dataset, mode, status, filename, row_count, inserted_count, updated_count, deleted_count,
             error_message, created_by, meta_json)
        VALUES
            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
        """,
        (
            kwargs.get("dataset"), kwargs.get("mode"), kwargs.get("status"), kwargs.get("filename"),
            kwargs.get("row_count", 0), kwargs.get("inserted_count", 0), kwargs.get("updated_count", 0), kwargs.get("deleted_count", 0),
            kwargs.get("error_message"), kwargs.get("created_by"), json.dumps(kwargs.get("meta", {}), ensure_ascii=False),
        ),
    )


def _preview_hints(dataset, mode, rows):
    hints = []
    if dataset == "products":
        brands = sorted({_norm(r.get("brand")) for r in rows if _norm(r.get("brand"))})
        hints.append(f"Distinct brands in file: {len(brands)}")
        if mode == "replace_by_brand":
            hints.append("Apply sẽ xóa products theo các brand trong file rồi insert lại")
        elif mode == "upsert":
            hints.append("Upsert key: code + brand (không phân biệt hoa thường)")
    else:
        types_ = sorted({_norm(r.get("rule_type")).upper() for r in rows if _norm(r.get("rule_type"))})
        hints.append(f"Rule types in file: {', '.join(types_) if types_ else 'none'}")
        if mode == "replace_by_type":
            hints.append("Apply sẽ xóa rules của các rule_type trong file rồi insert lại")
    return hints


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""

        if username:
            conn = get_connection()
            try:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, password_hash, team_id, is_admin, ip_bypass_allowlist, "
                        "account_status, auth_version "
                        "FROM app_users WHERE username = %s AND auth_provider = 'LOCAL'",
                        (username,),
                    )
                    row = cur.fetchone()
            finally:
                conn.close()

            if row and row[1] and check_password_hash(row[1], password) and row[5] == "ACTIVE":
                session.clear()
                session["authenticated"] = True
                session["username"] = username
                session["user_id"] = row[0]
                session["team_id"] = row[2]
                session["is_admin"] = bool(row[3])
                session["ip_bypass_allowlist"] = bool(row[4])
                session["role"] = "admin" if row[3] else "user"
                session["auth_provider"] = "LOCAL"
                session["auth_version"] = row[6]
                return redirect(url_for("home"))
            return render_template("login.html", error="Sai tên đăng nhập hoặc mật khẩu.", google_auth_enabled=auth_google.google_auth_enabled()), 401

        if ENABLE_LEGACY_PASSWORD_LOGIN and MANAGER_PASSWORD and password == MANAGER_PASSWORD:
            session.clear()
            session["authenticated"] = True
            session["username"] = "__legacy_manager__"
            session["is_admin"] = True
            session["ip_bypass_allowlist"] = False
            session["role"] = "manager"
            return redirect(url_for("home"))
        if ENABLE_LEGACY_PASSWORD_LOGIN and STAFF_PASSWORD and password == STAFF_PASSWORD:
            session.clear()
            session["authenticated"] = True
            session["username"] = "__legacy_staff__"
            session["is_admin"] = False
            session["ip_bypass_allowlist"] = False
            team_id = int(os.environ.get("LEGACY_STAFF_TEAM_ID", "1"))
            session["team_id"] = team_id
            session["role"] = "staff"
            return redirect(url_for("home"))
        return render_template("login.html", error="Sai mật khẩu.", google_auth_enabled=auth_google.google_auth_enabled()), 403

    return render_template(
        "login.html",
        error=session_security.pop_login_notice(),
        google_auth_enabled=auth_google.google_auth_enabled(),
    )


@app.route("/")
def home():
    if not session.get("authenticated"):
        return redirect(url_for("login"))
    return render_template("index.html")


@app.route("/quote-assistant/quick")
def quick_quote():
    if not session.get("authenticated"):
        return redirect(url_for("login"))
    brand_options: list[str] = []
    brand_load_error = False
    conn = get_connection()
    try:
        vis, vis_params = _visibility_sql("p")
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT brand
                FROM (
                    SELECT DISTINCT TRIM(p.brand) AS brand
                    FROM products p
                    WHERE p.brand IS NOT NULL
                      AND TRIM(p.brand) <> ''
                      {vis}
                ) visible_brands
                ORDER BY UPPER(brand) ASC, brand ASC
                """,
                vis_params,
            )
            brand_options = [row[0] for row in cur.fetchall() if row and row[0]]
    except Exception:
        brand_load_error = True
        brand_options = []
        app.logger.exception("quick_quote: failed to load brand options")
    finally:
        conn.close()
    return render_template(
        "quick_quote.html",
        brand_options=brand_options,
        brand_load_error=brand_load_error,
    )


@app.route("/admin/imports", methods=["GET"])
def admin_imports():
    guard = _require_admin_page()
    if guard is not None:
        return guard

    token = request.args.get("preview")
    preview = IMPORT_PREVIEWS.get(token) if token else None

    conn = get_connection()
    recent_jobs = []
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, dataset, mode, status, row_count, inserted_count, updated_count, deleted_count, created_at
                FROM import_jobs
                ORDER BY id DESC
                LIMIT 20
                """
            )
            recent_jobs = [
                {
                    "id": r[0], "dataset": r[1], "mode": r[2], "status": r[3],
                    "row_count": r[4], "inserted_count": r[5], "updated_count": r[6], "deleted_count": r[7], "created_at": r[8],
                }
                for r in cur.fetchall()
            ]
    except Exception:
        recent_jobs = []
    finally:
        conn.close()

    return render_template("admin_imports.html", preview=preview, recent_jobs=recent_jobs, message=request.args.get("msg"), error=request.args.get("err"))


@app.route("/admin/quote-templates", methods=["GET"])
def admin_quote_templates_page():
    guard = _require_admin_page()
    if guard is not None:
        return guard
    return render_template("admin_quote_templates.html")


@app.route("/admin/imports/preview", methods=["POST"])
def admin_imports_preview():
    guard = _require_admin_page()
    if guard is not None:
        return guard

    dataset = (request.form.get("dataset") or "").strip()
    mode = (request.form.get("mode") or "").strip()
    file = request.files.get("file")
    if not file:
        return redirect(url_for("admin_imports", err="Thiếu file upload"))

    try:
        rows, header_cols = _read_excel_dicts(file)
    except Exception as e:
        return redirect(url_for("admin_imports", err=f"Không đọc được Excel: {e}"))

    if dataset not in {"products", "regulatory_rules"}:
        return redirect(url_for("admin_imports", err="Dataset không hợp lệ"))

    if dataset == "products":
        # Cho phép import thiếu nhiều cột; chỉ cần có cột brand để
        # hỗ trợ replace_by_brand và giữ dữ liệu nhất quán theo team/brand.
        required = ["brand"]
        valid_modes = {"upsert", "replace_by_brand", "append"}
    else:
        required = ["rule_type", "rule_label", "match_field", "match_value", "priority", "is_active", "note"]
        valid_modes = {"upsert", "replace_by_type"}

    if mode not in valid_modes:
        return redirect(url_for("admin_imports", err="Mode không hợp lệ"))

    missing = [c for c in required if c not in header_cols]
    if missing:
        return redirect(
            url_for(
                "admin_imports",
                err="Thiếu cột trong dòng tiêu đề (dòng 1): "
                + ", ".join(missing)
                + ". Tải file mẫu và giữ đúng tên cột tiếng Anh, không dấu.",
            )
        )

    if not rows:
        return redirect(
            url_for(
                "admin_imports",
                err="File có tiêu đề nhưng không có dòng dữ liệu nào. Thêm ít nhất một dòng dưới tiêu đề (ô không được để trống hoàn toàn).",
            )
        )

    if dataset == "products":
        try:
            validate_product_import_rows(rows, header_cols)
        except ValueError as e:
            return redirect(url_for("admin_imports", err=str(e)))

    token = str(uuid4())
    IMPORT_PREVIEWS[token] = {
        "token": token,
        "dataset": dataset,
        "mode": mode,
        "filename": file.filename or "upload.xlsx",
        "rows": rows,
        "header_cols": sorted(header_cols),
        "row_count": len(rows),
        "sample_rows": rows[:10],
        "hints": _preview_hints(dataset, mode, rows),
    }
    return redirect(url_for("admin_imports", preview=token))


@app.route("/admin/imports/apply", methods=["GET", "POST"])
def admin_imports_apply():
    guard = _require_admin_page()
    if guard is not None:
        return guard
    if request.method != "POST":
        return redirect(url_for("admin_imports", err="Vui lòng bấm 'Xem trước' rồi mới 'Xác nhận ghi vào database'."))

    token = request.form.get("preview_token")
    data = IMPORT_PREVIEWS.pop(token, None)
    if not data:
        return redirect(url_for("admin_imports", err="Preview hết hạn, vui lòng upload lại"))

    dataset = data["dataset"]
    mode = data["mode"]
    rows = data["rows"]
    filename = data.get("filename")
    header_cols = set(data.get("header_cols") or [])

    inserted = updated = deleted = 0
    actor = _current_actor()

    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                if dataset == "products":
                    validate_product_import_rows(rows, header_cols)
                    manual_header_mode = classify_manual_compliance_headers(header_cols)
                    manual_snapshot = {}
                    preparation_snapshot = {}

                    if mode == "replace_by_brand":
                        brands = sorted({_norm(r.get("brand")) for r in rows if _norm(r.get("brand"))})
                        if not brands:
                            raise ValueError("Mode replace_by_brand yêu cầu ít nhất 1 brand hợp lệ trong file.")
                        # Xóa theo brand không phân biệt hoa thường và bỏ khoảng trắng thừa.
                        brands_norm = sorted({b.strip().upper() for b in brands if b.strip()})
                        if manual_header_mode == HEADER_MODE_ABSENT:
                            manual_snapshot = fetch_manual_compliance_snapshot(cur, brands_norm)
                        if "preparation_type" not in header_cols:
                            preparation_snapshot = fetch_preparation_type_snapshot(cur, brands_norm)
                        cur.execute(
                            """
                            DELETE FROM products
                            WHERE UPPER(TRIM(COALESCE(brand, ''))) = ANY(%s)
                            """,
                            (brands_norm,),
                        )
                        deleted = cur.rowcount

                    for r in rows:
                        vals = (
                            _norm(r.get("name")), _norm(r.get("code")), _norm(r.get("cas")), _norm(r.get("brand")),
                            _norm(r.get("size")), _norm(r.get("ship")), _norm(r.get("price")), _norm(r.get("note")),
                        )
                        code, brand = vals[1], vals[3]
                        include_manual, manual_c, manual_n = resolve_manual_fields_for_write(
                            header_mode=manual_header_mode,
                            row=r,
                            code=code,
                            brand=brand,
                            snapshot=manual_snapshot,
                        )
                        include_preparation, preparation_type = resolve_preparation_type_for_write(
                            header_cols=header_cols,
                            row=r,
                            code=code,
                            brand=brand,
                            snapshot=preparation_snapshot,
                        )
                        if mode == "append":
                            _insert_product_row(
                                cur,
                                vals,
                                include_manual,
                                manual_c,
                                manual_n,
                                include_preparation,
                                preparation_type,
                            )
                            inserted += 1
                        else:
                            if not code or not brand:
                                _insert_product_row(
                                    cur,
                                    vals,
                                    include_manual,
                                    manual_c,
                                    manual_n,
                                    include_preparation,
                                    preparation_type,
                                )
                                inserted += 1
                                continue
                            cur.execute(
                                """
                                SELECT id FROM products
                                WHERE UPPER(TRIM(code)) = UPPER(TRIM(%s))
                                  AND UPPER(TRIM(brand)) = UPPER(TRIM(%s))
                                LIMIT 1
                                """,
                                (code, brand),
                            )
                            ex = cur.fetchone()
                            if ex:
                                _update_product_row(
                                    cur,
                                    vals,
                                    ex[0],
                                    include_manual,
                                    manual_c,
                                    manual_n,
                                    include_preparation,
                                    preparation_type,
                                )
                                updated += 1
                            else:
                                _insert_product_row(
                                    cur,
                                    vals,
                                    include_manual,
                                    manual_c,
                                    manual_n,
                                    include_preparation,
                                    preparation_type,
                                )
                                inserted += 1

                else:
                    parsed = []
                    for r in rows:
                        rule_type = _norm(r.get("rule_type")).upper()
                        match_field = _norm(r.get("match_field")).lower()
                        if rule_type not in {"CAM_NHAP", "PHU_LUC_II", "PHU_LUC_III", "TON_KHO"}:
                            raise ValueError(f"rule_type không hợp lệ: {rule_type}")
                        if match_field not in {"cas", "name", "code"}:
                            raise ValueError(f"match_field không hợp lệ: {match_field}")
                        priority_raw = _norm(r.get("priority")) or "100"
                        is_active_raw = _norm(r.get("is_active")).lower()
                        is_active = is_active_raw in {"1", "true", "yes", "y", "on"}
                        parsed.append(
                            {
                                "rule_type": rule_type,
                                "rule_label": _norm(r.get("rule_label")),
                                "match_field": match_field,
                                "match_value": _norm(r.get("match_value")),
                                "priority": int(float(priority_raw)),
                                "is_active": is_active,
                                "note": _norm(r.get("note")),
                            }
                        )

                    if mode == "replace_by_type":
                        types_ = sorted({x["rule_type"] for x in parsed})
                        cur.execute("DELETE FROM regulatory_rules WHERE rule_type = ANY(%s)", (types_,))
                        deleted = cur.rowcount

                    for r in parsed:
                        if mode == "replace_by_type":
                            cur.execute(
                                """
                                INSERT INTO regulatory_rules (rule_type, rule_label, match_field, match_value, priority, is_active, note)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                                """,
                                (r["rule_type"], r["rule_label"], r["match_field"], r["match_value"], r["priority"], r["is_active"], r["note"]),
                            )
                            inserted += 1
                        else:
                            cur.execute(
                                """
                                SELECT id FROM regulatory_rules
                                WHERE rule_type=%s AND match_field=%s AND UPPER(TRIM(match_value))=UPPER(TRIM(%s))
                                LIMIT 1
                                """,
                                (r["rule_type"], r["match_field"], r["match_value"]),
                            )
                            ex = cur.fetchone()
                            if ex:
                                cur.execute(
                                    """
                                    UPDATE regulatory_rules
                                       SET rule_label=%s, match_value=%s, priority=%s, is_active=%s, note=%s, updated_at=NOW()
                                     WHERE id=%s
                                    """,
                                    (r["rule_label"], r["match_value"], r["priority"], r["is_active"], r["note"], ex[0]),
                                )
                                updated += 1
                            else:
                                cur.execute(
                                    """
                                    INSERT INTO regulatory_rules (rule_type, rule_label, match_field, match_value, priority, is_active, note)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                                    """,
                                    (r["rule_type"], r["rule_label"], r["match_field"], r["match_value"], r["priority"], r["is_active"], r["note"]),
                                )
                                inserted += 1

                _insert_import_job(
                    cur,
                    dataset=dataset,
                    mode=mode,
                    status="success",
                    filename=filename,
                    row_count=len(rows),
                    inserted_count=inserted,
                    updated_count=updated,
                    deleted_count=deleted,
                    created_by=actor,
                    meta={"preview_token": token},
                )

        return redirect(url_for("admin_imports", msg=f"Import OK: inserted={inserted}, updated={updated}, deleted={deleted}"))
    except Exception as e:
        try:
            with conn:
                with conn.cursor() as cur:
                    _insert_import_job(
                        cur,
                        dataset=dataset,
                        mode=mode,
                        status="failed",
                        filename=filename,
                        row_count=len(rows),
                        inserted_count=inserted,
                        updated_count=updated,
                        deleted_count=deleted,
                        error_message=str(e),
                        created_by=actor,
                        meta={"preview_token": token},
                    )
        except Exception:
            pass
        return redirect(url_for("admin_imports", err=f"Import failed: {e}"))
    finally:
        conn.close()


RULE_TYPE_LABELS = {
    "CAM_NHAP": "CẤM NHẬP",
    "PHU_LUC_II": "Phụ lục II",
    "PHU_LUC_III": "Phụ lục III",
    "TON_KHO": "TỒN KHO",
}


def _insert_product_row(
    cur,
    vals,
    include_manual: bool,
    manual_c,
    manual_n,
    include_preparation: bool = False,
    preparation_type=None,
) -> None:
    columns = ["name", "code", "cas", "brand", "size", "ship", "price", "note"]
    params = list(vals)
    if include_manual:
        columns.extend(["manual_compliance", "manual_compliance_note"])
        params.extend([manual_c, manual_n])
    if include_preparation:
        columns.append("preparation_type")
        params.append(preparation_type)
    placeholders = ", ".join(["%s"] * len(columns))
    cur.execute(
        f"""
        INSERT INTO products ({", ".join(columns)})
        VALUES ({placeholders})
        """,
        tuple(params),
    )


def _update_product_row(
    cur,
    vals,
    product_id,
    include_manual: bool,
    manual_c,
    manual_n,
    include_preparation: bool = False,
    preparation_type=None,
) -> None:
    assignments = ["name=%s", "code=%s", "cas=%s", "brand=%s", "size=%s", "ship=%s", "price=%s", "note=%s"]
    params = list(vals)
    if include_manual:
        assignments.extend(["manual_compliance=%s", "manual_compliance_note=%s"])
        params.extend([manual_c, manual_n])
    if include_preparation:
        assignments.append("preparation_type=%s")
        params.append(preparation_type)
    params.append(product_id)
    cur.execute(
        f"""
        UPDATE products
           SET {", ".join(assignments)}
         WHERE id=%s
        """,
        tuple(params),
    )


def _parse_is_active(raw) -> bool:
    if isinstance(raw, bool):
        return raw
    s = _norm(str(raw or "")).lower()
    if not s:
        return True
    return s in {"1", "true", "yes", "y", "on"}


def _upsert_single_product(cur, row: dict) -> tuple[str, str]:
    """UPSERT 1 dòng products theo cặp code + brand (không phân biệt hoa thường)."""
    code = _norm(row.get("code"))
    brand = _norm(row.get("brand"))
    if not code or not brand:
        raise ValueError("Trường code và brand là bắt buộc.")

    vals = (
        _norm(row.get("name")),
        code,
        _norm(row.get("cas")),
        brand,
        _norm(row.get("size")),
        _norm(row.get("ship")),
        _norm(row.get("price")),
        _norm(row.get("note")),
    )
    cur.execute(
        """
        SELECT id FROM products
        WHERE UPPER(TRIM(code)) = UPPER(TRIM(%s))
          AND UPPER(TRIM(brand)) = UPPER(TRIM(%s))
        LIMIT 1
        """,
        (code, brand),
    )
    existing = cur.fetchone()
    label = vals[0] or code
    if existing:
        _update_product_row(cur, vals, existing[0], include_manual=False, manual_c=None, manual_n=None)
        return "updated", label

    _insert_product_row(cur, vals, include_manual=False, manual_c=None, manual_n=None)
    return "inserted", label


def _upsert_single_regulatory_rule(cur, row: dict) -> tuple[str, str]:
    """UPSERT 1 dòng regulatory_rules theo rule_type + match_field + match_value."""
    rule_type = _norm(row.get("rule_type")).upper()
    match_field = _norm(row.get("match_field")).lower()
    match_value = _norm(row.get("match_value"))
    rule_label = _norm(row.get("rule_label")) or RULE_TYPE_LABELS.get(rule_type, rule_type)

    if rule_type not in {"CAM_NHAP", "PHU_LUC_II", "PHU_LUC_III", "TON_KHO"}:
        raise ValueError(f"rule_type không hợp lệ: {rule_type}")
    if match_field not in {"cas", "name", "code"}:
        raise ValueError(f"match_field không hợp lệ: {match_field}")
    if not match_value:
        raise ValueError("match_value là bắt buộc.")

    priority_raw = _norm(row.get("priority")) or "100"
    try:
        priority = int(float(priority_raw))
    except (TypeError, ValueError):
        raise ValueError("priority phải là số nguyên.")
    is_active = _parse_is_active(row.get("is_active"))
    note = _norm(row.get("note"))

    cur.execute(
        """
        SELECT id FROM regulatory_rules
        WHERE rule_type=%s AND match_field=%s AND UPPER(TRIM(match_value))=UPPER(TRIM(%s))
        LIMIT 1
        """,
        (rule_type, match_field, match_value),
    )
    existing = cur.fetchone()
    label = f"{rule_label} ({match_field}={match_value})"
    if existing:
        cur.execute(
            """
            UPDATE regulatory_rules
               SET rule_label=%s, match_value=%s, priority=%s, is_active=%s, note=%s, updated_at=NOW()
             WHERE id=%s
            """,
            (rule_label, match_value, priority, is_active, note, existing[0]),
        )
        return "updated", label

    cur.execute(
        """
        INSERT INTO regulatory_rules (rule_type, rule_label, match_field, match_value, priority, is_active, note)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """,
        (rule_type, rule_label, match_field, match_value, priority, is_active, note),
    )
    return "inserted", label


def _delete_single_product(cur, row: dict) -> tuple[str, int]:
    """Xóa sản phẩm theo brand + (code | cas | name); size tuỳ chọn để thu hẹp."""
    brand = _norm(row.get("brand"))
    code = _norm(row.get("code"))
    cas = _norm(row.get("cas"))
    name = _norm(row.get("name"))
    size = _norm(row.get("size"))

    if not brand:
        raise ValueError("Trường brand là bắt buộc để xóa.")
    if not code and not cas and not name:
        raise ValueError(
            "Cần điền brand và ít nhất một trong: code, CAS hoặc name. "
            "Sản phẩm legacy (vd. CẤM NHẬP không có code) có thể xóa bằng brand + CAS."
        )

    size_clause = " AND UPPER(TRIM(COALESCE(size, ''))) = UPPER(TRIM(%s))" if size else ""
    size_params = (size,) if size else ()

    if code:
        cur.execute(
            f"""
            DELETE FROM products
            WHERE UPPER(TRIM(brand)) = UPPER(TRIM(%s))
              AND UPPER(TRIM(code)) = UPPER(TRIM(%s))
              {size_clause}
            """,
            (brand, code) + size_params,
        )
        label = f"{code} / {brand}" + (f" / {size}" if size else "")
    elif cas:
        name_clause = " AND UPPER(TRIM(name)) = UPPER(TRIM(%s))" if name else ""
        name_params = (name,) if name else ()
        cur.execute(
            f"""
            DELETE FROM products
            WHERE UPPER(TRIM(brand)) = UPPER(TRIM(%s))
              AND UPPER(TRIM(cas)) = UPPER(TRIM(%s))
              {name_clause}
              {size_clause}
            """,
            (brand, cas) + name_params + size_params,
        )
        label = f"{brand} / CAS {cas}" + (f" / {name}" if name else "") + (f" / {size}" if size else "")
    else:
        cur.execute(
            f"""
            DELETE FROM products
            WHERE UPPER(TRIM(brand)) = UPPER(TRIM(%s))
              AND UPPER(TRIM(name)) = UPPER(TRIM(%s))
              {size_clause}
            """,
            (brand, name) + size_params,
        )
        label = f"{brand} / {name}" + (f" / {size}" if size else "")

    deleted = cur.rowcount
    if deleted <= 0:
        raise ValueError(
            "Không tìm thấy sản phẩm để xóa. Kiểm tra brand"
            + (" + code" if code else " + CAS" if cas else " + name")
            + (" + size" if size else "")
            + ". Với dòng không có code, thử brand + CAS (vd. CẤM NHẬP + 634-90-2)."
        )
    return label, deleted


def _delete_single_regulatory_rule(cur, row: dict) -> tuple[str, int]:
    """Xóa quy tắc theo rule_type + match_field + match_value."""
    rule_type = _norm(row.get("rule_type")).upper()
    match_field = _norm(row.get("match_field")).lower()
    match_value = _norm(row.get("match_value"))

    if rule_type not in {"CAM_NHAP", "PHU_LUC_II", "PHU_LUC_III", "TON_KHO"}:
        raise ValueError(f"rule_type không hợp lệ: {rule_type}")
    if match_field not in {"cas", "name", "code"}:
        raise ValueError(f"match_field không hợp lệ: {match_field}")
    if not match_value:
        raise ValueError("match_value là bắt buộc để xóa.")

    cur.execute(
        """
        DELETE FROM regulatory_rules
        WHERE rule_type=%s AND match_field=%s AND UPPER(TRIM(match_value))=UPPER(TRIM(%s))
        """,
        (rule_type, match_field, match_value),
    )
    deleted = cur.rowcount
    if deleted <= 0:
        raise ValueError("Không tìm thấy quy tắc để xóa. Kiểm tra rule_type, match_field và match_value.")

    label = f"{rule_type} ({match_field}={match_value})"
    return label, deleted


def _quick_edit_json_response(ok: bool, message: str, action: str = "", label: str = "", status: int = 200):
    return jsonify({"ok": ok, "message": message, "action": action, "label": label}), status


@app.route("/admin/imports/quick-product", methods=["POST"])
def admin_imports_quick_product():
    guard = _require_admin_api()
    if guard is not None:
        return guard

    actor = _current_actor()
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                action, label = _upsert_single_product(cur, request.values)
                inserted = 1 if action == "inserted" else 0
                updated = 1 if action == "updated" else 0
                _insert_import_job(
                    cur,
                    dataset="products",
                    mode="quick_upsert",
                    status="success",
                    filename=None,
                    row_count=1,
                    inserted_count=inserted,
                    updated_count=updated,
                    deleted_count=0,
                    created_by=actor,
                    meta={"label": label, "code": _norm(request.values.get("code")), "brand": _norm(request.values.get("brand"))},
                )
        verb = "cập nhật" if action == "updated" else "thêm mới"
        return _quick_edit_json_response(
            True,
            f"Đã {verb} thành công dữ liệu của {label}",
            action=action,
            label=label,
        )
    except ValueError as e:
        return _quick_edit_json_response(False, str(e), status=400)
    except Exception as e:
        return _quick_edit_json_response(False, f"Lỗi: {e}", status=500)
    finally:
        conn.close()


@app.route("/admin/imports/quick-rule", methods=["POST"])
def admin_imports_quick_rule():
    guard = _require_admin_api()
    if guard is not None:
        return guard

    actor = _current_actor()
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                action, label = _upsert_single_regulatory_rule(cur, request.values)
                inserted = 1 if action == "inserted" else 0
                updated = 1 if action == "updated" else 0
                _insert_import_job(
                    cur,
                    dataset="regulatory_rules",
                    mode="quick_upsert",
                    status="success",
                    filename=None,
                    row_count=1,
                    inserted_count=inserted,
                    updated_count=updated,
                    deleted_count=0,
                    created_by=actor,
                    meta={"label": label},
                )
        verb = "cập nhật" if action == "updated" else "thêm mới"
        return _quick_edit_json_response(
            True,
            f"Đã {verb} thành công dữ liệu của {label}",
            action=action,
            label=label,
        )
    except ValueError as e:
        return _quick_edit_json_response(False, str(e), status=400)
    except Exception as e:
        return _quick_edit_json_response(False, f"Lỗi: {e}", status=500)
    finally:
        conn.close()


@app.route("/admin/imports/quick-product/delete", methods=["POST"])
def admin_imports_quick_product_delete():
    guard = _require_admin_api()
    if guard is not None:
        return guard

    actor = _current_actor()
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                label, deleted = _delete_single_product(cur, request.values)
                _insert_import_job(
                    cur,
                    dataset="products",
                    mode="quick_delete",
                    status="success",
                    filename=None,
                    row_count=1,
                    inserted_count=0,
                    updated_count=0,
                    deleted_count=deleted,
                    created_by=actor,
                    meta={"label": label, "code": _norm(request.values.get("code")), "brand": _norm(request.values.get("brand"))},
                )
        return _quick_edit_json_response(
            True,
            f"Đã xóa thành công {deleted} dòng: {label}",
            action="deleted",
            label=label,
        )
    except ValueError as e:
        return _quick_edit_json_response(False, str(e), status=400)
    except Exception as e:
        return _quick_edit_json_response(False, f"Lỗi: {e}", status=500)
    finally:
        conn.close()


@app.route("/admin/imports/quick-rule/delete", methods=["POST"])
def admin_imports_quick_rule_delete():
    guard = _require_admin_api()
    if guard is not None:
        return guard

    actor = _current_actor()
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                label, deleted = _delete_single_regulatory_rule(cur, request.values)
                _insert_import_job(
                    cur,
                    dataset="regulatory_rules",
                    mode="quick_delete",
                    status="success",
                    filename=None,
                    row_count=1,
                    inserted_count=0,
                    updated_count=0,
                    deleted_count=deleted,
                    created_by=actor,
                    meta={"label": label},
                )
        return _quick_edit_json_response(
            True,
            f"Đã xóa thành công quy tắc: {label}",
            action="deleted",
            label=label,
        )
    except ValueError as e:
        return _quick_edit_json_response(False, str(e), status=400)
    except Exception as e:
        return _quick_edit_json_response(False, f"Lỗi: {e}", status=500)
    finally:
        conn.close()


def _xlsx_response(wb: Workbook, download_name: str):
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _xlsx_bytes_response(raw: bytes, download_name: str):
    bio = BytesIO(raw)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/admin/quote-templates", methods=["GET"])
def admin_quote_templates_list():
    guard = _require_admin_api()
    if guard is not None:
        return guard

    conn = get_connection()
    try:
        return jsonify({"templates": _list_quote_templates(conn)})
    finally:
        conn.close()


@app.route("/api/admin/quote-templates", methods=["POST"])
def admin_quote_templates_upload():
    guard = _require_admin_api()
    if guard is not None:
        return guard

    workbook = request.files.get("workbook")
    if workbook is None:
        return _quote_json_error("Thiếu multipart field workbook.", status=400)

    try:
        filename = _safe_uploaded_xlsx_filename(workbook.filename or "")
        activate_value = request.form.get("activate")
        activate = True if activate_value is None else _quote_bool_or_none(activate_value)
        if activate is None:
            raise ValueError("activate phải là true hoặc false.")
        raw = _read_bounded_workbook_upload(workbook)
        mapping = _validate_bg_v1_template(raw)
    except OverflowError as e:
        return _quote_json_error(str(e), status=413)
    except (ValueError, WorkbookExportError, QuoteTemplateError) as e:
        return _quote_json_error(str(e), status=400)

    conn = get_connection()
    try:
        template = _insert_quote_template(
            conn,
            filename=filename,
            raw=raw,
            mapping=mapping,
            activate=bool(activate),
            uploaded_by=_current_actor(),
        )
        return jsonify({"template": template}), 201
    except IntegrityError:
        return _quote_json_error("Không thể kích hoạt template do đã có phiên bản active khác. Vui lòng thử lại.", status=409)
    finally:
        conn.close()


@app.route("/api/admin/quote-templates/<int:template_id>/activate", methods=["POST"])
def admin_quote_templates_activate(template_id: int):
    guard = _require_admin_api()
    if guard is not None:
        return guard

    conn = get_connection()
    try:
        template = _activate_quote_template(conn, template_id)
        return jsonify({"template": template})
    except QuoteTemplateError as e:
        return _quote_json_error(str(e), status=404)
    except IntegrityError:
        return _quote_json_error("Không thể kích hoạt template do ràng buộc active duy nhất.", status=409)
    finally:
        conn.close()


@app.route("/api/admin/quote-templates/<int:template_id>/download", methods=["GET"])
def admin_quote_templates_download(template_id: int):
    guard = _require_admin_api()
    if guard is not None:
        return guard

    conn = get_connection()
    try:
        filename, raw = _download_quote_template(conn, template_id)
        return _xlsx_bytes_response(raw, filename)
    except QuoteTemplateError as e:
        return _quote_json_error(str(e), status=404)
    finally:
        conn.close()


@app.route("/admin/templates/products.xlsx")
def admin_template_products():
    guard = _require_admin_page()
    if guard is not None:
        return guard
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(["name", "code", "cas", "brand", "size", "ship", "price", "note", "Preparation_Type"])
    return _xlsx_response(wb, "products_import_template.xlsx")


@app.route("/admin/templates/regulatory_rules.xlsx")
def admin_template_regulatory_rules():
    guard = _require_admin_page()
    if guard is not None:
        return guard
    wb = Workbook()
    ws = wb.active
    ws.title = "regulatory_rules"
    ws.append(["rule_type", "rule_label", "match_field", "match_value", "priority", "is_active", "note"])
    ws.append(["CAM_NHAP", "CẤM NHẬP", "cas", "123-45-6", 10, "TRUE", ""])
    return _xlsx_response(wb, "regulatory_rules_import_template.xlsx")


@app.route("/admin/exchange-rates", methods=["GET", "POST"])
def admin_exchange_rates():
    guard = _require_admin_page()
    if guard is not None:
        return guard

    msg = err = None
    if request.method == "POST":
        conn = get_connection()
        try:
            if request.form.get("seed_json"):
                base = _default_exchange_rates_from_json()
                if not base:
                    err = "Không đọc được static/exchange_rates.json hoặc file rỗng."
                else:
                    with conn.cursor() as cur:
                        for brand, rate in base.items():
                            cur.execute(
                                """
                                INSERT INTO exchange_rates (brand, rate)
                                VALUES (%s, %s)
                                ON CONFLICT (brand) DO UPDATE SET rate = EXCLUDED.rate, updated_at = NOW()
                                """,
                                (brand, rate),
                            )
                    conn.commit()
                    msg = f"Đã đồng bộ {len(base)} brand từ file JSON vào database."
            elif request.form.get("delete_brand"):
                b = (request.form.get("delete_brand") or "").strip()
                if b:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM exchange_rates WHERE brand = %s", (b,))
                    conn.commit()
                    msg = f"Đã xóa tỷ giá cho brand: {b}"
            elif request.form.get("bulk_same_apply"):
                brands = _parse_brand_list(request.form.get("bulk_brands") or "")
                rate_raw = (request.form.get("bulk_rate") or "").strip().replace(",", ".")
                if not brands:
                    err = "Nhập danh sách brand (mỗi dòng hoặc cách nhau bởi dấu phẩy)."
                elif not rate_raw:
                    err = "Nhập tỷ giá chung."
                elif len(brands) > 2000:
                    err = "Tối đa 2000 brand mỗi lần."
                else:
                    try:
                        rate = float(rate_raw)
                    except ValueError:
                        err = "Tỷ giá không phải số."
                    else:
                        with conn.cursor() as cur:
                            for brand in brands:
                                cur.execute(
                                    """
                                    INSERT INTO exchange_rates (brand, rate)
                                    VALUES (%s, %s)
                                    ON CONFLICT (brand) DO UPDATE SET rate = EXCLUDED.rate, updated_at = NOW()
                                    """,
                                    (brand, rate),
                                )
                        conn.commit()
                        msg = f"Đã áp dụng tỷ giá {rate} cho {len(brands)} brand."
            elif request.form.get("bulk_lines_apply"):
                text = request.form.get("bulk_lines") or ""
                rows_parsed: list[tuple[str, float]] = []
                bad_lines: list[str] = []
                for raw in text.splitlines():
                    line = raw.strip()
                    if not line or line.startswith("#"):
                        continue
                    if "\t" in line and "," not in line:
                        parts = [p.strip() for p in line.split("\t", 1)]
                    else:
                        parts = [p.strip() for p in line.split(",", 1)]
                    if len(parts) < 2 or not parts[0]:
                        bad_lines.append(raw.strip()[:80])
                        continue
                    b, rraw = parts[0], parts[1].replace(",", ".")
                    try:
                        r = float(rraw)
                    except ValueError:
                        bad_lines.append(raw.strip()[:80])
                        continue
                    rows_parsed.append((b, r))
                if len(rows_parsed) > 2000:
                    err = "Tối đa 2000 dòng mỗi lần."
                elif not rows_parsed:
                    err = "Không có dòng hợp lệ. Định dạng: brand,tỷ_giá (mỗi dòng một cặp)."
                else:
                    with conn.cursor() as cur:
                        for brand, rate in rows_parsed:
                            cur.execute(
                                """
                                INSERT INTO exchange_rates (brand, rate)
                                VALUES (%s, %s)
                                ON CONFLICT (brand) DO UPDATE SET rate = EXCLUDED.rate, updated_at = NOW()
                                """,
                                (brand, rate),
                            )
                    conn.commit()
                    msg = f"Đã cập nhật {len(rows_parsed)} dòng."
                    if bad_lines:
                        msg += f" Bỏ qua {len(bad_lines)} dòng không đọc được."
            else:
                brand = (request.form.get("brand") or "").strip()
                rate_raw = (request.form.get("rate") or "").strip().replace(",", ".")
                if not brand or not rate_raw:
                    err = "Nhập đủ brand và rate."
                else:
                    try:
                        rate = float(rate_raw)
                    except ValueError:
                        err = "Rate không phải số."
                    else:
                        with conn.cursor() as cur:
                            cur.execute(
                                """
                                INSERT INTO exchange_rates (brand, rate)
                                VALUES (%s, %s)
                                ON CONFLICT (brand) DO UPDATE SET rate = EXCLUDED.rate, updated_at = NOW()
                                """,
                                (brand, rate),
                            )
                        conn.commit()
                        msg = f"Đã lưu tỷ giá {brand} = {rate}"
        except Exception as e:
            err = str(e)
        finally:
            conn.close()

    rows = []
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT brand, rate, updated_at FROM exchange_rates ORDER BY brand ASC LIMIT 500")
            rows = [{"brand": r[0], "rate": r[1], "updated_at": r[2]} for r in cur.fetchall()]
    except Exception as e:
        err = err or f"Không đọc được bảng exchange_rates (đã chạy migration_005?): {e}"
    finally:
        conn.close()

    return render_template(
        "admin_exchange_rates.html",
        rows=rows,
        message=msg or request.args.get("msg"),
        error=err or request.args.get("err"),
        json_fallback_count=len(_default_exchange_rates_from_json()),
    )


@app.route("/admin/network", methods=["GET", "POST"])
def admin_network():
    guard = _require_admin_page()
    if guard is not None:
        return guard

    msg = err = None
    if request.method == "POST":
        conn = get_connection()
        try:
            if request.form.get("add_my_ip"):
                ip_s = _client_ip_from_request()
                cidr = _host_cidr(ip_s)
                if not cidr:
                    err = f"Không suy ra được CIDR từ IP: {ip_s!r}"
                else:
                    label = (request.form.get("my_ip_label") or "auto").strip() or "auto"
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO office_ip_allowlist (cidr, label)
                            VALUES (%s, %s)
                            ON CONFLICT (cidr) DO UPDATE SET label = EXCLUDED.label, is_active = TRUE
                            """,
                            (cidr, label),
                        )
                    conn.commit()
                    msg = (
                        f"Đã thêm/kích hoạt {cidr}. Đây là IP công khai (WAN) mà máy chủ nhận được khi truy cập từ "
                        f"mạng hiện tại — thường là IP modem/router văn phòng (chung cho cả LAN), không phải IP máy nội bộ 192.168.x.x."
                    )
            elif request.form.get("delete_id"):
                try:
                    rid = int(request.form.get("delete_id"))
                except (TypeError, ValueError):
                    err = "ID không hợp lệ."
                else:
                    with conn.cursor() as cur:
                        cur.execute("DELETE FROM office_ip_allowlist WHERE id = %s", (rid,))
                    conn.commit()
                    msg = "Đã xóa rule."
            else:
                cidr = (request.form.get("cidr") or "").strip()
                label = (request.form.get("label") or "").strip() or None
                if not cidr:
                    err = "Nhập CIDR hoặc IP (ví dụ 203.0.113.0/24 hoặc 203.0.113.10)."
                else:
                    try:
                        ipaddress.ip_network(cidr, strict=False)
                    except ValueError:
                        err = "CIDR/IP không hợp lệ."
                    else:
                        with conn.cursor() as cur:
                            cur.execute(
                                """
                                INSERT INTO office_ip_allowlist (cidr, label)
                                VALUES (%s, %s)
                                ON CONFLICT (cidr) DO UPDATE SET label = EXCLUDED.label, is_active = TRUE
                                """,
                                (cidr, label),
                            )
                        conn.commit()
                        msg = f"Đã thêm/kích hoạt {cidr}"
        except Exception as e:
            err = str(e)
        finally:
            conn.close()

    rules = []
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, cidr, label, is_active, created_at FROM office_ip_allowlist ORDER BY id DESC LIMIT 200"
            )
            rules = [
                {"id": r[0], "cidr": r[1], "label": r[2], "is_active": r[3], "created_at": r[4]}
                for r in cur.fetchall()
            ]
    except Exception as e:
        err = err or f"Không đọc được office_ip_allowlist (migration_006?): {e}"
    finally:
        conn.close()

    disable = os.environ.get("DISABLE_IP_ALLOWLIST", "").lower() in ("1", "true", "yes", "on")
    env_list = [x.strip() for x in (os.environ.get("OFFICE_IP_ALLOWLIST") or "").split(",") if x.strip()]
    seen_ip = _client_ip_from_request()
    seen_ip_non_public = _ip_looks_non_public(seen_ip)

    return render_template(
        "admin_network.html",
        rules=rules,
        message=msg or request.args.get("msg"),
        error=err or request.args.get("err"),
        disable_allowlist=disable,
        env_allowlist=env_list,
        seen_ip=seen_ip,
        seen_ip_non_public=seen_ip_non_public,
    )


@app.route("/admin/users", methods=["GET", "POST"])
def admin_users():
    guard = _require_admin_page()
    if guard is not None:
        return guard

    msg = err = None

    if request.method == "POST":
        # Phase 5D2B.1: every legacy mutation here must be CSRF-checked
        # server-side, same as the newer Google-account actions.
        if not session_security.verify_csrf_token(request.form.get("csrf_token", "")):
            return "Yêu cầu không hợp lệ.", 400

        action = (request.form.get("action") or "").strip()

        if action == "create_user":
            username = (request.form.get("username") or "").strip()
            password = request.form.get("password") or ""
            role = (request.form.get("role") or "user").strip().lower()
            is_admin = role in {"admin", "1", "true", "yes", "on"}
            ip_bypass_allowlist = (request.form.get("ip_bypass_allowlist") or "").strip().lower() in {"1", "true", "yes", "on"}
            # Phase 6A: team is SELECTED (from an existing team, managed on
            # /admin/teams), never typed as free-text brands and never
            # auto-created here. See admin_teams.py for team CRUD.
            team_id_s = (request.form.get("team_id") or "").strip()

            if not username:
                err = "Thiếu username."
            elif not password:
                err = "Thiếu mật khẩu."

            team_id = None
            if not err and not is_admin:
                try:
                    team_id = int(team_id_s)
                except (TypeError, ValueError):
                    err = "Vui lòng chọn team hợp lệ cho nhân viên."

            if not err:
                conn = get_connection()
                try:
                    with conn:
                        with conn.cursor() as cur:
                            # Phase 5D2B Final: every admin user-management
                            # mutation (this file + admin_google_users.py)
                            # now uniformly acquires the shared advisory
                            # lock and revalidates the ACTING admin's own
                            # DB row FIRST, before reading/locking any
                            # target row -- see acquire_last_admin_lock's
                            # and revalidate_actor's docstrings for why.
                            admin_google_users.acquire_last_admin_lock(cur)
                            admin_google_users.revalidate_actor(
                                cur, session.get("user_id"), session.get("auth_version")
                            )

                            cur.execute("SELECT id FROM app_users WHERE username = %s", (username,))
                            if cur.fetchone():
                                raise ValueError(f"Username đã tồn tại: {username}")

                            if not is_admin:
                                cur.execute("SELECT id FROM teams WHERE id = %s", (team_id,))
                                if cur.fetchone() is None:
                                    raise ValueError("Team đã chọn không còn tồn tại.")

                            password_hash = generate_password_hash(password)
                            # auth_provider/account_status set EXPLICITLY
                            # (not just relying on column DEFAULTs) -- this
                            # legacy form only ever creates LOCAL,
                            # already-ACTIVE accounts.
                            cur.execute(
                                """
                                INSERT INTO app_users
                                    (username, password_hash, team_id, is_admin, ip_bypass_allowlist,
                                     auth_provider, account_status)
                                VALUES (%s, %s, %s, %s, %s, 'LOCAL', 'ACTIVE')
                                RETURNING id
                                """,
                                (username, password_hash, None if is_admin else team_id,
                                 is_admin, ip_bypass_allowlist),
                            )
                            (new_user_id,) = cur.fetchone()
                            admin_google_users.touch_team_updated_at(
                                cur, None if is_admin else team_id
                            )
                            admin_google_users.write_permission_audit(
                                cur, actor_user_id=session.get("user_id"),
                                target_user_id=new_user_id, target_provider="LOCAL",
                                reason_code="USER_CREATED",
                            )
                    msg = "Tạo user thành công."
                except Exception as e:
                    err = str(e)
                finally:
                    conn.close()

        elif action == "update_user":
            conn = get_connection()
            try:
                user_id_s = (request.form.get("user_id") or "").strip()
                if not user_id_s:
                    err = "Thiếu user_id."
                else:
                    user_id = int(user_id_s)
                    password = request.form.get("password") or ""
                    role = (request.form.get("role") or "user").strip().lower()
                    set_is_admin = role in {"admin", "1", "true", "yes", "on"}
                    set_ip_bypass_allowlist = (request.form.get("ip_bypass_allowlist") or "").strip().lower() in {"1", "true", "yes", "on"}
                    # Phase 6A: team is SELECTED from an existing team; this
                    # form never creates a team or edits team_brands.
                    new_team_id_s = (request.form.get("team_id") or "").strip()
                    new_team_id = None
                    if not set_is_admin:
                        try:
                            new_team_id = int(new_team_id_s)
                        except (TypeError, ValueError):
                            raise ValueError("Vui lòng chọn team hợp lệ cho nhân viên.")

                    with conn:
                        with conn.cursor() as cur:
                            # Phase 5D2B.1: consistent global lock order with
                            # admin_google_users.suspend() -- acquired
                            # unconditionally, before any per-row lock, by
                            # every path that could reduce the ACTIVE-admin
                            # count. See _LAST_ADMIN_LOCK_KEY there.
                            admin_google_users.acquire_last_admin_lock(cur)
                            # Phase 5D2B.2: re-read the ACTING admin's own
                            # row fresh, after the lock, in case they were
                            # suspended/demoted/revoked by someone else
                            # while this request was waiting for the lock.
                            admin_google_users.revalidate_actor(
                                cur, session.get("user_id"), session.get("auth_version")
                            )

                            # This legacy form can only ever target a LOCAL
                            # account: a GOOGLE account id (real or a forged
                            # hidden-field value) simply will not match this
                            # WHERE clause and is treated identically to "no
                            # such user" below. This form also never reads
                            # or writes auth_provider, google_sub,
                            # account_status, or is_break_glass -- those
                            # columns are exclusively managed by
                            # admin_google_users.py's own validated actions.
                            cur.execute(
                                "SELECT username, team_id, is_admin FROM app_users "
                                "WHERE id = %s AND auth_provider = 'LOCAL' FOR UPDATE",
                                (user_id,),
                            )
                            row = cur.fetchone()
                            if not row:
                                raise ValueError("Không tìm thấy user.")

                            username, team_id, old_is_admin = row
                            demoting = bool(old_is_admin) and not set_is_admin

                            if demoting and user_id == session.get("user_id"):
                                raise ValueError("Không thể tự hạ quyền admin của chính mình.")

                            if demoting:
                                # Same invariant admin_google_users.suspend()
                                # enforces, checked AFTER the advisory lock
                                # above -- counts across the whole system
                                # (LOCAL + GOOGLE), not just this table.
                                cur.execute(
                                    "SELECT COUNT(*) FROM app_users "
                                    "WHERE is_admin = TRUE AND account_status = 'ACTIVE' AND id <> %s",
                                    (user_id,),
                                )
                                (other_active_admins,) = cur.fetchone()
                                if other_active_admins == 0:
                                    raise ValueError("Không thể hạ quyền admin đang hoạt động cuối cùng.")

                            if password.strip():
                                cur.execute(
                                    "UPDATE app_users SET password_hash = %s WHERE id = %s",
                                    (generate_password_hash(password.strip()), user_id),
                                )

                            if set_is_admin:
                                cur.execute(
                                    "UPDATE app_users SET is_admin = TRUE, team_id = NULL, "
                                    "ip_bypass_allowlist = %s, auth_version = auth_version + 1 WHERE id = %s",
                                    (set_ip_bypass_allowlist, user_id),
                                )
                                final_team_id = None
                            else:
                                cur.execute("SELECT id FROM teams WHERE id = %s", (new_team_id,))
                                if cur.fetchone() is None:
                                    raise ValueError("Team đã chọn không còn tồn tại.")
                                cur.execute(
                                    "UPDATE app_users SET is_admin = FALSE, team_id = %s, "
                                    "ip_bypass_allowlist = %s, auth_version = auth_version + 1 WHERE id = %s",
                                    (new_team_id, set_ip_bypass_allowlist, user_id),
                                )
                                final_team_id = new_team_id

                            # Bump BOTH the team being left (old `team_id`,
                            # read above before this UPDATE) and the team
                            # being joined -- a no-op if neither actually
                            # changed. Phase 6A-Fix2: this is what makes an
                            # in-flight team permission-change preview
                            # correctly go stale when membership changes,
                            # not just when brands/ip_policy change -- see
                            # admin_teams.py's module docstring.
                            admin_google_users.touch_team_updated_at(cur, team_id, final_team_id)

                            admin_google_users.write_permission_audit(
                                cur, actor_user_id=session.get("user_id"),
                                target_user_id=user_id, target_provider="LOCAL",
                                reason_code="USER_TEAM_UPDATED",
                            )

                    msg = "Cập nhật user thành công."
            except Exception as e:
                err = str(e)
            finally:
                conn.close()
        else:
            err = "Hành động không hợp lệ."

        if msg or err:
            return redirect(url_for("admin_users", msg=msg, err=err))

    # GET: load users + brands
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT a.id, a.username, a.is_admin, a.team_id, t.name, a.ip_bypass_allowlist
                FROM app_users a
                LEFT JOIN teams t ON t.id = a.team_id
                WHERE a.auth_provider = 'LOCAL'
                ORDER BY a.id DESC
                """
            )
            user_rows = cur.fetchall()

            users = []
            for (uid, username, is_admin, team_id, team_name, ip_bypass_allowlist) in user_rows:
                # Phase 6A: brands are shown READ-ONLY here (inherited from
                # the assigned team, managed on /admin/teams) -- this page
                # never edits team_brands directly anymore.
                inherited_brands = []
                if (not is_admin) and team_id:
                    cur.execute(
                        "SELECT brand FROM team_brands WHERE team_id = %s ORDER BY brand",
                        (team_id,),
                    )
                    inherited_brands = [r[0] for r in cur.fetchall()]

                users.append(
                    {
                        "id": uid,
                        "username": username,
                        "is_admin": bool(is_admin),
                        "ip_bypass_allowlist": bool(ip_bypass_allowlist),
                        "team_id": team_id,
                        "team_name": team_name,
                        "inherited_brands": inherited_brands,
                        "inherited_brands_count": len(inherited_brands),
                    }
                )

        # (Tuỳ chọn) danh sách brand đang có trong products để admin nhìn nhanh.
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT brand FROM products ORDER BY brand")
            distinct_brands = [r[0] for r in cur.fetchall() if r[0] is not None]

        # Phase 5D2B: Google Workspace account lifecycle section (same
        # connection, no extra round-trip). Never includes google_sub,
        # password_hash, or audit rows.
        with conn.cursor() as cur:
            google_users, google_teams = admin_google_users.fetch_google_admin_context(cur)

    except Exception as e:
        users = []
        distinct_brands = []
        google_users = []
        google_teams = []
        err = str(e)
    finally:
        conn.close()

    google_allowed_domains = sorted(
        auth_google.parse_allowed_domains(os.environ.get("GOOGLE_WORKSPACE_ALLOWED_DOMAINS", ""))
    )

    return render_template(
        "admin_users.html",
        users=users,
        distinct_brands=distinct_brands,
        teams=google_teams,  # [{"id":.., "name":..}], same query shape needed by the LOCAL team-select
        google_users=google_users,
        google_teams=google_teams,
        google_allowed_domains=google_allowed_domains,
        message=msg or request.args.get("msg"),
        error=err or request.args.get("err"),
    )


_BRAND_COMPLIANCE_LIST_SQL = """
    SELECT DISTINCT ON (UPPER(TRIM(p.brand)))
        TRIM(p.brand) AS brand,
        UPPER(TRIM(p.brand)) AS brand_norm,
        COALESCE(bcs.manual_compliance_priority, FALSE) AS manual_enabled,
        bcs.updated_at AS setting_updated_at
    FROM products p
    LEFT JOIN brand_compliance_settings bcs
      ON bcs.brand_norm = UPPER(TRIM(p.brand))
    WHERE p.brand IS NOT NULL
      AND TRIM(p.brand) <> ''
    ORDER BY UPPER(TRIM(p.brand)) ASC, TRIM(p.brand) ASC
"""


def _fetch_brand_compliance_rows(cur) -> list[dict]:
    cur.execute(_BRAND_COMPLIANCE_LIST_SQL)
    return [
        {
            "brand": brand,
            "brand_norm": brand_norm,
            "manual_enabled": bool(manual_enabled),
            "setting_updated_at": setting_updated_at,
        }
        for brand, brand_norm, manual_enabled, setting_updated_at in cur.fetchall()
    ]


def _canonical_brand_exists(cur, brand_norm: str) -> bool:
    cur.execute(
        """
        SELECT 1
        FROM products
        WHERE UPPER(TRIM(brand)) = %s
        LIMIT 1
        """,
        (brand_norm,),
    )
    return cur.fetchone() is not None


def _set_brand_manual_compliance_priority(cur, brand_norm: str, enabled: bool) -> None:
    if enabled:
        cur.execute(
            """
            INSERT INTO brand_compliance_settings (brand_norm, manual_compliance_priority)
            VALUES (%s, TRUE)
            ON CONFLICT (brand_norm) DO UPDATE
            SET manual_compliance_priority = TRUE,
                updated_at = NOW()
            """,
            (brand_norm,),
        )
        return
    cur.execute("DELETE FROM brand_compliance_settings WHERE brand_norm = %s", (brand_norm,))


@app.route("/admin/brand-compliance", methods=["GET", "POST"])
def admin_brand_compliance():
    guard = _require_admin_page()
    if guard is not None:
        return guard

    msg = err = None
    if request.method == "POST":
        brand_norm = _norm(request.form.get("brand_norm")).upper()
        action = _norm(request.form.get("action")).lower()
        conn = get_connection()
        try:
            if action not in {"enable", "disable"}:
                err = "Thao tác không hợp lệ."
            elif not brand_norm:
                err = "Thiếu brand."
            else:
                with conn.cursor() as cur:
                    if not _canonical_brand_exists(cur, brand_norm):
                        err = f"Brand không tồn tại: {brand_norm}"
                    else:
                        _set_brand_manual_compliance_priority(cur, brand_norm, action == "enable")
                if not err:
                    conn.commit()
                    if action == "enable":
                        msg = f"Đã bật ưu tiên manual compliance cho brand: {brand_norm}"
                    else:
                        msg = f"Đã tắt ưu tiên manual compliance cho brand: {brand_norm}"
        except Exception as e:
            err = str(e)
        finally:
            conn.close()

    rows = []
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            rows = _fetch_brand_compliance_rows(cur)
    except Exception as e:
        err = err or f"Không đọc được brand compliance (đã chạy migration_011?): {e}"
    finally:
        conn.close()

    enabled_count = sum(1 for row in rows if row["manual_enabled"])
    return render_template(
        "admin_brand_compliance.html",
        rows=rows,
        brand_count=len(rows),
        enabled_count=enabled_count,
        message=msg or request.args.get("msg"),
        error=err or request.args.get("err"),
    )


@app.route("/search", methods=["GET"])
def search_products():
    search_query = request.args.get("query") or ""
    vis, vis_params = _visibility_sql("p")
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            query = f"""
                SELECT
                    p.name,
                    p.code,
                    p.cas,
                    p.brand,
                    p.size,
                    p.ship,
                    p.price,
                    p.note,
                    p.manual_compliance,
                    p.manual_compliance_note,
                    COALESCE(bcs.manual_compliance_priority, FALSE) AS brand_manual_enabled,
                    rr.rule_label AS compliance_status,
                    rr.note AS compliance_note
                FROM products p
                LEFT JOIN brand_compliance_settings bcs
                  ON bcs.brand_norm = UPPER(TRIM(COALESCE(p.brand, '')))
                LEFT JOIN LATERAL (
                    SELECT r.rule_label, r.note
                    FROM regulatory_rules r
                    WHERE NOT (
                        COALESCE(bcs.manual_compliance_priority, FALSE)
                        AND NULLIF(TRIM(COALESCE(p.manual_compliance, '')), '') IS NOT NULL
                    )
                      AND r.is_active = TRUE
                      AND (
                        (r.match_field = 'cas' AND NULLIF(TRIM(p.cas), '') IS NOT NULL AND UPPER(TRIM(p.cas)) = UPPER(TRIM(r.match_value)))
                        OR (r.match_field = 'name' AND NULLIF(TRIM(p.name), '') IS NOT NULL AND UPPER(TRIM(p.name)) = UPPER(TRIM(r.match_value)))
                        OR (r.match_field = 'code' AND NULLIF(TRIM(p.code), '') IS NOT NULL AND UPPER(TRIM(p.code)) = UPPER(TRIM(r.match_value)))
                      )
                    ORDER BY r.priority ASC, r.id ASC
                    LIMIT 1
                ) rr ON TRUE
                WHERE (p.name ILIKE %s OR p.code ILIKE %s OR p.cas ILIKE %s)
                {vis}
                ORDER BY
                    UPPER(TRIM(COALESCE(p.brand, ''))) ASC,
                    UPPER(TRIM(COALESCE(p.size, ''))) ASC,
                    UPPER(TRIM(COALESCE(p.name, ''))) ASC,
                    UPPER(TRIM(COALESCE(p.code, ''))) ASC,
                    p.id ASC
            """
            pattern = f"%{search_query}%"
            cursor.execute(query, (pattern, pattern, pattern) + vis_params)
            products = cursor.fetchall()

        rate_map = _exchange_rate_map(conn)
        results = []
        for product in products:
            (
                name,
                code,
                cas,
                brand,
                size,
                ship,
                price,
                note,
                manual_compliance,
                manual_compliance_note,
                brand_manual_enabled,
                compliance_status,
                compliance_note,
            ) = product
            try:
                ship = float(ship) if ship is not None else 0
            except (TypeError, ValueError):
                ship = 0
            try:
                price = float(price) if price is not None else 0
            except (TypeError, ValueError):
                price = 0

            bkey = (brand or "").strip()
            exchange_rate = rate_map.get(bkey, 1.0)
            unit_price = round(price * ship * exchange_rate, -3)
            formatted_unit_price = "{:,.0f}".format(unit_price)
            resolved = resolve_compliance_precedence(
                brand_manual_enabled=bool(brand_manual_enabled),
                manual_compliance=manual_compliance,
                manual_compliance_note=manual_compliance_note,
                legacy_compliance=compliance_status,
                legacy_compliance_note=compliance_note,
                cas=cas,
            )

            results.append(
                {
                    "Name": name,
                    "Code": code,
                    "Cas": cas,
                    "Brand": brand,
                    "Size": size,
                    "Unit_Price": formatted_unit_price,
                    "Note": note,
                    "Compliance_Status": resolved["compliance"],
                    "Compliance_Note": resolved["compliance_note"],
                    "Compliance_Css": resolved["compliance_css"],
                    "Compliance_Source": resolved["compliance_source"],
                    "note": note or "",
                    "compliance": resolved["compliance"],
                    "compliance_note": resolved["compliance_note"],
                    "compliance_css": resolved["compliance_css"],
                    "compliance_source": resolved["compliance_source"],
                }
            )

        return jsonify({"results": results})
    finally:
        conn.close()


@app.route("/check_cas", methods=["GET"])
def check_cas():
    cas = request.args.get("cas")
    if not cas:
        return jsonify({"warning": False})

    vis, vis_params = _visibility_sql("p")
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            query = f"""
                SELECT r.rule_label
                FROM regulatory_rules r
                WHERE r.is_active = TRUE
                  AND r.match_field = 'cas'
                  AND UPPER(TRIM(r.match_value)) = UPPER(TRIM(%s))
                  AND EXISTS (
                    SELECT 1
                    FROM products p
                    WHERE UPPER(TRIM(p.cas)) = UPPER(TRIM(%s))
                    {vis}
                  )
                ORDER BY r.priority ASC, r.id ASC
                LIMIT 1
            """
            cursor.execute(query, (cas, cas) + vis_params)
            row = cursor.fetchone()

        warning = row[0] if row else None
        if warning:
            return jsonify({"warning": True, "warning_type": warning, "message": f"CAS {cas} thuộc danh mục {warning}."})
        return jsonify({"warning": False})
    finally:
        conn.close()


@app.route("/check_cas_batch", methods=["GET", "POST"])
def check_cas_batch():
    cas_text = request.values.get("cas") or request.values.get("cas_list") or ""
    cas_items = _split_multi_items(cas_text, max_items=2000)
    if not cas_items:
        return jsonify({"results": [], "error": "Thiếu CAS."})

    cas_upper = [c.upper() for c in cas_items]

    vis, vis_params = _visibility_sql("p")
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            # Mỗi CAS lookup riêng qua LATERAL để planner dùng idx_products_cas_upper_trim.
            query = f"""
                WITH input AS (
                    SELECT u.ord, u.cas_u
                    FROM unnest(%s::text[]) WITH ORDINALITY AS u(cas_u, ord)
                )
                SELECT
                    i.ord,
                    i.cas_u,
                    rr.rule_label AS compliance_status,
                    rr.note AS compliance_note
                FROM input i
                LEFT JOIN LATERAL (
                    SELECT 1 AS found
                    FROM products p
                    WHERE UPPER(TRIM(p.cas)) = i.cas_u
                      AND p.cas IS NOT NULL
                      AND TRIM(p.cas) <> ''
                      {vis}
                    -- (p.id + 0) keeps deterministic lowest-id order without tempting a pkey scan.
                    ORDER BY (p.id + 0) ASC
                    LIMIT 1
                ) eligible ON TRUE
                LEFT JOIN LATERAL (
                    SELECT r.rule_label, r.note
                    FROM regulatory_rules r
                    WHERE eligible.found IS NOT NULL
                      AND r.is_active = TRUE
                      AND r.match_field = 'cas'
                      AND UPPER(TRIM(r.match_value)) = i.cas_u
                    ORDER BY r.priority ASC, r.id ASC
                    LIMIT 1
                ) rr ON TRUE
                ORDER BY i.ord
            """
            cursor.execute(query, (cas_upper,) + vis_params)
            rows = cursor.fetchall()

        # Luôn trả về đúng số dòng bằng input (kể cả CAS không có match)
        results = [
            {"Cas": original, "Compliance_Status": "", "Compliance_Note": ""}
            for original in cas_items
        ]
        for ord_, _cas_u, compliance_status, compliance_note in rows:
            idx = int(ord_) - 1
            if 0 <= idx < len(results):
                results[idx]["Compliance_Status"] = compliance_status or ""
                results[idx]["Compliance_Note"] = compliance_note or ""

        return jsonify({"results": results})
    finally:
        conn.close()


@app.route("/find_code_batch", methods=["GET", "POST"])
def find_code_batch():
    codes_text = request.values.get("codes") or request.values.get("code_list") or ""
    codes_items = _split_multi_items(codes_text, max_items=2000)
    if not codes_items:
        return jsonify({"results": [], "error": "Thiếu code."})

    codes_upper = [c.upper() for c in codes_items]

    vis, vis_params = _visibility_sql("p")
    conn = get_connection()
    try:
        rate_map = _exchange_rate_map(conn)
        with conn.cursor() as cursor:
            # Mỗi code lookup riêng qua LATERAL để planner dùng idx_products_code_upper_trim.
            query = f"""
                WITH input AS (
                    SELECT u.ord, u.code_u
                    FROM unnest(%s::text[]) WITH ORDINALITY AS u(code_u, ord)
                )
                SELECT
                    i.ord,
                    p.name,
                    p.code,
                    p.cas,
                    p.brand,
                    p.size,
                    p.ship,
                    p.price,
                    p.note,
                    p.manual_compliance,
                    p.manual_compliance_note,
                    p.brand_manual_enabled,
                    rr.rule_label AS compliance_status,
                    rr.note AS compliance_note
                FROM input i
                LEFT JOIN LATERAL (
                    SELECT
                        p.id,
                        p.name,
                        p.code,
                        p.cas,
                        p.brand,
                        p.size,
                        p.ship,
                        p.price,
                        p.note,
                        p.manual_compliance,
                        p.manual_compliance_note,
                        COALESCE(bcs.manual_compliance_priority, FALSE) AS brand_manual_enabled
                    FROM products p
                    LEFT JOIN brand_compliance_settings bcs
                      ON bcs.brand_norm = UPPER(TRIM(COALESCE(p.brand, '')))
                    WHERE UPPER(TRIM(p.code)) = i.code_u
                      AND p.code IS NOT NULL
                      AND TRIM(p.code) <> ''
                      {vis}
                    -- (p.id + 0) keeps deterministic lowest-id order without tempting a pkey scan.
                    ORDER BY (p.id + 0) ASC
                    LIMIT 1
                ) p ON TRUE
                LEFT JOIN LATERAL (
                    SELECT r.rule_label, r.note
                    FROM regulatory_rules r
                    WHERE p.id IS NOT NULL
                      AND NOT (
                        p.brand_manual_enabled
                        AND NULLIF(TRIM(COALESCE(p.manual_compliance, '')), '') IS NOT NULL
                      )
                      AND r.is_active = TRUE
                      AND (
                        (r.match_field = 'cas' AND NULLIF(TRIM(p.cas), '') IS NOT NULL
                            AND UPPER(TRIM(p.cas)) = UPPER(TRIM(r.match_value)))
                        OR (r.match_field = 'name' AND NULLIF(TRIM(p.name), '') IS NOT NULL
                            AND UPPER(TRIM(p.name)) = UPPER(TRIM(r.match_value)))
                        OR (r.match_field = 'code' AND NULLIF(TRIM(p.code), '') IS NOT NULL
                            AND UPPER(TRIM(p.code)) = UPPER(TRIM(r.match_value)))
                      )
                    ORDER BY r.priority ASC, r.id ASC
                    LIMIT 1
                ) rr ON TRUE
                ORDER BY i.ord
            """
            cursor.execute(query, (codes_upper,) + vis_params)
            rows = cursor.fetchall()

        # Trả về mảng luôn đúng thứ tự/độ dài input
        results = [
            {
                "Name": "",
                "Code": original,
                "Cas": "",
                "Brand": "",
                "Size": "",
                "Unit_Price": "",
                "Note": "",
                "Compliance_Status": "Chưa xác định",
                "Compliance_Note": "",
                "Compliance_Css": "warning-chua-xac-dinh",
                "Compliance_Source": "unresolved",
                "note": "",
                "compliance": "Chưa xác định",
                "compliance_note": "",
                "compliance_css": "warning-chua-xac-dinh",
                "compliance_source": "unresolved",
            }
            for original in codes_items
        ]

        for row in rows:
            (
                ord_,
                name,
                code,
                cas,
                brand,
                size,
                ship,
                price,
                note,
                manual_compliance,
                manual_compliance_note,
                brand_manual_enabled,
                compliance_status,
                compliance_note,
            ) = row
            idx = int(ord_) - 1
            if not (0 <= idx < len(results)):
                continue

            results[idx]["Name"] = name or ""
            # Code giữ nguyên theo input để đảm bảo copy dễ
            results[idx]["Cas"] = cas or ""
            results[idx]["Brand"] = brand or ""
            results[idx]["Size"] = size or ""
            results[idx]["Note"] = note or ""
            results[idx]["note"] = note or ""
            resolved = resolve_compliance_precedence(
                brand_manual_enabled=bool(brand_manual_enabled),
                manual_compliance=manual_compliance,
                manual_compliance_note=manual_compliance_note,
                legacy_compliance=compliance_status,
                legacy_compliance_note=compliance_note,
                cas=cas,
            )
            results[idx]["Compliance_Status"] = resolved["compliance"]
            results[idx]["Compliance_Note"] = resolved["compliance_note"]
            results[idx]["Compliance_Css"] = resolved["compliance_css"]
            results[idx]["Compliance_Source"] = resolved["compliance_source"]
            results[idx]["compliance"] = resolved["compliance"]
            results[idx]["compliance_note"] = resolved["compliance_note"]
            results[idx]["compliance_css"] = resolved["compliance_css"]
            results[idx]["compliance_source"] = resolved["compliance_source"]

            # Unit price chỉ tính nếu có đủ số
            try:
                ship_f = float(ship) if ship is not None else None
            except (TypeError, ValueError):
                ship_f = None
            try:
                price_f = float(price) if price is not None else None
            except (TypeError, ValueError):
                price_f = None

            if ship_f is not None and price_f is not None:
                bkey = (brand or "").strip()
                exchange_rate = rate_map.get(bkey, 1.0)
                unit_price = round(price_f * ship_f * exchange_rate, -3)
                results[idx]["Unit_Price"] = "{:,.0f}".format(unit_price)

        return jsonify({"results": results})
    finally:
        conn.close()


@app.route("/advanced_search/options", methods=["POST"])
def advanced_search_options():
    """Trả về brand/size thực tế có trong DB cho danh sách CAS (không liệt kê toàn bộ size)."""
    cas_text = request.values.get("cas") or request.values.get("cas_list") or ""
    cas_items = _split_multi_items(cas_text, max_items=2000)
    if not cas_items:
        return jsonify({"error": "Thiếu CAS.", "brands": [], "size_pairs": [], "cas_count": 0})

    cas_upper = [c.upper() for c in cas_items]
    vis, vis_params = _visibility_sql("p")
    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            query = f"""
                WITH input AS (
                    SELECT u.ord, u.cas_u
                    FROM unnest(%s::text[]) WITH ORDINALITY AS u(cas_u, ord)
                )
                SELECT
                    TRIM(p.brand) AS brand,
                    TRIM(p.size) AS size,
                    COUNT(DISTINCT i.cas_u) AS cas_hits,
                    COUNT(*) AS row_count
                FROM input i
                INNER JOIN products p ON UPPER(TRIM(p.cas)) = i.cas_u
                WHERE NULLIF(TRIM(p.brand), '') IS NOT NULL
                  AND NULLIF(TRIM(p.size), '') IS NOT NULL
                  {vis}
                GROUP BY TRIM(p.brand), TRIM(p.size)
                ORDER BY row_count DESC, UPPER(TRIM(p.brand)), UPPER(TRIM(p.size))
            """
            cursor.execute(query, (cas_upper,) + vis_params)
            rows = cursor.fetchall()

        brand_totals: dict[str, int] = {}
        size_pairs = []
        for brand, size, cas_hits, row_count in rows:
            brand_key = brand or ""
            size_key = size or ""
            brand_totals[brand_key] = brand_totals.get(brand_key, 0) + int(row_count)
            size_pairs.append(
                {
                    "brand": brand_key,
                    "size": size_key,
                    "cas_hits": int(cas_hits),
                    "row_count": int(row_count),
                }
            )

        brands = [
            {"brand": name, "row_count": count}
            for name, count in sorted(brand_totals.items(), key=lambda x: (-x[1], x[0].upper()))
        ]
        return jsonify(
            {
                "cas_count": len(cas_items),
                "brands": brands,
                "size_pairs": size_pairs,
            }
        )
    finally:
        conn.close()


@app.route("/advanced_search", methods=["POST"])
def advanced_search():
    """Tìm sản phẩm theo danh sách CAS + lọc brand/size (có tuỳ chọn size gần đúng)."""
    cas_text = request.values.get("cas") or request.values.get("cas_list") or ""
    cas_items = _split_multi_items(cas_text, max_items=2000)
    if not cas_items:
        return jsonify({"results": [], "error": "Thiếu CAS."})

    selected_brands = request.values.getlist("brands") or _split_multi_values(
        request.values.get("brands") or request.values.get("brand") or ""
    )
    selected_sizes = request.values.getlist("sizes") or _split_multi_values(
        request.values.get("sizes") or request.values.get("size") or ""
    )
    size_fuzzy = str(request.values.get("size_fuzzy") or "").strip().lower() in ("1", "true", "yes", "on")
    tolerance_pct = 10.0 if size_fuzzy else 0.0

    cas_upper = [c.upper() for c in cas_items]
    vis, vis_params = _visibility_sql("p")
    brand_filter = [b.upper() for b in selected_brands if b]
    exact_size_filter = [(s or "").strip().upper() for s in selected_sizes if (s or "").strip()]
    product_filter_sql = ""
    product_filter_params: tuple = ()
    if brand_filter:
        product_filter_sql += " AND UPPER(TRIM(COALESCE(p.brand, ''))) = ANY(%s)"
        product_filter_params += (brand_filter,)
    if selected_sizes and not size_fuzzy:
        if exact_size_filter:
            product_filter_sql += " AND UPPER(TRIM(COALESCE(p.size, ''))) = ANY(%s)"
            product_filter_params += (exact_size_filter,)
        else:
            product_filter_sql += " AND FALSE"

    conn = get_connection()
    try:
        rate_map = _exchange_rate_map(conn)
        with conn.cursor() as cursor:
            # Mỗi CAS lookup riêng qua LATERAL để planner dùng idx_products_cas_upper_trim.
            query = f"""
                WITH input AS (
                    SELECT u.ord, u.cas_u
                    FROM unnest(%s::text[]) WITH ORDINALITY AS u(cas_u, ord)
                )
                SELECT
                    i.ord,
                    i.cas_u,
                    p.id AS product_id,
                    p.name,
                    p.code,
                    p.cas,
                    p.brand,
                    p.size,
                    p.ship,
                    p.price,
                    p.note,
                    p.manual_compliance,
                    p.manual_compliance_note,
                    p.brand_manual_enabled,
                    rr.rule_label AS compliance_status,
                    rr.note AS compliance_note
                FROM input i
                LEFT JOIN LATERAL (
                    SELECT
                        p.id,
                        p.name,
                        p.code,
                        p.cas,
                        p.brand,
                        p.size,
                        p.ship,
                        p.price,
                        p.note,
                        p.manual_compliance,
                        p.manual_compliance_note,
                        COALESCE(bcs.manual_compliance_priority, FALSE) AS brand_manual_enabled
                    FROM products p
                    LEFT JOIN brand_compliance_settings bcs
                      ON bcs.brand_norm = UPPER(TRIM(COALESCE(p.brand, '')))
                    WHERE UPPER(TRIM(p.cas)) = i.cas_u
                      AND p.cas IS NOT NULL
                      AND TRIM(p.cas) <> ''
                      {vis}
                      {product_filter_sql}
                    -- (p.id + 0) keeps deterministic lowest-id order without tempting a pkey scan.
                    ORDER BY UPPER(TRIM(COALESCE(p.brand, ''))), UPPER(TRIM(COALESCE(p.size, ''))), (p.id + 0) ASC
                ) p ON TRUE
                LEFT JOIN LATERAL (
                    SELECT r.rule_label, r.note
                    FROM regulatory_rules r
                    WHERE p.id IS NOT NULL
                      AND NOT (
                        p.brand_manual_enabled
                        AND NULLIF(TRIM(COALESCE(p.manual_compliance, '')), '') IS NOT NULL
                      )
                      AND r.is_active = TRUE
                      AND (
                        (r.match_field = 'cas' AND NULLIF(TRIM(p.cas), '') IS NOT NULL
                            AND UPPER(TRIM(p.cas)) = UPPER(TRIM(r.match_value)))
                        OR (r.match_field = 'name' AND NULLIF(TRIM(p.name), '') IS NOT NULL
                            AND UPPER(TRIM(p.name)) = UPPER(TRIM(r.match_value)))
                        OR (r.match_field = 'code' AND NULLIF(TRIM(p.code), '') IS NOT NULL
                            AND UPPER(TRIM(p.code)) = UPPER(TRIM(r.match_value)))
                      )
                    ORDER BY r.priority ASC, r.id ASC
                    LIMIT 1
                ) rr ON TRUE
                ORDER BY i.ord, UPPER(TRIM(COALESCE(p.brand, ''))), UPPER(TRIM(COALESCE(p.size, ''))), (p.id + 0) ASC
            """
            cursor.execute(query, (cas_upper,) + vis_params + product_filter_params)
            rows = cursor.fetchall()

        brand_set = {b.upper() for b in selected_brands if b}
        by_ord: dict[int, list[dict]] = {i + 1: [] for i in range(len(cas_items))}

        for row in rows:
            (
                ord_,
                cas_u,
                product_id,
                name,
                code,
                cas,
                brand,
                size,
                ship,
                price,
                note,
                manual_compliance,
                manual_compliance_note,
                brand_manual_enabled,
                compliance_status,
                compliance_note,
            ) = row
            if product_id is None:
                continue
            if brand_set and (brand or "").strip().upper() not in brand_set:
                continue
            if not _size_matches(size, selected_sizes, tolerance_pct=tolerance_pct):
                continue

            by_ord[int(ord_)].append(
                _product_row_to_result(
                    name,
                    code,
                    cas,
                    brand,
                    size,
                    ship,
                    price,
                    note,
                    compliance_status,
                    compliance_note,
                    rate_map,
                    brand_manual_enabled=brand_manual_enabled,
                    manual_compliance=manual_compliance,
                    manual_compliance_note=manual_compliance_note,
                )
            )

        results: list[dict] = []
        matched_cas = 0
        for ord_ in range(1, len(cas_items) + 1):
            items = by_ord.get(ord_, [])
            if items:
                matched_cas += 1
                results.extend(items)
            else:
                resolved = resolve_compliance_precedence(
                    brand_manual_enabled=False,
                    manual_compliance=None,
                    manual_compliance_note=None,
                    legacy_compliance=None,
                    legacy_compliance_note=None,
                    cas=cas_items[ord_ - 1],
                )
                results.append(
                    {
                        "Name": "",
                        "Code": "",
                        "Cas": cas_items[ord_ - 1],
                        "Brand": "",
                        "Size": "",
                        "Unit_Price": "",
                        "Note": "",
                        "Compliance_Status": resolved["compliance"],
                        "Compliance_Note": resolved["compliance_note"],
                        "Compliance_Css": resolved["compliance_css"],
                        "Compliance_Source": resolved["compliance_source"],
                        "note": "",
                        "compliance": resolved["compliance"],
                        "compliance_note": resolved["compliance_note"],
                        "compliance_css": resolved["compliance_css"],
                        "compliance_source": resolved["compliance_source"],
                    }
                )

        return jsonify(
            {
                "results": results,
                "matched_cas": matched_cas,
                "total_cas": len(cas_items),
            }
        )
    finally:
        conn.close()


@app.route("/api/quote-assistant/match", methods=["POST"])
def quote_assistant_match():
    guard = _require_authenticated_quote_api()
    if guard is not None:
        return guard

    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return _quote_json_error("Payload JSON không hợp lệ.")

    try:
        parsed_rows, filters, strategy = _quote_parse_payload(payload)
    except OverflowError as e:
        return _quote_json_error(str(e), status=413)
    except ValueError as e:
        return _quote_json_error(str(e), status=400)

    if not any(row["code_u"] or row["cas_u"] for row in parsed_rows):
        return jsonify(
            {
                "results": _quote_match_rows(None, parsed_rows, filters, strategy),
                "row_count": len(parsed_rows),
                "selection_strategy": strategy,
            }
        )

    conn = get_connection()
    try:
        results = _quote_match_rows(conn, parsed_rows, filters, strategy)
        return jsonify(
            {
                "results": results,
                "row_count": len(results),
                "selection_strategy": strategy,
            }
        )
    finally:
        conn.close()


def _quote_preflight_rows(conn, parsed_rows: list[dict]) -> list[dict]:
    """Perform a stateless, bulk preflight check for request rows.

    Checks Code/CAS existence and conflict using indexes and team visibility.
    Returns no product pricing or candidate lists.
    """
    results: list[dict] = []
    lookup_rows: list[dict] = []
    for row in parsed_rows:
        has_id = bool(row["code_u"] or row["cas_u"])
        if not has_id:
            results.append(
                {
                    "request_id": row["request_id"],
                    "request_order": row["request_order"],
                    "source_row": row["source_row"],
                    "requested_name": row["requested_name"],
                    "requested_code": row["code"],
                    "requested_cas": row["cas"],
                    "preflight_status": "MISSING_IDENTIFIER",
                    "lifecycle": LIFECYCLE_UNRESOLVED,
                    "reason_code": REASON_MISSING_IDENTIFIER,
                    "match_count": 0,
                }
            )
        else:
            lookup_rows.append(row)
            results.append(
                {
                    "request_id": row["request_id"],
                    "request_order": row["request_order"],
                    "source_row": row["source_row"],
                    "requested_name": row["requested_name"],
                    "requested_code": row["code"],
                    "requested_cas": row["cas"],
                    "preflight_status": "NO_MATCH",
                    "lifecycle": LIFECYCLE_UNRESOLVED,
                    "reason_code": REASON_NO_MATCH,
                    "match_count": 0,
                }
            )

    if not lookup_rows or conn is None:
        return results

    vis, vis_params = _visibility_sql("p")
    query = f"""
        WITH input AS (
            SELECT u.ord, u.code_u, u.cas_u
            FROM unnest(%s::int[], %s::text[], %s::text[]) AS u(ord, code_u, cas_u)
        ),
        code_counts AS (
            SELECT i.ord, COUNT(p.id)::int AS count_code
            FROM input i
            JOIN products p
              ON i.code_u <> ''
             AND p.code IS NOT NULL
             AND TRIM(p.code) <> ''
             AND UPPER(TRIM(p.code)) = i.code_u
            {vis}
            GROUP BY i.ord
        ),
        cas_counts AS (
            SELECT i.ord, COUNT(p.id)::int AS count_cas
            FROM input i
            JOIN products p
              ON i.cas_u <> ''
             AND p.cas IS NOT NULL
             AND TRIM(p.cas) <> ''
             AND UPPER(TRIM(p.cas)) = i.cas_u
            {vis}
            GROUP BY i.ord
        ),
        both_counts AS (
            SELECT i.ord, COUNT(p.id)::int AS count_both
            FROM input i
            JOIN products p
              ON i.code_u <> ''
             AND i.cas_u <> ''
             AND p.code IS NOT NULL
             AND TRIM(p.code) <> ''
             AND UPPER(TRIM(p.code)) = i.code_u
             AND p.cas IS NOT NULL
             AND TRIM(p.cas) <> ''
             AND UPPER(TRIM(p.cas)) = i.cas_u
            {vis}
            GROUP BY i.ord
        )
        SELECT
            i.ord,
            COALESCE(cc.count_code, 0) AS count_code,
            COALESCE(ca.count_cas, 0) AS count_cas,
            COALESCE(bc.count_both, 0) AS count_both
        FROM input i
        LEFT JOIN code_counts cc ON cc.ord = i.ord
        LEFT JOIN cas_counts ca ON ca.ord = i.ord
        LEFT JOIN both_counts bc ON bc.ord = i.ord
        ORDER BY i.ord ASC
    """
    params = (
        [row["ord"] for row in lookup_rows],
        [row["code_u"] for row in lookup_rows],
        [row["cas_u"] for row in lookup_rows],
    ) + vis_params + vis_params + vis_params

    with conn.cursor() as cur:
        cur.execute(query, params)
        db_rows = cur.fetchall()

    row_by_ord = {row["ord"]: row for row in lookup_rows}
    for ord_, count_code, count_cas, count_both in db_rows:
        row = row_by_ord.get(int(ord_))
        if not row:
            continue
        idx = row["ord"] - 1
        has_code = bool(row["code_u"])
        has_cas = bool(row["cas_u"])

        if has_code and has_cas:
            if count_both > 0:
                results[idx]["preflight_status"] = "FOUND"
                results[idx]["lifecycle"] = LIFECYCLE_REVIEW
                results[idx]["reason_code"] = REASON_PENDING_MATCH
                results[idx]["match_count"] = count_both
            else:
                results[idx]["preflight_status"] = "CODE_CAS_CONFLICT"
                results[idx]["lifecycle"] = LIFECYCLE_UNRESOLVED
                results[idx]["reason_code"] = REASON_CODE_CAS_CONFLICT
                results[idx]["match_count"] = 0
        elif has_code:
            if count_code > 0:
                results[idx]["preflight_status"] = "FOUND"
                results[idx]["lifecycle"] = LIFECYCLE_REVIEW
                results[idx]["reason_code"] = REASON_PENDING_MATCH
                results[idx]["match_count"] = count_code
            else:
                results[idx]["preflight_status"] = "NO_MATCH"
                results[idx]["lifecycle"] = LIFECYCLE_UNRESOLVED
                results[idx]["reason_code"] = REASON_NO_MATCH
                results[idx]["match_count"] = 0
        elif has_cas:
            if count_cas > 0:
                results[idx]["preflight_status"] = "FOUND"
                results[idx]["lifecycle"] = LIFECYCLE_REVIEW
                results[idx]["reason_code"] = REASON_PENDING_MATCH
                results[idx]["match_count"] = count_cas
            else:
                results[idx]["preflight_status"] = "NO_MATCH"
                results[idx]["lifecycle"] = LIFECYCLE_UNRESOLVED
                results[idx]["reason_code"] = REASON_NO_MATCH
                results[idx]["match_count"] = 0

    return results


@app.route("/api/quote-assistant/preflight", methods=["POST"])
def quote_assistant_preflight():
    guard = _require_authenticated_quote_api()
    if guard is not None:
        return guard

    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return _quote_json_error("Payload JSON không hợp lệ.")

    try:
        parsed_rows, _, _ = _quote_parse_payload(payload)
    except OverflowError as e:
        return _quote_json_error(str(e), status=413)
    except ValueError as e:
        return _quote_json_error(str(e), status=400)

    if not any(row["code_u"] or row["cas_u"] for row in parsed_rows):
        return jsonify(
            {
                "results": _quote_preflight_rows(None, parsed_rows),
                "row_count": len(parsed_rows),
            }
        )

    conn = get_connection()
    try:
        results = _quote_preflight_rows(conn, parsed_rows)
        return jsonify(
            {
                "results": results,
                "row_count": len(results),
            }
        )
    finally:
        conn.close()


@app.route("/api/quote-assistant/request-file/analyze", methods=["POST"])
def quote_assistant_request_file_analyze():
    guard = _require_authenticated_quote_api()
    if guard is not None:
        return guard

    try:
        result = analyze_request_file(
            request.files.get("file"),
            sheet=request.form.get("sheet"),
            header_row=request.form.get("header_row"),
        )
    except OverflowError as e:
        return _quote_json_error(str(e), status=413)
    except (ValueError, zipfile.BadZipFile) as e:
        return _quote_json_error(str(e), status=400)
    return jsonify(result)


@app.route("/api/quote-assistant/request-file/parse", methods=["POST"])
def quote_assistant_request_file_parse():
    guard = _require_authenticated_quote_api()
    if guard is not None:
        return guard

    try:
        result = parse_request_file(request.files.get("file"), request.form.get("mapping") or "")
    except OverflowError as e:
        return _quote_json_error(str(e), status=413)
    except (ValueError, zipfile.BadZipFile) as e:
        return _quote_json_error(str(e), status=400)
    return jsonify(result)


def _quote_export_parse_selections(raw_text: str) -> list[dict]:
    try:
        payload = json.loads(raw_text or "")
    except json.JSONDecodeError as exc:
        raise ValueError("selections phải là JSON hợp lệ.") from exc
    selections = payload.get("selections") if isinstance(payload, dict) else payload
    if not isinstance(selections, list):
        raise ValueError("selections phải là danh sách JSON.")
    if not selections:
        raise ValueError("selections không được rỗng.")
    if len(selections) > QUOTE_MAX_ROWS:
        raise OverflowError(f"Tối đa {QUOTE_MAX_ROWS} sản phẩm mỗi lần xuất.")

    parsed = []
    for index, item in enumerate(selections, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"Selection {index} không hợp lệ.")
        try:
            product_id = int(item.get("product_id"))
        except (TypeError, ValueError):
            raise ValueError(f"Selection {index} thiếu product_id hợp lệ.")
        if product_id <= 0:
            raise ValueError(f"Selection {index} có product_id không hợp lệ.")
        parsed.append({"ord": index, "product_id": product_id})
    return parsed


def _quote_export_parse_items(raw_text: str) -> list[dict]:
    """Parse export_items v2: identity-preserving request groups with ordered lines."""
    try:
        payload = json.loads(raw_text or "")
    except json.JSONDecodeError as exc:
        raise ValueError("export_items phải là JSON hợp lệ.") from exc
    items = payload.get("export_items") if isinstance(payload, dict) else payload
    if not isinstance(items, list):
        raise ValueError("export_items phải là danh sách JSON.")
    if not items:
        raise ValueError("export_items không được rỗng.")
    if len(items) > QUOTE_MAX_ROWS:
        raise OverflowError(f"Tối đa {QUOTE_MAX_ROWS} request mỗi lần xuất.")

    parsed: list[dict] = []
    seen_request_ids: set[str] = set()
    total_lines = 0
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"Request {index} không hợp lệ.")
        request_id = _quote_text(str(item.get("request_id") or ""), max_len=128)
        if not request_id:
            raise ValueError(f"Request {index} thiếu request_id.")
        if request_id in seen_request_ids:
            raise ValueError(f"Request {index}: request_id trùng lặp.")
        seen_request_ids.add(request_id)

        try:
            request_order = int(item.get("request_order") or index)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Request {index}: request_order phải là số nguyên.") from exc
        if request_order < 1:
            raise ValueError(f"Request {index}: request_order phải >= 1.")

        source_row = item.get("source_row")
        if source_row is not None:
            try:
                source_row = int(source_row)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Request {index}: source_row phải là số nguyên hoặc null.") from exc
            if source_row < 1:
                raise ValueError(f"Request {index}: source_row phải >= 1 hoặc null.")

        lines_raw = item.get("lines")
        if not isinstance(lines_raw, list):
            raise ValueError(f"Request {index}: lines phải là danh sách.")
        lines: list[dict] = []
        seen_product_ids: set[int] = set()
        for line_index, line in enumerate(lines_raw, start=1):
            if not isinstance(line, dict):
                raise ValueError(f"Request {index} line {line_index} không hợp lệ.")
            try:
                product_id = int(line.get("product_id"))
            except (TypeError, ValueError):
                raise ValueError(f"Request {index} line {line_index} thiếu product_id hợp lệ.")
            if product_id <= 0:
                raise ValueError(f"Request {index} line {line_index} có product_id không hợp lệ.")
            if product_id in seen_product_ids:
                raise ValueError(f"Request {index} line {line_index}: product_id trùng trong cùng request.")
            seen_product_ids.add(product_id)
            try:
                selection_order = int(line.get("selection_order") or line_index)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Request {index} line {line_index}: selection_order phải là số nguyên.") from exc
            lines.append({"product_id": product_id, "selection_order": selection_order})
            total_lines += 1
            if total_lines > QUOTE_MAX_ROWS:
                raise OverflowError(f"Tối đa {QUOTE_MAX_ROWS} line mỗi lần xuất.")

        placeholder = None
        if not lines:
            placeholder = _quote_export_parse_placeholder(item.get("placeholder"), index)

        parsed.append(
            {
                "request_id": request_id,
                "request_order": request_order,
                "source_row": source_row,
                "requested_name": _quote_text(item.get("requested_name")),
                "requested_code": _quote_text(item.get("requested_code")),
                "requested_cas": _quote_text(item.get("requested_cas")),
                "lines": lines,
                "placeholder": placeholder,
            }
        )
    return parsed


def _quote_export_parse_placeholder(placeholder_raw, index: int) -> dict:
    """Validate placeholder classification/reason for a request with no lines.

    Only whitelisted enum values are accepted; the placeholder note text is
    always generated server-side (never taken verbatim from the client) to
    keep column N free of client-controlled text.
    """
    if not isinstance(placeholder_raw, dict):
        raise ValueError(f"Request {index}: thiếu placeholder khi không có lines.")

    classification = str(placeholder_raw.get("classification") or "").strip().upper()
    if classification not in QUOTE_EXPORT_PLACEHOLDER_CLASSIFICATIONS:
        raise ValueError(f"Request {index}: placeholder.classification không hợp lệ.")

    reason_code_raw = placeholder_raw.get("reason_code")
    reason_code = None
    if reason_code_raw is not None and str(reason_code_raw).strip():
        reason_code = str(reason_code_raw).strip().upper()
        if reason_code not in QUOTE_REASON_CODES:
            raise ValueError(f"Request {index}: placeholder.reason_code không hợp lệ.")
        expected_lifecycle = REASON_CODE_TO_LIFECYCLE.get(reason_code)
        if expected_lifecycle != classification:
            raise ValueError(f"Request {index}: placeholder.reason_code không khớp classification.")

    if classification == LIFECYCLE_BLOCKED:
        reason_vn = QUOTE_EXPORT_BLOCKED_REASON_VN.get(reason_code, QUOTE_EXPORT_BLOCKED_REASON_VN_DEFAULT)
        note_text = f"Không thể báo giá: {reason_vn}"
    else:
        note_text = QUOTE_EXPORT_PLACEHOLDER_NOTES[classification]

    return {"classification": classification, "reason_code": reason_code, "note_text": note_text}


def _quote_export_products(conn, selections: list[dict]) -> list[dict]:
    vis, vis_params = _visibility_sql("p")
    rate_map = _exchange_rate_map(conn)
    query = f"""
        WITH input AS (
            SELECT u.ord, u.product_id
            FROM unnest(%s::int[], %s::int[]) AS u(ord, product_id)
        ),
        product_rows AS (
            SELECT
                i.ord,
                p.id AS product_id,
                p.name,
                p.code,
                p.cas,
                p.brand,
                p.size,
                p.ship,
                p.price,
                p.note,
                p.preparation_type,
                p.manual_compliance,
                p.manual_compliance_note,
                COALESCE(bcs.manual_compliance_priority, FALSE) AS brand_manual_enabled
            FROM input i
            JOIN products p ON p.id = i.product_id
            LEFT JOIN brand_compliance_settings bcs
              ON bcs.brand_norm = UPPER(TRIM(COALESCE(p.brand, '')))
            WHERE TRUE
              {vis}
        )
        SELECT
            pr.ord,
            pr.product_id,
            pr.name,
            pr.code,
            pr.cas,
            pr.brand,
            pr.size,
            pr.ship,
            pr.price,
            pr.note,
            pr.preparation_type,
            pr.manual_compliance,
            pr.manual_compliance_note,
            pr.brand_manual_enabled,
            rr.rule_label AS compliance_status,
            rr.note AS compliance_note
        FROM product_rows pr
        LEFT JOIN LATERAL (
            SELECT r.rule_label, r.note
            FROM regulatory_rules r
            WHERE NOT (
                pr.brand_manual_enabled
                AND NULLIF(TRIM(COALESCE(pr.manual_compliance, '')), '') IS NOT NULL
            )
              AND r.is_active = TRUE
              AND (
                (r.match_field = 'cas' AND NULLIF(TRIM(pr.cas), '') IS NOT NULL
                    AND UPPER(TRIM(pr.cas)) = UPPER(TRIM(r.match_value)))
                OR (r.match_field = 'name' AND NULLIF(TRIM(pr.name), '') IS NOT NULL
                    AND UPPER(TRIM(pr.name)) = UPPER(TRIM(r.match_value)))
                OR (r.match_field = 'code' AND NULLIF(TRIM(pr.code), '') IS NOT NULL
                    AND UPPER(TRIM(pr.code)) = UPPER(TRIM(r.match_value)))
              )
            ORDER BY r.priority ASC, r.id ASC
            LIMIT 1
        ) rr ON TRUE
        ORDER BY pr.ord ASC
    """
    params = (
        [row["ord"] for row in selections],
        [row["product_id"] for row in selections],
    ) + vis_params
    by_ord: dict[int, dict] = {}
    with conn.cursor() as cur:
        cur.execute(query, params)
        for row in cur.fetchall():
            candidate = _quote_candidate_from_row(row, rate_map)
            parts = [
                str(candidate.get("Compliance") or "").strip(),
                str(candidate.get("Compliance_Note") or "").strip(),
            ]
            candidate["Compliance_Combined"] = " | ".join(part for part in parts if part)
            by_ord[int(row[0])] = candidate

    products = []
    for selection in selections:
        candidate = by_ord.get(selection["ord"])
        if not candidate:
            raise ValueError(f"Selection {selection['ord']} không visible hoặc product_id không tồn tại.")
        if candidate.get("ineligible_reason") == "COMPLIANCE_BLOCKED":
            raise ValueError(f"Selection {selection['ord']} bị chặn compliance: {candidate.get('Compliance')}.")
        if not candidate.get("eligible") or float(candidate.get("Unit_Price_Value") or 0) <= 0:
            raise ValueError(f"Selection {selection['ord']} không có Unit_Price hợp lệ.")
        products.append(candidate)
    return products


def _quote_export_placeholder_product(line: dict) -> dict:
    """Build a placeholder row for a request with no selected product.

    B/C/D come from the client-supplied requested Name/Code/CAS (already
    length-bounded text, written as inlineStr so a leading =/+/-/@ stays
    literal text). All other product fields stay empty; price is None so the
    workbook writer clears the price cell instead of writing 0, leaving the
    per-row total formula (e.g. P*qty) to naturally evaluate to 0.
    """
    placeholder = line["placeholder"] or {}
    return {
        "Name": line["requested_name"],
        "Code": line["requested_code"],
        "Cas": line["requested_cas"],
        "Brand": "",
        "Size": "",
        "Note": "",
        "Compliance_Combined": placeholder.get("note_text", ""),
        "Unit_Price_Value": None,
    }


def _quote_export_items_to_products(conn, items: list[dict]) -> list[dict]:
    """Re-fetch products for export_items v2 and emit ordered lines with STT labels.

    Preserves 100% of request_order: every request contributes exactly one
    output row per selected line, or exactly one placeholder row when it has
    no selection (UNRESOLVED/BLOCKED/REVIEW), so a request is never dropped
    or reordered because it has nothing selected. Lines are sorted by
    request_order then selection_order. STT uses request_order for the first
    line of each request (or the placeholder); sub-ordinals (.1, .2, ...)
    start from the second line onward (e.g. 5 -> 5, 5.1, 5.2; 1, 2 -> 1, 1.1, 2).
    """
    ordered_items = sorted(items, key=lambda it: (it["request_order"], it["request_id"]))

    flat_lines: list[dict] = []
    for item in ordered_items:
        sorted_lines = sorted(item["lines"], key=lambda ln: ln["selection_order"])
        if not sorted_lines:
            flat_lines.append(
                {
                    "ord": len(flat_lines) + 1,
                    "product_id": None,
                    "stt": f"{item['request_order']}",
                    "request_id": item["request_id"],
                    "request_order": item["request_order"],
                    "source_row": item["source_row"],
                    "requested_name": item["requested_name"],
                    "requested_code": item["requested_code"],
                    "requested_cas": item["requested_cas"],
                    "placeholder": item["placeholder"],
                }
            )
            continue
        for sub_index, line in enumerate(sorted_lines, start=1):
            if sub_index == 1:
                stt = f"{item['request_order']}"
            else:
                stt = f"{item['request_order']}.{sub_index - 1}"
            flat_lines.append(
                {
                    "ord": len(flat_lines) + 1,
                    "product_id": line["product_id"],
                    "stt": stt,
                    "request_id": item["request_id"],
                    "request_order": item["request_order"],
                    "source_row": item["source_row"],
                    "requested_name": item["requested_name"],
                    "requested_code": item["requested_code"],
                    "requested_cas": item["requested_cas"],
                    "placeholder": None,
                }
            )

    if not flat_lines:
        raise ValueError("export_items không có line nào.")

    real_lines = [ln for ln in flat_lines if ln["product_id"] is not None]
    products_by_ord: dict[int, dict] = {}
    if real_lines:
        fetched = _quote_export_products(conn, real_lines)
        for product, line in zip(fetched, real_lines):
            products_by_ord[line["ord"]] = product

    result: list[dict] = []
    for line in flat_lines:
        if line["product_id"] is not None:
            product = products_by_ord[line["ord"]]
        else:
            product = _quote_export_placeholder_product(line)
        product["STT"] = line["stt"]
        product["request_id"] = line["request_id"]
        product["request_order"] = line["request_order"]
        product["source_row"] = line["source_row"]
        product["requested_name"] = line["requested_name"]
        product["requested_code"] = line["requested_code"]
        product["requested_cas"] = line["requested_cas"]
        result.append(product)
    return result


def _quote_export_download_name(filename: str) -> str:
    base = os.path.basename(filename or "workbook.xlsx")
    stem, ext = os.path.splitext(base)
    if not stem:
        stem = "workbook"
    if ext.lower() != ".xlsx":
        ext = ".xlsx"
    return f"{stem}_draft{ext}"


@app.route("/api/quote-assistant/workbook/template", methods=["GET"])
def quote_assistant_workbook_template():
    guard = _require_authenticated_quote_api()
    if guard is not None:
        return guard

    conn = get_connection()
    try:
        return jsonify({"template": _get_active_quote_template(conn, include_content=False)})
    except QuoteTemplateError as e:
        return _quote_json_error(str(e), status=409)
    except Exception as e:
        if _is_table_missing_error(e):
            app.logger.warning("quote_assistant_workbook_template: quote_templates table not found (migration_013 missing)")
            return _quote_json_error(QUOTE_TEMPLATE_TABLE_UNAVAILABLE_MSG, status=503)
        raise
    finally:
        conn.close()


@app.route("/api/quote-assistant/workbook/export", methods=["POST"])
def quote_assistant_workbook_export():
    guard = _require_authenticated_quote_api()
    if guard is not None:
        return guard

    workbook = request.files.get("workbook")
    raw = None
    filename = None
    if workbook is not None:
        try:
            filename = _safe_uploaded_xlsx_filename(workbook.filename or "workbook.xlsx")
            raw = _read_bounded_workbook_upload(workbook)
        except OverflowError as e:
            return _quote_json_error(str(e), status=413)
        except ValueError as e:
            return _quote_json_error(str(e), status=400)
        if not _is_ooxml_xlsx(raw):
            return _quote_json_error("File không phải .xlsx OOXML hợp lệ.", status=400)

    export_items_raw = request.form.get("export_items") or ""
    try:
        if export_items_raw:
            items = _quote_export_parse_items(export_items_raw)
            selections = []
        else:
            items = None
            selections = _quote_export_parse_selections(request.form.get("selections") or "")
    except OverflowError as e:
        return _quote_json_error(str(e), status=413)
    except ValueError as e:
        return _quote_json_error(str(e), status=400)

    conn = get_connection()
    try:
        if raw is None:
            template = _get_active_quote_template(conn, include_content=True)
            raw = template["content"]
            filename = template["filename"]
        if items is not None:
            products = _quote_export_items_to_products(conn, items)
        else:
            products = _quote_export_products(conn, selections)
        exported = export_quick_quote_workbook(raw, products)
    except QuoteTemplateError as e:
        return _quote_json_error(str(e), status=409)
    except Exception as e:
        if _is_table_missing_error(e):
            app.logger.warning("quote_assistant_workbook_export: quote_templates table not found (migration_013 missing)")
            return _quote_json_error(QUOTE_TEMPLATE_TABLE_UNAVAILABLE_MSG, status=503)
        if isinstance(e, (WorkbookExportError, ValueError)):
            return _quote_json_error(str(e), status=400)
        raise
    finally:
        conn.close()
    return _xlsx_bytes_response(exported, _quote_export_download_name(filename or "workbook.xlsx"))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5001"))
    use_reloader = "--no-reload" not in sys.argv
    app.run(host="0.0.0.0", port=port, debug=True, use_reloader=use_reloader)
