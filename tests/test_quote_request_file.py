import io
import json
import unittest
import zipfile
from unittest.mock import patch

from openpyxl import Workbook

import quote_request_file as qrf
import search
from auth_test_helpers import start_auth_db_patch
from quote_workbook_export import WorkbookExportError


def xlsx_bytes(sheets):
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for title, rows in sheets:
        ws = wb.create_sheet(title)
        for row in rows:
            ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def post_file(client, path, raw, filename, mapping=None, extra=None):
    data = {"file": (io.BytesIO(raw), filename)}
    if mapping is not None:
        data["mapping"] = json.dumps(mapping, ensure_ascii=False)
    if extra:
        data.update(extra)
    return client.post(path, data=data, content_type="multipart/form-data")


class QuoteRequestFileApiTests(unittest.TestCase):
    def setUp(self):
        search.app.testing = True
        self.client = search.app.test_client()
        # Phase 5D2A: stub the per-request session-liveness DB check with an
        # in-memory fake (no real Postgres touched) for every test here.
        start_auth_db_patch(self)

    def auth(self, *, admin=True, team_id=1):
        with self.client.session_transaction() as sess:
            sess["authenticated"] = True
            sess["user_id"] = 1
            sess["auth_version"] = 1
            sess["is_admin"] = admin
            if team_id is not None:
                sess["team_id"] = team_id

    def test_auth_team_guard(self):
        raw = xlsx_bytes([("Sheet1", [["Code"], ["C1"]])])
        response = post_file(self.client, "/api/quote-assistant/request-file/analyze", raw, "sales.xlsx")
        self.assertEqual(response.status_code, 401)

        self.auth(admin=False, team_id=None)
        response = post_file(self.client, "/api/quote-assistant/request-file/analyze", raw, "sales.xlsx")
        self.assertEqual(response.status_code, 403)

        self.auth(admin=False, team_id=7)
        response = post_file(self.client, "/api/quote-assistant/request-file/analyze", raw, "sales.xlsx")
        self.assertEqual(response.status_code, 200)

    def test_analyze_xlsx_multi_sheet_header_offset_vietnamese_aliases_and_preview_limit(self):
        long_cell = "x" * 200
        raw = xlsx_bytes(
            [
                ("Intro", [["not", "headers"], ["still", "intro"]]),
                (
                    "Daily",
                    [
                        ["Report", "", ""],
                        ["Tên hàng", "Mã", "Số CAS", "Extra"],
                        ["Display", 123, "50-00-0", long_cell],
                    ],
                ),
            ]
        )
        self.auth()
        response = post_file(self.client, "/api/quote-assistant/request-file/analyze", raw, "sales.xlsx")
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data["type"], "xlsx")
        self.assertEqual(data["sheets"], ["Intro", "Daily"])
        self.assertEqual(data["suggested_sheet"], "Daily")
        self.assertEqual(data["suggested_header_row"], 2)
        self.assertEqual(data["suggested_mapping"]["requested_name"]["column"], 0)
        self.assertEqual(data["suggested_mapping"]["code"]["column"], 1)
        self.assertEqual(data["suggested_mapping"]["cas"]["column"], 2)
        self.assertEqual(data["columns"][1]["letter"], "B")
        self.assertLessEqual(len(data["preview"]), 20)
        self.assertLessEqual(len(data["preview"][0]), 30)
        self.assertTrue(data["preview"][2][3].endswith("..."))

    def test_ambiguous_mapping_does_not_choose_column(self):
        raw = xlsx_bytes([("Sheet1", [["Code", "Product Code", "CAS"], ["A", "B", "C"]])])
        self.auth()
        response = post_file(self.client, "/api/quote-assistant/request-file/analyze", raw, "sales.xlsx")
        self.assertEqual(response.status_code, 200)
        code = response.get_json()["suggested_mapping"]["code"]
        self.assertTrue(code["ambiguous"])
        self.assertIsNone(code["column"])
        self.assertEqual([c["index"] for c in code["candidates"]], [0, 1])

    def test_analyze_accepts_optional_sheet_and_header_row(self):
        raw = xlsx_bytes(
            [
                ("Auto", [["Code"], ["AUTO"]]),
                ("Manual", [["intro"], ["Product Name", "Catalog Number", "CAS Number"], ["Name", "M1", "64-17-5"]]),
            ]
        )
        self.auth()
        response = post_file(
            self.client,
            "/api/quote-assistant/request-file/analyze",
            raw,
            "sales.xlsx",
            extra={"sheet": "Manual", "header_row": "2"},
        )
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data["suggested_sheet"], "Manual")
        self.assertEqual(data["suggested_header_row"], 2)
        self.assertEqual(data["suggested_mapping"]["code"]["column"], 1)
        self.assertEqual(data["columns"][2]["header"], "CAS Number")

    def test_parse_csv_keeps_blank_name_only_order_duplicates_and_numeric_text(self):
        raw = "\n".join(
            [
                "Tên sản phẩm,Code,CAS",
                "First,00123,50-00-0",
                ",,",
                "Name only,,",
                "First,00123,50-00-0",
            ]
        ).encode("utf-8")
        mapping = {"sheet": qrf.CSV_SHEET_NAME, "header_row": 1, "requested_name": 0, "code": 1, "cas": 2}
        self.auth(admin=False, team_id=3)
        with patch("search.get_connection", side_effect=AssertionError("DB must not be used")):
            response = post_file(self.client, "/api/quote-assistant/request-file/parse", raw, "sales.csv", mapping)
        self.assertEqual(response.status_code, 200)
        rows = response.get_json()["rows"]
        self.assertEqual(
            rows,
            [
                {"source_row": 2, "requested_name": "First", "code": "00123", "cas": "50-00-0"},
                {"source_row": 4, "requested_name": "Name only", "code": "", "cas": ""},
                {"source_row": 5, "requested_name": "First", "code": "00123", "cas": "50-00-0"},
            ],
        )

    def test_parse_xlsx_numeric_values_and_letter_mapping(self):
        raw = xlsx_bytes([("Daily", [["Product Name", "Catalog No", "CAS RN"], ["Numeric", 123.0, "64-17-5"]])])
        mapping = {"sheet": "Daily", "header_row": 1, "requested_name": "A", "code": "B", "cas": "C"}
        self.auth()
        response = post_file(self.client, "/api/quote-assistant/request-file/parse", raw, "sales.xlsx", mapping)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["rows"][0]["code"], "123")
        self.assertEqual(response.get_json()["rows"][0]["cas"], "64-17-5")

    def test_parse_mapping_validation_errors(self):
        raw = xlsx_bytes([("Daily", [["Name", "Code"], ["Only name", ""]])])
        self.auth()

        missing_identifier = post_file(
            self.client,
            "/api/quote-assistant/request-file/parse",
            raw,
            "sales.xlsx",
            {"sheet": "Daily", "header_row": 1, "requested_name": 0, "code": None, "cas": None},
        )
        self.assertEqual(missing_identifier.status_code, 400)
        self.assertIn("ít nhất Code hoặc CAS", missing_identifier.get_json()["error"])

        bad_sheet = post_file(
            self.client,
            "/api/quote-assistant/request-file/parse",
            raw,
            "sales.xlsx",
            {"sheet": "Missing", "header_row": 1, "code": 1},
        )
        self.assertEqual(bad_sheet.status_code, 400)
        self.assertIn("không tồn tại", bad_sheet.get_json()["error"])

        bad_column = post_file(
            self.client,
            "/api/quote-assistant/request-file/parse",
            raw,
            "sales.xlsx",
            {"sheet": "Daily", "header_row": 1, "code": 9},
        )
        self.assertEqual(bad_column.status_code, 400)
        self.assertIn("Cột code", bad_column.get_json()["error"])

    def test_data_row_limit_allows_2000_and_rejects_2001(self):
        ok_raw = xlsx_bytes([("Daily", [["Code"]] + [[f"C{i}"] for i in range(2000)])])
        too_many_raw = xlsx_bytes([("Daily", [["Code"]] + [[f"C{i}"] for i in range(2001)])])
        mapping = {"sheet": "Daily", "header_row": 1, "code": 0}
        self.auth()

        ok = post_file(self.client, "/api/quote-assistant/request-file/parse", ok_raw, "ok.xlsx", mapping)
        self.assertEqual(ok.status_code, 200)
        self.assertEqual(len(ok.get_json()["rows"]), 2000)

        too_many = post_file(self.client, "/api/quote-assistant/request-file/parse", too_many_raw, "too_many.xlsx", mapping)
        self.assertEqual(too_many.status_code, 413)

    def test_rejects_large_legacy_macro_fake_xlsx_and_zip_bomb(self):
        self.auth()
        with patch.object(qrf, "MAX_REQUEST_FILE_BYTES", 5):
            large = post_file(self.client, "/api/quote-assistant/request-file/analyze", b"x" * 6, "large.csv")
        self.assertEqual(large.status_code, 413)

        for filename in ("bad.xls", "bad.xlsm"):
            response = post_file(self.client, "/api/quote-assistant/request-file/analyze", b"x", filename)
            self.assertEqual(response.status_code, 400)

        fake = post_file(self.client, "/api/quote-assistant/request-file/analyze", b"not a zip", "fake.xlsx")
        self.assertEqual(fake.status_code, 400)

        macro_raw = xlsx_bytes([("Daily", [["Code"], ["C1"]])])
        macro_io = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(macro_raw), "r") as src, zipfile.ZipFile(macro_io, "w") as dst:
            for info in src.infolist():
                dst.writestr(info, src.read(info.filename))
            dst.writestr("xl/vbaProject.bin", b"macro")
        macro = post_file(self.client, "/api/quote-assistant/request-file/analyze", macro_io.getvalue(), "macro.xlsx")
        self.assertEqual(macro.status_code, 400)

        with patch.object(qrf, "_read_valid_xlsx_entries", side_effect=WorkbookExportError("zip bomb")):
            bomb = post_file(self.client, "/api/quote-assistant/request-file/analyze", macro_raw, "bomb.xlsx")
        self.assertEqual(bomb.status_code, 400)
        self.assertIn("zip bomb", bomb.get_json()["error"])

    def test_formula_without_cached_value_warns_for_code_and_cas(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily"
        ws.append(["Code", "CAS"])
        ws.append(["=A3", "=B3"])
        ws.append(["CACHED_CODE", "50-00-0"])
        bio = io.BytesIO()
        wb.save(bio)

        self.auth()
        analyze = post_file(self.client, "/api/quote-assistant/request-file/analyze", bio.getvalue(), "formula.xlsx")
        self.assertEqual(analyze.status_code, 200)
        self.assertTrue(any("formula" in warning for warning in analyze.get_json()["warnings"]))

        mapping = {"sheet": "Daily", "header_row": 1, "code": 0, "cas": 1}
        parse = post_file(self.client, "/api/quote-assistant/request-file/parse", bio.getvalue(), "formula.xlsx", mapping)
        self.assertEqual(parse.status_code, 200)
        self.assertTrue(any("formula" in warning for warning in parse.get_json()["warnings"]))

    def test_no_global_file_token_storage_in_module(self):
        with open(qrf.__file__, "r", encoding="utf-8") as fh:
            source = fh.read()
        self.assertNotIn("IMPORT_PREVIEWS", source)
        self.assertNotIn("uuid", source.lower())
        self.assertFalse(hasattr(qrf, "REQUEST_FILE_PREVIEWS"))


if __name__ == "__main__":
    unittest.main()
